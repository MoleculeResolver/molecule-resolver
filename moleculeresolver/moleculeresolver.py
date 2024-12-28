import collections
from concurrent.futures import ThreadPoolExecutor
from contextlib import closing, contextmanager
import copy
from datetime import datetime
from functools import cache
import gzip
import html
import json
import os
from PIL import ImageFont, ImageDraw
import platform
import requests
import subprocess
import tempfile
import time
from types import SimpleNamespace
from typing import Any, Generator, Optional, Union
import traceback
import unicodedata
import urllib
import uuid
import warnings
import ssl

import openpyxl
from prompt_toolkit import PromptSession
from prompt_toolkit.shortcuts import yes_no_dialog
import regex
from rdkit import Chem
from rdkit.Chem import Descriptors
from rdkit.Chem import Draw
from rdkit.Chem import rdmolops
from rdkit.Chem import rdMolDescriptors
from rdkit.Chem.MolStandardize import rdMolStandardize
from rdkit.Chem.rdchem import ResonanceMolSupplierCallback
from tqdm import tqdm
import urllib3
import xmltodict
from moleculeresolver.rdkitmods import disabling_rdkit_logger
from moleculeresolver.molecule import Molecule
from moleculeresolver.SqliteMoleculeCache import SqliteMoleculeCache


class EmptyResonanceMolSupplierCallback(ResonanceMolSupplierCallback):
    """
    A callback class that does nothing when called.

    This class is a workaround for the issue described in
    https://github.com/rdkit/rdkit/issues/6704.

    It inherits from ResonanceMolSupplierCallback but overrides the __call__
    method to do nothing, effectively suppressing any callback behavior.
    """

    def __call__(self) -> None:
        """
        Override the call method to do nothing.

        This method is called when an instance of this class is invoked as a function.
        It intentionally does nothing as a workaround for the RDKit issue.

        Returns:
            None
        """
        pass


class CustomHttpAdapter(requests.adapters.HTTPAdapter):
    """
    A custom HTTP adapter that allows for specifying an SSL context.

    This class is a workaround for the SSL error described in
    https://stackoverflow.com/questions/71603314/ssl-error-unsafe-legacy-renegotiation-disabled/71646353#71646353

    It extends the HTTPAdapter class from the requests library, allowing the use of a custom SSL context
    when making HTTP requests.

    Attributes:
        ssl_context (Optional[ssl.SSLContext]): The SSL context to use for HTTPS connections.
    """

    def __init__(self, ssl_context: Optional[ssl.SSLContext] = None, **kwargs) -> None:
        """
        Initialize the CustomHttpAdapter.

        Args:
            ssl_context (Optional[ssl.SSLContext]): The SSL context to use. Defaults to None.

            **kwargs: Additional keyword arguments to pass to the parent HTTPAdapter.

        Returns:
            None
        """
        self.ssl_context = ssl_context
        super().__init__(**kwargs)

    def init_poolmanager(
        self, connections: int, maxsize: int, block: bool = False
    ) -> None:
        """
        Initialize the connection pool manager with the custom SSL context.

        This method overrides the parent class method to use the custom SSL context
        when initializing the pool manager.

        Args:
            connections (int): The number of connection pools to cache.

            maxsize (int): The maximum number of connections to save in the pool.

            block (bool): Whether the connection pool should block for connections. Defaults to False.

        Returns:
            None
        """
        self.poolmanager = urllib3.poolmanager.PoolManager(
            num_pools=connections,
            maxsize=maxsize,
            block=block,
            ssl_context=self.ssl_context,
        )


class MoleculeResolver:

    def chunker(self, seq: list, size: int) -> set:
        """
        Split a sequence into chunks of a specified size.

        Args:
            seq (list): The sequence to be chunked.

            size (int): The size of each chunk.

        Returns:
            set: A set containing subsequences (chunks) from the input sequence.

        Example:
            >>> list(self.chunker([1, 2, 3, 4, 5, 6], 2))
            [(1, 2), (3, 4), (5, 6)]
        """
        return (seq[pos : pos + size] for pos in range(0, len(seq), size))

    def take_most_common(
        self, container: list, number_to_take: Optional[int] = None
    ) -> list:
        """
        Select the most common elements from a container.

        Identifies and returns the most frequently occurring elements in the given container.
        Handles both case-sensitive and case-insensitive comparisons for string elements.

        Args:
            container (list): The input list of elements to process.

            number_to_take (Optional[int]): The number of most common elements to return.
            If None, returns all elements sorted by frequency. Defaults to None.

        Returns:
            list: A list of the most common elements, preserving the original case for strings.

        Notes:
            - If the container has fewer than 2 elements, it returns the container as is.
            - For string elements, comparisons are case-insensitive, but the original case is preserved in the output.
            - The method maintains the order of elements based on their frequency,
              with ties broken by the order of appearance in the original container.
            - Whitespace is stripped from string elements before comparison.
        """
        if not number_to_take:
            number_to_take = len(container)

        if len(container) < 2:
            return container

        original_container = container

        is_str = isinstance(container[0], str)
        if is_str:
            original_container = [item.strip() for item in container]
            container = [item.strip().lower() for item in container]

        index_container = []
        items_seen = []
        for item_index, item in enumerate(container):
            if item in items_seen:
                continue
            n = container.count(item)
            for i in range(n):
                index_container.append(item_index)
            items_seen.append(item)

        counter = collections.Counter(index_container)
        sorted_item_indices_to_take = [
            item_index for item_index, _ in counter.most_common(number_to_take)
        ]

        final_container = []
        for item_index in sorted_item_indices_to_take:
            final_container.append(original_container[item_index])

        return final_container

    def __init__(
        self,
        available_service_API_keys: Optional[dict[str, Optional[str]]] = None,
        molecule_cache_db_path: Optional[str] = None,
        molecule_cache_expiration: Optional[datetime] = None,
        standardization_options: Optional[dict] = None,
        differentiate_isomers: Optional[bool] = True,
        differentiate_tautomers: Optional[bool] = True,
        differentiate_isotopes: Optional[bool] = True,
        check_for_resonance_structures: Optional[bool] = None,
        show_warning_if_non_unique_structure_was_found: Optional[bool] = False,
    ) -> None:
        """
        Initialize a MoleculeResolver instance.

        Args:
            available_service_API_keys (Optional[dict[str, Optional[str]]]): A dictionary of API keys for various services. Defaults to None.

            molecule_cache_db_path (Optional[str]): Path to the molecule cache database. Defaults to using the same cache for all MoleculeResolver instances
            on a specific environment.

            molecule_cache_expiration (Optional[datetime]): Expiration time for cached molecules. Defaults to None.

            standardization_options (Optional[dict]): Options for molecule standardization. Defaults to None.

            differentiate_isomers (Optional[bool]): Whether to differentiate between isomers. Defaults to True.

            differentiate_tautomers (Optional[bool]): Whether to differentiate between tautomers. Defaults to True.

            differentiate_isotopes (Optional[bool]): Whether to differentiate between isotopes. Defaults to True.

            check_for_resonance_structures (Optional[bool]): Whether to check for resonance structures. Defaults to None.

            show_warning_if_non_unique_structure_was_found (Optional[bool]): Whether to show a warning if a non-unique structure was found. Defaults to False.

        Notes:
            - Sets up the MoleculeResolver with the provided configuration options.
            - Initializes various attributes and sets up the molecule cache if a database path is provided.
        """

        if not available_service_API_keys:
            available_service_API_keys = {}

        if "chemeo" not in available_service_API_keys:
            available_service_API_keys["chemeo"] = None

        if "comptox" not in available_service_API_keys:
            available_service_API_keys["comptox"] = None

        if not molecule_cache_db_path:
            module_path = os.path.dirname(__file__)
            molecule_cache_db_path = os.path.join(module_path, "molecule_cache.db")

        self.molecule_cache_db_path = molecule_cache_db_path
        self.molecule_cache_expiration = molecule_cache_expiration

        self.available_service_API_keys = available_service_API_keys

        default_standardization_options = {
            "disconnect_metals": False,
            "disconnect_more_metals_for_salts": True,
            "normalize": True,
            "reionize": True,
            "uncharge": False,
            "try_assign_sterochemistry": True,
            "remove_atom_mapping_number": True,
        }
        if standardization_options:
            missing_entries = set(default_standardization_options) - set(
                standardization_options
            )
        else:
            standardization_options = {}
            missing_entries = set(default_standardization_options)

        for missing_entry in missing_entries:
            standardization_options[missing_entry] = default_standardization_options[
                missing_entry
            ]

        self._standardization_options = SimpleNamespace(**standardization_options)
        self._differentiate_isomers = differentiate_isomers
        self._differentiate_tautomers = differentiate_tautomers
        self._differentiate_isotopes = differentiate_isotopes
        self._check_for_resonance_structures = check_for_resonance_structures
        self._show_warning_if_non_unique_structure_was_found = (
            show_warning_if_non_unique_structure_was_found
        )

        self._available_services_with_batch_capabilities = ["srs", "comptox", "pubchem"]
        self._message_slugs_shown = []
        self._session = None
        self._session_CompTox = None
        self._java_path = self.get_java_path()
        if self._java_path:
            self._available_services_with_batch_capabilities.insert(0, "opsin")
        self._OPSIN_tempfolder = None
        self.supported_modes_by_services = {
            "cas_registry": ["name", "smiles", "inchi", "cas"],
            "chebi": ["name", "cas", "formula", "smiles", "inchi", "inchikey"],
            "chemeo": ["name", "cas", "smiles", "inchi", "inchikey"],
            "cir": ["formula", "name", "cas", "smiles", "inchi", "inchikey"],
            "comptox": ["name", "cas", "inchikey"],
            "cts": [
                "cas",
                "smiles",
            ],  # "name", I have taken out name because cts works less than 5% of the time
            "nist": ["formula", "name", "cas", "smiles"],
            "opsin": ["name"],
            "pubchem": ["name", "cas", "smiles", "formula", "inchi", "inchikey", "cid"],
            "srs": ["name", "cas"],
        }
        self._available_services = sorted(list(self.supported_modes_by_services.keys()))
        self.supported_modes = []
        self.supported_services_by_mode = {}
        for service, service_modes in self.supported_modes_by_services.items():
            self.supported_modes.extend(service_modes)
            for mode in service_modes:
                if mode not in self.supported_services_by_mode:
                    self.supported_services_by_mode[mode] = []
                self.supported_services_by_mode[mode].append(service)

        self.supported_services_by_mode = {
            k: sorted(self.supported_services_by_mode[k])
            for k in sorted(self.supported_services_by_mode)
        }
        self.supported_modes = sorted(list(set(self.supported_modes)))

        self.CAS_regex_with_groups = regex.compile(r"^(\d{2,7})-(\d{2})-(\d)$")
        self.CAS_regex = r"(\d{2,7}-\d{2}-\d)"
        self.empirical_formula_regex_compiled = regex.compile(
            r"([A-IK-Z][a-ik-z]*)([0-9]+(?:[.][0-9]+)?)?"
        )
        self.formula_bracket_group_regex_compiled = regex.compile(
            r"(\((?:[^()]|(?R))*\))(\d+(?:\.\d+)?)"
        )
        self.non_generic_SMILES_regex_compiled = regex.compile(
            r"^[a-ik-zA-IK-Z0-9%=#$@+\-\[\]\(\)\\\/\:\.]*$"
        )  # non-generic SMILES, no wildcards or unspecified bonds
        self.InChI_regex_compiled = regex.compile(
            r"^InChI=\dS?\/[0-9a-ik-zA-IK-Z]+\/[0-9a-ik-zA-IK-Z+\-\(\)\\\/,\?]*$"
        )
        self.InChIKey_regex_compiled = regex.compile(
            r"^[A-Z]{14}\-[A-Z]{8}[SN][A-Z]\-[A-Z]$"
        )

        self.chemeo_API_token_regex_compiled = regex.compile(r"[a-zA-Z0-9_]+")
        self.comptox_API_token_regex_compiled = regex.compile(r"[a-z0-9\-]+")
        self.html_tag_regex_compiled = regex.compile(r"<.*?>")

        self._init_session()

    def __enter__(self) -> "MoleculeResolver":
        """
        Enter the runtime context for the MoleculeResolver.

        This method is called when entering a 'with' statement. It sets up the
        necessary resources for the MoleculeResolver to function.

        Returns:
            MoleculeResolver: The instance of the class (self).

        Raises:
            Exception: Any exceptions raised during the setup process.

        Notes:
            - Performs the following actions:
                1. Disables the RDKit logger.
                2. Initializes the molecule cache.
                3. Sets up a temporary folder for OPSIN if it's available.
        """
        self._disabling_rdkit_logger = disabling_rdkit_logger()
        self._disabling_rdkit_logger.__enter__()
        self.molecule_cache = SqliteMoleculeCache(
            self.molecule_cache_db_path, self.molecule_cache_expiration
        )
        self.molecule_cache.__enter__()
        if "opsin" in self._available_services_with_batch_capabilities:
            self._OPSIN_tempfolder = tempfile.TemporaryDirectory(
                prefix="OPSIN_tempfolder_"
            )
        return self

    def __exit__(self, exception_type, exception_value, exception_traceback) -> None:
        """
        Exit the runtime context for the MoleculeResolver.

        This method is called when exiting a 'with' statement. It cleans up
        resources used by the MoleculeResolver.

        Args:
            exception_type (Type[BaseException] or None): The type of the exception that caused the context to be exited.

            exception_value (BaseException or None): The instance of the exception that caused the context to be exited.

            exception_traceback (TracebackType or None): A traceback object encoding the stack trace.

        Returns:
            None

        Notes:
            - Performs the following cleanup actions:
                1. Determines if an error occurred during execution.
                2. Exits the RDKit logger disabling context.
                3. Exits the molecule cache context.
                4. Cleans up the OPSIN temporary folder if no error occurred.
        """
        error_ocurred = (
            exception_type is not None
            or exception_value is not None
            or exception_traceback is not None
        )

        self._disabling_rdkit_logger.__exit__(None, None, None)
        self.molecule_cache.__exit__(None, None, None)
        self._disabling_rdkit_logger.__exit__(None, None, None)
        if self._OPSIN_tempfolder and not error_ocurred:
            self._OPSIN_tempfolder.cleanup()

    @contextmanager
    def query_molecule_cache(
        self, service: str, identifier_mode: str, identifier: str
    ) -> Generator[tuple[bool, list[Molecule]], None, None]:
        """
        Query the molecule cache for a given identifier and yield the results.

        Searches the molecule cache for a specific identifier using the provided service and identifier mode.
        Yields whether an entry is available and the list of molecules found. After the context is exited,
        it handles saving new molecules to the cache if necessary.

        Args:
            service (str): The service used for querying (e.g., "cts", "cir").

            identifier_mode (str): The mode of identification used.

            identifier (str): The specific identifier to search for.

        Returns:
            tuple[bool, list[Molecule]]: A tuple containing:

            - bool: True if an entry is available in the cache, False otherwise.
            - list[Molecule]: The list of molecules found in the cache (empty if not found).

        Raises:
            Exception: Any exceptions raised by the underlying cache operations.

        Notes:
            - Uses a context manager to ensure proper handling of cache operations.
            - Handles special cases for CTS and CIR services when they are down.
            - New molecules are saved to the cache after the context is exited, unless
              specific conditions prevent saving (e.g., service is down).
        """
        molecules = self.molecule_cache.search(service, identifier_mode, identifier)
        entry_available = molecules is not None

        if entry_available:
            yield entry_available, molecules
        else:
            molecules = []
            yield entry_available, molecules
            for molecule in molecules:
                molecule.identifier = identifier

            should_save = True
            if service == "cts" and "CTS_is_down" in self._message_slugs_shown:
                should_save = False
            elif service == "cir" and "CIR_is_down" in self._message_slugs_shown:
                should_save = False

            if should_save:
                self.molecule_cache.save(
                    service,
                    identifier_mode,
                    identifier,
                    molecules if molecules else None,
                )

        return

    @contextmanager
    def query_molecule_cache_batchmode(
        self,
        service: str,
        identifier_mode: str,
        identifiers: list[str],
        save_not_found: Optional[bool] = True,
    ) -> Generator[
        tuple[list[str], list[int], Optional[list[Optional[list[Molecule]]]]],
        None,
        None,
    ]:
        """
        Query the molecule cache for multiple identifiers in batch mode.

        Searches the cache for molecules matching the given service, identifier mode, and list of identifiers.
        Yields information about identifiers to search, their indices, and the results. After the context is exited,
        it saves new molecules to the cache.

        Args:
            service (str): The service used for querying (e.g., "cts", "cir").

            identifier_mode (str): The mode of identification used.

            identifiers (list[str]): The list of identifiers to search for.

            save_not_found (Optional[bool]): Whether to save entries for identifiers not found. Defaults to True.

        Returns:
            tuple[list[str], list[int], Optional[list[Optional[list[Molecule]]]]]: A tuple containing:
                - list[str]: Identifiers that need to be searched (not found in cache).
                - list[int]: Indices of the identifiers to be searched.
                - Optional[list[Optional[list[Molecule]]]]: Results from the cache search.
        """
        results = self.molecule_cache.search(
            [service] * len(identifiers),
            [identifier_mode] * len(identifiers),
            identifiers,
        )

        identifiers_to_search = []
        indices_of_identifiers_to_search = []
        for (molecule_index, molecule), identifier in zip(
            enumerate(results), identifiers, strict=True
        ):
            if molecule is None:
                identifiers_to_search.append(identifier)
                indices_of_identifiers_to_search.append(molecule_index)

        yield identifiers_to_search, indices_of_identifiers_to_search, results
        identifiers_to_save = []
        molecules_to_save = []
        for molecule_index, identifier in zip(
            indices_of_identifiers_to_search, identifiers_to_search, strict=True
        ):
            molecules = results[molecule_index]
            if molecules:
                for molecule in molecules:
                    identifiers_to_save.append(identifier)
                    molecules_to_save.append(molecule)
            else:
                if save_not_found:
                    identifiers_to_save.append(identifier)
                    molecules_to_save.append(None)

        self.molecule_cache.save(
            [service] * len(identifiers_to_save),
            [identifier_mode] * len(identifiers_to_save),
            identifiers_to_save,
            molecules_to_save,
        )
        return

    def _init_session(
        self, pool_connections: Optional[int] = None, pool_maxsize: Optional[int] = None
    ) -> None:
        """
        Initialize HTTP sessions for making requests.

        Sets up two sessions: a general session and a specific session for CompTox.
        Configures connection pooling and SSL contexts for these sessions.

        Args:
            pool_connections (Optional[int]): The number of connection pools to cache.
            If None, it's set to twice the number of available services.

            pool_maxsize (Optional[int]): The maximum number of connections to save in the pool.
            If None, it's set to 10. The minimum value is always 10.

        Notes:
            - This method is idempotent; it will not reinitialize existing sessions.
            - The CompTox session uses a custom SSL context to handle specific SSL requirements.
        """
        if self._session is not None:
            return

        if pool_connections is None:
            # reserve two connections for each service in case of different hosts are used.
            pool_connections = len(self._available_services) * 2

        if pool_maxsize is None:
            # default from documentation
            pool_maxsize = 10
        else:
            pool_maxsize = max(pool_maxsize, 10)

        self._session = requests.Session()
        self._session.mount(
            "https://",
            requests.adapters.HTTPAdapter(
                pool_connections=pool_connections,
                pool_maxsize=pool_maxsize,
                max_retries=2,
            ),
        )

        self._session_CompTox = requests.Session()
        ctx = ssl.create_default_context(ssl.Purpose.SERVER_AUTH)
        ctx.options |= 0x4
        self._session_CompTox.mount(
            "https://",
            CustomHttpAdapter(
                ctx,
                pool_connections=pool_connections,
                pool_maxsize=pool_maxsize,
                max_retries=2,
            ),
        )

    def _resilient_request(
        self,
        url: str,
        kwargs: Optional[dict[str, Any]] = None,
        request_type: Optional[str] = "get",
        accepted_status_codes: list[int] = [200],
        rejected_status_codes: list[int] = [404],
        offline_status_codes: list[int] = [],
        max_retries: Optional[int] = 10,
        sleep_time: Union[int, float] = 2,
        allow_redirects: Optional[bool] = False,
        json: Optional[str] = None,
        return_response: Optional[bool] = False,
    ) -> Optional[str]:
        """
        Make a resilient HTTP request with retry logic.

        Attempts to make an HTTP request, handling various error conditions and retrying the request if necessary.

        Args:
            url (str): The URL to send the request to.

            kwargs (Optional[dict[str, Any]]): Additional keyword arguments for the request.

            request_type (Optional[str]): The type of HTTP request ('get' or 'post'). Defaults to 'get'.

            accepted_status_codes (list[int]): List of HTTP status codes to accept. Defaults to [200].

            rejected_status_codes (list[int]): List of HTTP status codes to reject. Defaults to [404].

            max_retries (Optional[int]): Maximum number of retry attempts. Defaults to 10.

            sleep_time (Union[int, float]): Time to sleep between retries in seconds. Defaults to 2.

            allow_redirects (Optional[bool]): Whether to allow URL redirection. Defaults to False.

            json (Optional[str]): JSON data to send in the request body. Defaults to None.

            return_response (Optional[bool]): If True, return the full response object instead of the text. Defaults to False.

        Returns:
            Optional[str]: The response text if successful, or None if the request failed.

        Raises:
            ValueError: If an invalid request_type is provided.
            requests.exceptions.ConnectionError: If connection errors persist after maximum retries.

        Notes:
            - Automatically sets a user agent if not provided in the headers.
            - Uses different sessions for CompTox and other services.
            - Implements exponential backoff for retries.
        """
        if not kwargs:
            kwargs = {}

        if request_type not in ["get", "post"]:
            raise ValueError("The request_type must be either 'get' or 'post'.")

        if "timeout" not in kwargs:
            kwargs["timeout"] = 5

        headers = {}
        user_agent_is_set = False
        if "headers" in kwargs:
            headers = kwargs["headers"]
            user_agent_is_set = "user-agent" in [key.lower() for key in headers.keys()]

        if user_agent_is_set is False:
            headers["user-agent"] = (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0"
            )

        kwargs["headers"] = headers

        session = (
            self._session if "comptox" not in url.lower() else self._session_CompTox
        )
        n_try = 0
        while True:
            n_try += 1
            try:
                response = None

                if request_type == "get":
                    response = session.get(
                        url, **kwargs, json=json, allow_redirects=allow_redirects
                    )
                if request_type == "post":
                    response = session.post(
                        url, **kwargs, json=json, allow_redirects=allow_redirects
                    )

                if response.status_code in accepted_status_codes:
                    if return_response:
                        return response
                    response_text = response.content.decode("utf8", errors="ignore")
                    response_text = response_text.replace("\u200b", "")
                    # find encoding errors
                    if response_text != response.text:
                        if response_text.count(r"\u") + response_text.count(r"\x") > 0:
                            raise UnicodeEncodeError(
                                "Wrong character encoding was used."
                            )
                    return response_text
                elif response.status_code in rejected_status_codes:
                    return None
                else:
                    raise requests.exceptions.HTTPError("Wrong status_code.")

            except requests.exceptions.RequestException as error:
                time.sleep(sleep_time)

                if n_try > max_retries:
                    if isinstance(error, requests.exceptions.ConnectionError):
                        raise error

                    if response.status_code in offline_status_codes:
                        raise requests.exceptions.ConnectionError(
                            "The service is probably offline."
                        )

                    if response is not None:
                        print(
                            f"ERROR ocurred during request to:\n{url}\n{kwargs}\n HTTP request status code: {response.status_code}\n"
                        )
                    else:
                        print(
                            f"ERROR ocurred during request to:\n{url}\n{kwargs}\n",
                            traceback.format_exc(),
                        )

                    return None

    @cache
    def try_disconnect_more_metals(self, SMILES: str) -> str:
        """
        Attempt to disconnect additional metal atoms in a molecule represented by a SMILES string.

        This method performs a more extensive metal disconnection process than the standard
        RDKit metal disconnector. It includes more metals and handles specific cases like
        mercury (Hg) correctly: https://github.com/rdkit/rdkit/discussions/6729
        if the issue for Hg is a bug, this should be replaced by the metal disconnector from rdkit

        Args:
            SMILES (str): The input SMILES string representing the molecule.

        Returns:
            str: The SMILES string of the molecule after attempting to disconnect metals.
            If no changes are made, the original SMILES string is returned.

        Notes:
            - Uses caching to improve performance for repeated calls with the same input.
            - Employs a custom SMARTS pattern to identify metal-nonmetal bonds.
            - If the input SMILES is invalid or no metal disconnection is possible, the original SMILES is returned.
            - Particularly useful for handling cases where the standard RDKit metal disconnector may not be sufficient
              or may have known issues (e.g., with mercury compounds).
        """
        metals = "[#3,#11,#19,#37,#55,#87,#4,#12,#20,#38,#56,#88,#21,#22,#23,#24,#25,#26,#27,#28,#29,#30,#13,#31,#39,#40,#41,#42,#43,#44,#45,#46,#47,#48,#49,#50,#72,#73,#74,#75,#76,#77,#78,#79,#80,#81,#82,#83]"
        SMARTS = f"{metals}~[B,C,#14,P,#33,#51,S,#34,#52,F,Cl,Br,I,#85]"
        mol = Chem.MolFromSmiles(SMILES)
        if not mol:
            return SMILES
        patt = Chem.MolFromSmarts(SMARTS)
        hit_ats = list(mol.GetSubstructMatches(patt))
        if not hit_ats:
            return SMILES

        property_name = "molecule_resolver_charge"
        cation_charges = {}
        bonds_to_be_broken = []
        for cation_id, anion_id in hit_ats:
            cation = mol.GetAtomWithIdx(cation_id)
            if cation_id not in cation_charges:
                cation_charge = cation.GetTotalDegree()
                cation_charges[cation_id] = cation_charge
                cation.SetIntProp(property_name, cation_charge)
            else:
                cation_charge = cation_charges[cation_id]

            bond_to_be_broken = mol.GetBondBetweenAtoms(cation_id, anion_id)
            bonds_to_be_broken.append(bond_to_be_broken.GetIdx())

            anion_charge = -1 * int(bond_to_be_broken.GetBondType())
            anion = mol.GetAtomWithIdx(anion_id)
            anion.SetIntProp(property_name, anion_charge)
            cation_charges[cation_id] += anion_charge

        if not all(v == 0 for v in cation_charges.values()):
            return SMILES  # charge missmatch

        fragment_mols = rdmolops.GetMolFrags(
            rdmolops.FragmentOnBonds(mol, bonds_to_be_broken), asMols=True
        )

        ion_mols = []
        for m in fragment_mols:
            rwmol = Chem.RWMol(m)
            ai_dummy_atoms = []
            for a in rwmol.GetAtoms():
                if a.HasProp(property_name):
                    a.SetFormalCharge(a.GetIntProp(property_name))
                if a.GetSymbol() == "*":
                    ai_dummy_atoms.append(a.GetIdx())

            for ai in sorted(ai_dummy_atoms, reverse=True):
                rwmol.RemoveAtom(ai)

            ion_mols.append(rwmol.GetMol())

        return self.standardize_SMILES(
            ".".join([Chem.MolToSmiles(m) for m in ion_mols])
        )

    @cache
    def standardize_SMILES(
        self,
        SMILES: str,
        /,
        disconnect_metals: Optional[bool] = None,
        normalize: Optional[bool] = None,
        reionize: Optional[bool] = None,
        uncharge: Optional[bool] = None,
        try_assign_sterochemistry: Optional[bool] = None,
        remove_atom_mapping_number: Optional[bool] = None,
    ) -> Optional[str]:
        """
        Standardize a SMILES string representation of a molecule.

        Applies various standardization procedures to a given SMILES string, including metal disconnection,
        normalization, reionization, uncharging, stereochemistry assignment, and atom mapping number removal.

        Args:
            SMILES (str): The input SMILES string to be standardized.

            disconnect_metals (Optional[bool]): Whether to disconnect metals. Defaults to None.

            normalize (Optional[bool]): Whether to normalize the molecule. Defaults to None.

            reionize (Optional[bool]): Whether to reionize the molecule. Defaults to None.

            uncharge (Optional[bool]): Whether to uncharge the molecule. Defaults to None.

            try_assign_sterochemistry (Optional[bool]): Whether to attempt stereochemistry assignment. Defaults to None.

            remove_atom_mapping_number (Optional[bool]): Whether to remove atom mapping numbers. Defaults to None.

        Returns:
            Optional[str]: The standardized SMILES string, or None if standardization fails.

        Notes:
            - Uses caching to improve performance for repeated calls with the same input.
            - The standardization process uses the RDKit library for molecular operations.
        """
        mol = self.get_from_SMILES(SMILES)
        if mol is None:
            return None

        return Chem.MolToSmiles(
            self.standardize_mol(
                mol,
                disconnect_metals,
                normalize,
                reionize,
                uncharge,
                try_assign_sterochemistry,
                remove_atom_mapping_number,
            )
        )

    def standardize_mol(
        self,
        mol: Chem.rdchem.Mol,
        /,
        disconnect_metals: Optional[bool] = None,
        normalize: Optional[bool] = None,
        reionize: Optional[bool] = None,
        uncharge: Optional[bool] = None,
        try_assign_sterochemistry: Optional[bool] = None,
        remove_atom_mapping_number: Optional[bool] = None,
    ) -> Optional[Chem.rdchem.Mol]:
        """
        Standardize an RDKit molecule object.

        Applies various standardization procedures to a given RDKit molecule, including metal disconnection,
        normalization, reionization, uncharging, stereochemistry assignment, and atom mapping number removal.

        Args:
            mol (Chem.rdchem.Mol): The input RDKit molecule to be standardized.

            disconnect_metals (Optional[bool]): Whether to disconnect metals. Defaults to None.

            normalize (Optional[bool]): Whether to normalize the molecule. Defaults to None.

            reionize (Optional[bool]): Whether to reionize the molecule. Defaults to None.

            uncharge (Optional[bool]): Whether to uncharge the molecule. Defaults to None.

            try_assign_sterochemistry (Optional[bool]): Whether to attempt stereochemistry assignment. Defaults to None.

            remove_atom_mapping_number (Optional[bool]): Whether to remove atom mapping numbers. Defaults to None.

        Returns:
            Optional[Chem.rdchem.Mol]: The standardized RDKit molecule, or None if standardization fails.

        Raises:
            Warning: If reionization step cannot be performed.

        Notes:
            - If any standardization option is None, it defaults to the values set on class creation or default values.
            - For molecules with multiple fragments, each fragment is standardized separately.
            - Special handling is implemented for certain molecules (e.g., DMSO) during normalization.
            - Uses RDKit's standardization tools (e.g., MetalDisconnector, Normalize, Reionizer).
        """
        if mol is None:
            return None

        if disconnect_metals is None:
            disconnect_metals = self._standardization_options.disconnect_metals

        if normalize is None:
            normalize = self._standardization_options.normalize

        if reionize is None:
            reionize = self._standardization_options.reionize

        if uncharge is None:
            uncharge = self._standardization_options.uncharge

        if try_assign_sterochemistry is None:
            try_assign_sterochemistry = (
                self._standardization_options.try_assign_sterochemistry
            )

        if remove_atom_mapping_number is None:
            remove_atom_mapping_number = (
                self._standardization_options.remove_atom_mapping_number
            )

        mol_fragments = Chem.GetMolFrags(mol, asMols=True, sanitizeFrags=False)

        if len(mol_fragments) > 1:
            all_mol_fragments = []
            for mol_fragment in mol_fragments:
                standardized_mol_fragment = self.standardize_mol(
                    mol_fragment,
                    disconnect_metals,
                    normalize,
                    reionize,
                    uncharge,
                    try_assign_sterochemistry,
                    remove_atom_mapping_number,
                )
                if standardized_mol_fragment is None:
                    return None
                all_mol_fragments.append(standardized_mol_fragment)

            standardize_mol = all_mol_fragments[0]
            for mol_fragment in all_mol_fragments[1:]:
                standardize_mol = Chem.CombineMols(standardize_mol, mol_fragment)

            return standardize_mol

        # based on datamol version on https://github.com/datamol-org/datamol
        mol = copy.deepcopy(mol)

        if disconnect_metals:
            md = rdMolStandardize.MetalDisconnector()
            mol = md.Disconnect(mol)

        if normalize:
            # for some molecules this gives weird results. e.g. DMSO
            if Chem.MolToSmiles(mol) not in ["CS(C)=O"]:
                mol = rdMolStandardize.Normalize(mol)

        if reionize:
            reionizer = rdMolStandardize.Reionizer()
            try:
                mol = reionizer.reionize(mol)
            except:
                warnings.warn(
                    "Reionization step could not be performed. It will be skipped."
                )

        if uncharge:
            uncharger = rdMolStandardize.Uncharger()
            mol = uncharger.uncharge(mol)

        if try_assign_sterochemistry:
            Chem.AssignStereochemistry(mol, force=False, cleanIt=True)

        if remove_atom_mapping_number:
            [a.SetAtomMapNum(0) for a in mol.GetAtoms()]

        return mol

    def has_isotopes(self, mol: Chem.rdchem.Mol):
        """
        Check if the molecule contains any isotopes.

        Examines each atom in the given molecule to determine if any of them have a non-zero isotope value.

        Args:
            mol (Chem.rdchem.Mol): The input RDKit molecule to check for isotopes.

        Returns:
            bool: True if the molecule contains at least one isotope, False otherwise.

        Notes:
            - Uses RDKit's GetIsotope() function to check each atom's isotope value.
            - An isotope value of 0 indicates the most common isotope for that element.
        """
        return any([atom.GetIsotope() != 0 for atom in mol.GetAtoms()])

    def remove_isotopes(self, mol: Chem.rdchem.Mol) -> Chem.rdchem.Mol:
        """
        Remove all isotope information from the molecule.

        Sets the isotope value of all atoms in the molecule to 0, effectively removing any isotope information.

        Args:
            mol (Chem.rdchem.Mol): The input RDKit molecule from which to remove isotopes.

        Returns:
            Chem.rdchem.Mol: A new RDKit molecule with all isotope information removed and explicit hydrogens removed.

        Notes:
            - Setting an atom's isotope to 0 indicates the most common isotope for that element.
            - Modifies the input molecule in-place before returning a new molecule with explicit hydrogens removed.
            - Removing explicit hydrogens can change the molecule's representation but not its chemical identity.
        """
        for atom in mol.GetAtoms():
            atom.SetIsotope(0)
        return mol

    def _check_and_flatten_identifiers_and_modes(
        self,
        identifiers: Union[str, list[str], list[list[str]]],
        modes: Union[str, list[str]],
    ) -> tuple[list[str], list[str], list[str], list[str], str]:
        """
        Validate and flatten the input identifiers and modes.

        Processes the input identifiers and modes to ensure they are in the correct format
        and flattens nested structures. It also performs validation checks on the inputs.

        Args:
            identifiers (Union[str, list[str], list[list[str]]]): The input identifiers,
            which can be a single string, a list of strings, or a list of lists of strings.

            modes (Union[str, list[str]]): The input modes, which can be a single string or a list of strings.

        Returns:
            tuple[list[str], list[str], list[str], list[str], str]: A tuple containing:
                - list[str]: Flattened list of identifiers.
                - list[str]: Flattened list of modes.
                - list[str]: List of unique identifiers.
                - list[str]: List of unique modes.
                - str: A string representation of the unique modes.

        Raises:
            TypeError: If an identifier is not of type int, str, list, or tuple.
            ValueError: If the number of modes doesn't match the number of identifier groups.

        Notes:
            - Handles various input formats and normalizes them for further processing.
            - Ensures that each identifier has a corresponding mode.
            - Strips whitespace from identifiers and converts them to strings.
            - Converts modes to lowercase for consistency.
        """
        if isinstance(identifiers, str):
            identifiers = [identifiers]
        if isinstance(modes, str):
            modes = [modes]

        flattened_identifiers = []
        flattened_modes = []
        for identifier, mode in zip(identifiers, modes, strict=True):
            if all([not isinstance(identifier, t) for t in [int, str, list, tuple]]):
                raise TypeError("An identifier can only be an int, str, list or tuple.")

            temp_identifiers = identifier
            if not isinstance(identifier, tuple) and not isinstance(identifier, list):
                temp_identifiers = [identifier]

            for i in temp_identifiers:
                flattened_identifiers.append(str(i).strip())
                flattened_modes.append(mode.strip().lower())

        synonyms = []
        CAS = set()
        given_SMILES = []
        for identifier, mode in zip(
            flattened_identifiers, flattened_modes, strict=True
        ):
            if mode == "name":
                synonyms.append(identifier)
            elif mode == "cas":
                if not self.is_valid_CAS(identifier):
                    raise ValueError("You provided an invalid CAS.")
                CAS.add(identifier)
            elif mode == "inchi":
                if not self.is_valid_InChI(identifier):
                    raise ValueError("You provided an invalid InChI.")
                given_SMILES.append(self.InChI_to_SMILES(identifier))
            elif mode == "smiles":
                if not self.is_valid_SMILES(identifier):
                    raise ValueError("You provided an invalid SMILES.")
                given_SMILES.append(identifier)

        if not len(given_SMILES) < 2:
            raise ValueError(
                "Only one valid structure is accepted to search for (this includes SMILES and InChI)."
            )
        if len(given_SMILES) > 0:
            given_SMILES = given_SMILES[0]
        else:
            given_SMILES = None

        self._check_parameters(
            modes=flattened_modes,
            identifiers=flattened_identifiers,
            context="find_single",
        )

        return flattened_identifiers, flattened_modes, synonyms, CAS, given_SMILES

    def _is_list_of_list_of_str(self, value: list[list[str]]) -> bool:
        """
        Check if the input is a valid list of lists of strings.

        Verifies that the input value is a list containing only lists,
        and that each nested list contains only string elements.

        Args:
            value (list[list[str]]): The input to be checked.

        Returns:
            bool: True if the input is a valid list of lists of strings, False otherwise.

        Notes:
            - Performs a two-level deep check on the input structure.
            - First ensures that all elements of the outer list are themselves lists.
            - Then checks that all elements within each nested list are strings.
            - An empty list or a list containing empty lists will return True.
        """
        for items in value:
            if not isinstance(items, list):
                return False
            if not all([isinstance(item, str) for item in items]):
                return False
        return True

    def _check_parameters(
        self,
        *,
        modes=None,
        services=None,
        identifiers=None,
        required_formulas=None,
        required_charges=None,
        required_structure_types=None,
        context="get_molecule",
    ):
        """
        Validate input parameters for molecule retrieval and processing.

        Performs extensive validation on various input parameters used in
        molecule retrieval and processing operations. It checks for type correctness,
        value validity, and consistency across different parameters based on the context
        of the operation.

        Args:
            modes (Optional[Union[str, list[str]]]): The mode(s) of molecule identification.

            services (Optional[Union[str, list[str]]]): The service(s) to be used for retrieval.

            identifiers (Optional[Union[str, list[str], list[list[str]]]]): The molecule identifier(s).

            required_formulas (Optional[Union[str, list[str]]]): Required chemical formula(s).

            required_charges (Optional[Union[str, int, list[Union[str, int]]]]): Required charge(s).

            required_structure_types (Optional[Union[str, list[str]]]): Required structure type(s).

            context (str): The context of the operation. Defaults to "get_molecule".

        Raises:
            TypeError: If any parameter is of an incorrect type.

            ValueError: If any parameter has an invalid value or if there are inconsistencies
            between parameters.

        Notes:
            - The method's behavior varies based on the 'context' parameter.
            - For 'get_molecule' and 'get_molecules_batch' contexts, it expects single values
              for modes and services.
            - For 'batch' and 'find_single' contexts, it expects lists for modes and services.
            - Checks for compatibility between modes and services.
            - Performs specific validations for charges, structure types, and formulas.
            - Ensures consistency in the lengths of certain parameters when applicable.
        """
        if modes is not None:
            if context == "get_molecule" or context == "get_molecules_batch":
                if not isinstance(modes, str):
                    raise TypeError("The mode parameter can only be a string.")
                if modes not in self.supported_modes:
                    raise ValueError("The mode parameter can only be a supported mode.")
                if (
                    services is not None
                    and modes not in self.supported_modes_by_services[services]
                ):
                    raise ValueError(
                        f"The chosen mode ({modes}) is not compatible with the service {services}."
                    )

            if context == "batch" or context == "find_single":
                if not isinstance(modes, list):
                    raise TypeError(
                        "The modes parameter can only be a list of strings."
                    )
                if not all([isinstance(temp_mode, str) for temp_mode in modes]):
                    if not self._is_list_of_list_of_str(modes):
                        raise TypeError(
                            "The modes parameter can only be a list of strings."
                        )
                if (
                    identifiers is not None
                    and isinstance(identifiers, list)
                    and not len(modes) == len(identifiers)
                ):  # you can give several identifiers e.g. [['sodium chloride', 'NaCl']] with modes = ['name']
                    raise ValueError(
                        "The modes and identifiers parameters must be of the same length."
                    )

            if context == "find_multiple" and not all(
                [len(mode) > 0 for mode in modes]
            ):
                raise ValueError(
                    "The list of modes cannot include an empty string or an empty list."
                )

        if identifiers is not None:
            if not identifiers:
                raise ValueError("The identifiers parameter cannot be empty.")

        if services is not None:
            if context == "get_molecule" or context == "get_molecules_batch":
                if not isinstance(services, str):
                    raise TypeError("The service parameter can only be a string.")
                if services not in self._available_services:
                    raise ValueError(f"The service {services} is not supported.")

            if context == "batch" or context == "find_single":
                if not isinstance(services, list):
                    raise TypeError(
                        "The services parameter can only be a list of strings."
                    )
                if not all(
                    [isinstance(temp_service, str) for temp_service in services]
                ):
                    raise TypeError(
                        "The services parameter can only be a list of strings."
                    )
                if not all(
                    [
                        temp_service in self._available_services
                        for temp_service in services
                    ]
                ):
                    raise ValueError(
                        "The list of services can only include supported services."
                    )

        if required_formulas is not None:
            if (
                context == "find_multiple"
                and identifiers is not None
                and not len(identifiers) == len(required_formulas)
            ):
                raise ValueError(
                    "The identifiers and required_formulas parameters must be of the same length."
                )

            if context == "get_molecule" and not isinstance(required_formulas, str):
                raise TypeError("required_formula must be a string.")

        if required_charges is not None:
            if context == "get_molecule":
                if not isinstance(required_charges, str) and not isinstance(
                    required_charges, int
                ):
                    raise TypeError(
                        "required_charge can be zero, not_zero, positive, negative or any integer."
                    )
                if isinstance(required_charges, str):
                    if required_charges not in [
                        "zero",
                        "non_zero",
                        "positive",
                        "negative",
                    ]:
                        raise ValueError(
                            "required_charge can be 'zero', 'non_zero', 'positive', 'negative' or any integer."
                        )

            if context == "find_multiple":
                if not isinstance(required_charges, list):
                    raise TypeError("The required_charges must be a list.")
                if identifiers is not None and not len(identifiers) == len(
                    required_charges
                ):
                    raise ValueError(
                        "The parameters identifiers and required_charges must be of the same length."
                    )

        if required_structure_types is not None:
            if (
                context == "find_multiple"
                and identifiers is not None
                and not len(identifiers) == len(required_structure_types)
            ):
                raise ValueError(
                    "The parameters identifiers and required_structure_types must be of the same length."
                )

            if context == "get_molecule":
                if not isinstance(required_structure_types, str):
                    raise TypeError("required_structure_type must be a string.")
                if required_structure_types not in [
                    "mixture_neutrals",
                    "mixture_ions",
                    "neutral",
                    "salt",
                    "ion",
                    "mixture_neutrals_salts",
                    "mixture_neutrals_ions",
                ]:
                    raise ValueError(
                        "required_structure_type must be one of the following: 'mixture_neutrals','mixture_ions', 'neutral', 'salt', 'ion','mixture_neutrals_salts', 'mixture_neutrals_ions'"
                    )

    def replace_non_printable_characters(self, string_input: str) -> str:
        """
        Replaces all non-printable characters from the given string.

        This function replaces different white space types by a simple white space,
        then it iterates through each character in the input string and
        keeps only the printable characters, effectively removing any non-printable
        characters such as control characters or certain Unicode characters.

        Args:
            string (str): The input string to be processed.

        Returns:
            str: A new string containing only the printable characters from the input.
        """
        string_input = regex.sub(r"[\s\u200B\u2060\uFEFF]+", " ", string_input)
        return "".join([c for c in string_input if c.isprintable()])

    @cache
    def clean_chemical_name(
        self,
        chemical_name: str,
        normalize: Optional[bool] = True,
        unescape_html: Optional[bool] = True,
        spell_out_greek_characters: Optional[bool] = False,
        for_filename: Optional[bool] = False,
    ) -> str:
        """
        Clean and standardize a chemical name string.

        Processes a chemical name to standardize its format, remove unwanted
        characters, and optionally convert special characters or prepare it for use in filenames.

        Args:
            chemical_name (str): The chemical name to be cleaned.

            normalize (Optional[bool]): If True, normalizes Unicode characters. Defaults to True.

            unescape_html (Optional[bool]): If True, unescapes HTML entities. Defaults to True.

            spell_out_greek_characters (Optional[bool]): If True, replaces Greek letters with
            their spelled-out names. Defaults to False.

            for_filename (Optional[bool]): If True, prepares the name for use as a filename
            by removing non-alphanumeric characters. Defaults to False.

        Returns:
            str: The cleaned and standardized chemical name.

        Notes:
            - Strips leading and trailing whitespace from the input.
            - Replaces various special characters with standardized alternatives.
            - Optionally unescapes HTML entities (e.g., '&#x27;' to "'").
            - Can spell out Greek letters (e.g., 'α' to 'alpha').
            - Normalizes Unicode characters to their closest ASCII representation if normalize is True.
            - When preparing for filenames, removes all non-alphanumeric characters and converts to lowercase.
            - This method is cached for performance optimization.

        Example:
            >>> clean_chemical_name("α-Pinene", spell_out_greek_characters=True)
            'alpha-Pinene'
            >>> clean_chemical_name("Sodium chloride, ≥99%", for_filename=True)
            'sodiumchloride99'
        """
        greek_letters = [
            "α",
            "β",
            "γ",
            "δ",
            "ε",
            "ϵ",
            "ζ",
            "η",
            "θ",
            "ι",
            "κ",
            "λ",
            "μ",
            "ν",
            "ξ",
            "ο",
            "π",
            "ρ",
            "σ",
            "τ",
            "υ",
            "φ",
            "χ",
            "ψ",
            "ω",
            "Α",
            "Β",
            "Γ",
            "Δ",
            "Ε",
            "Ζ",
            "Η",
            "Θ",
            "Ι",
            "Κ",
            "Λ",
            "Μ",
            "Ν",
            "Ξ",
            "Ο",
            "Π",
            "Ρ",
            "Σ",
            "Τ",
            "Υ",
            "Φ",
            "Χ",
            "Ψ",
            "Ω",
            "µ",
            "∆",
        ]
        spelled_out_versions = [
            "alpha",
            "beta",
            "gamma",
            "delta",
            "epsilon",
            "epsilon",
            "zeta",
            "eta",
            "theta",
            "iota",
            "kappa",
            "lambda",
            "mu",
            "nu",
            "xi",
            "omicron",
            "pi",
            "rho",
            "sigma",
            "tau",
            "upsilon",
            "phi",
            "chi",
            "psi",
            "omega",
            "alpha",
            "beta",
            "gamma",
            "delta",
            "epsilon",
            "zeta",
            "eta",
            "theta",
            "iota",
            "kappa",
            "lambda",
            "mu",
            "nu",
            "xi",
            "omicron",
            "pi",
            "rho",
            "sigma",
            "tau",
            "upsilon",
            "phi",
            "chi",
            "psi",
            "omega",
            "mu",
            "delta",
        ]

        map_to_replace = [
            ("’", "'"),
            ("′", "'"),
            ("‘", "'"),
            ("±", "+-"),
            ("→", "-->"),
            ("≥", ">="),
            ("≤", "<="),
            ("·", "."),
            ("#", "no. "),
            ("«", ""),
            ("»", ""),
        ]

        if spell_out_greek_characters:
            for greek_letter, spelled_out_version in zip(
                greek_letters, spelled_out_versions, strict=True
            ):
                map_to_replace.append((greek_letter, spelled_out_version))

        chemical_name = self.replace_non_printable_characters(chemical_name)
        chemical_name = chemical_name.strip()

        if unescape_html:
            chemical_name = html.unescape(chemical_name)

        for old, new in map_to_replace:
            chemical_name = chemical_name.replace(old, new)

        if normalize:
            nfkd_form = unicodedata.normalize("NFKD", chemical_name)
            chemical_name = "".join(
                [c for c in nfkd_form if not unicodedata.combining(c)]
            )

        chemical_name = regex.sub(r"\s+", " ", chemical_name)

        if for_filename:
            chemical_name = regex.sub(r"[^\w\s]", "", chemical_name.lower())
            chemical_name = chemical_name.encode("ascii", "ignore").decode("ascii")
            chemical_name = regex.sub(r"\s+", "", chemical_name)

        return chemical_name.strip()

    def filter_and_sort_synonyms(
        self,
        synonyms: list[str],
        number_of_synonyms_to_take: Optional[int] = 5,
        strict: Optional[bool] = False,
    ) -> list[str]:
        """
        Filter and sort a list of synonyms based on various criteria.

        Processes a list of synonyms, applying filters to remove unwanted entries
        and sorting the results to return the most relevant synonyms.

        Args:
            synonyms (list[str]): A list of synonym strings to process.

            number_of_synonyms_to_take (Optional[int]): The maximum number of synonyms to return.
            Defaults to 5.

            strict (Optional[bool]): If True, returns an empty list when no synonyms pass the
            filters. If False, returns ["Noname"] when no synonyms pass the filters.
            Defaults to False.

        Returns:
            list[str]: A list of filtered and sorted synonyms, with a maximum length of
            `number_of_synonyms_to_take`.

        Notes:
            - Synonyms are split on '|' characters and each part is processed separately.
            - Various heuristics are applied to filter out unwanted synonyms, including:
                * Rejecting synonyms containing specific tokens (e.g., "SCHEMBL", "AKOS", etc.)
                * Filtering based on character case and specific regex patterns.
            - Attempts to standardize the format of certain types of synonyms.
            - If no synonyms pass the filters and `strict` is False, returns ["Noname"].
        """
        # heuristics used here are mainly to remove synonyms from pubchem
        synonyms = [
            self.replace_non_printable_characters(synonym.strip())
            for synonym in synonyms
            if synonym is not None
        ]
        if len(synonyms) == 0:
            return []

        temp = []
        for synonym in synonyms:
            temp.extend(synonym.split("|"))

        synonyms_taken = []
        for synonym in temp:
            if synonym is None:
                continue

            synonym = synonym.strip()
            if len(synonym) == 0:
                continue

            if synonym is not None:
                # typical tokens to be rejected
                tokens = [
                    "SCHEMBL",
                    "AKOS",
                    "MFCD",
                    "AKOS",
                    "CYPHOS",
                    "DTXSID",
                    "UNII",
                    "NSC",
                    "DSSTox",
                    "FLUKA",
                    "ACMC",
                    "%",
                    "(at)",
                    "SIGMA",
                    "ALDRICH",
                    "RIEDEL",
                    "SIAL",
                ]
                n_tokens_found = sum([synonym.lower().count(t.lower()) for t in tokens])
                if n_tokens_found > 0:
                    continue

                if not synonym[-1].isalpha() and synonym[-1] != "-":
                    continue

                # if completely uppercase convert to lower
                if synonym.isupper():
                    synonym = synonym.lower()

                ratio_uppercase_letters = sum(1 for c in synonym if c.isupper()) / len(
                    synonym
                )

                if ratio_uppercase_letters < 0.5:
                    match = regex.match(
                        self.CAS_regex_with_groups, synonym
                    )  # filter CAS numbers
                    # filter names with 3 consecutive numbers or containing some keywords
                    match2 = regex.search(
                        r"\d{3}|\bin\b|\bgrade\b|\bsolution\b|\bstandard\b|\bstabilized\b|\bdimer\b|\bfor\b|\btablet\b|\btotal\b|\bcode\b|\bNo\.\b|\banhydrous\b",
                        synonym,
                        regex.IGNORECASE,
                    )
                    if not match and not match2:
                        synonym = synonym.replace("<em>", "").replace(
                            "</em>", ""
                        )  # replace(';', ', ')

                        temp_synonym = synonym.split(", ")
                        if len(temp_synonym) > 1:
                            n_parts_ending_in_hyphen = 0
                            for i, part in enumerate(temp_synonym):
                                if i == 0:
                                    continue
                                if part.endswith("-"):
                                    n_parts_ending_in_hyphen += 1
                            if n_parts_ending_in_hyphen == len(temp_synonym) - 1:
                                temp_synonym.reverse()
                                synonym = "".join(temp_synonym)

                        synonyms_taken.append(synonym)

        final_synonyms_taken = self.take_most_common(
            synonyms_taken, number_of_synonyms_to_take
        )

        if len(final_synonyms_taken) == 0:
            if not strict:
                if len(synonyms) == 0:
                    synonyms.append("Noname")
                final_synonyms_taken = [synonyms[0]]

        return final_synonyms_taken

    def expand_name_heuristically(
        self,
        name: str,
        prefixes_to_delete: Optional[list[str]] = None,
        suffixes_to_use_as_prefix: Optional[list[str]] = None,
        suffixes_to_delete: Optional[list[str]] = None,
        parts_to_delete: Optional[list[str]] = None,
        maps_to_replace: Optional[dict[str, str]] = None,
    ) -> str:
        """
        Expand and standardize a chemical name using various heuristics.

        Applies a series of transformations to standardize and potentially
        expand a given chemical name.

        Args:
            name (str): The chemical name to process.

            prefixes_to_delete (Optional[list[str]]): Prefixes to remove from the name.

            suffixes_to_use_as_prefix (Optional[list[str]]): Suffixes to move to the
            beginning of the name as prefixes.

            suffixes_to_delete (Optional[list[str]]): Suffixes to remove from the name.

            parts_to_delete (Optional[list[str]]): Specific parts to remove from the name.

            maps_to_replace (Optional[dict[str, str]]): A dictionary of string replacements
            to apply to the name.

        Returns:
            str: The processed chemical name after applying all specified transformations.

        Notes:
            - Unescapes HTML entities in the name.
            - Handles various formatting issues, such as reversing comma-separated parts.
            - Default lists are provided for prefixes, suffixes, and parts to delete if not specified.
            - Applies a series of regex-based transformations to standardize the name format.
            - Special handling is implemented for stereochemistry indicators (e.g., cis/trans, E/Z).
        """
        name = self.replace_non_printable_characters(name.strip())
        original_name = name
        names = [original_name]
        name = html.unescape(name)

        name_without_html_tags = self.html_tag_regex_compiled.sub("", name)
        if name != name_without_html_tags:
            if "()" not in name_without_html_tags:
                name = name_without_html_tags
                names.append(name_without_html_tags)

        # eliminate trailing wrong apostrophe or prime, double prime
        name = regex.sub(r"(?<!\d)([\"′″'‘’]+)$", "", name).strip()

        temp_name = name
        # sometimes names are like this 2,6,10-Trimethyldodecane (Farnesane)
        m = regex.match("(.*?) \((\w*)\)$", name)
        if m:
            temp_name = m.group(1).strip()
            names.append(temp_name)
            names.append(m.group(2).strip())

        temp = name

        # add
        # acid, alskjdalskjd acid
        # amine,N,N-  amine
        # delete empty spaces in some places 2, 3-dimethil
        # α -> alpha
        # β -> beta   (dl)

        name_parts = name.split(", ")
        if len(name_parts) > 1:
            n_parts_ending_in_hyphen = 0
            for i, part in enumerate(name_parts):
                if i == 0 and part.endswith("-"):
                    break
                elif part.endswith("-"):
                    n_parts_ending_in_hyphen += 1
            if n_parts_ending_in_hyphen == 0:
                name_parts.reverse()
                names.append(" ".join(name_parts))
            elif n_parts_ending_in_hyphen == len(name_parts) - 1:
                name_parts.reverse()
                names.append("".join(name_parts))

        if prefixes_to_delete is None:
            prefixes_to_delete = [
                "±",
                "\+\,-",
                "\+/-",
                "\+-",
                "\.\+\,-\.",
                "\.\+/-\.",
                "\.\+-\.",
                "(dl)",
                "DL",
                "RS",
                "\(<\+\->\)\-",
                "\(2H\d+\)",
                "\[2H\d+\]",
                "[Nn]-",
            ]

        if suffixes_to_delete is None:
            suffixes_to_delete = [
                "mixed isomers",
                "and isomers",
                "isomers",
                "tautomers",
                "dl and meso",
                "±",
                "\+\,-",
                "\+/-",
                "\+-",
                "\.\+\,-\.",
                "\.\+/-\.",
                "\.\+-\.",
                "cis and trans",
                "-d2",
            ]

        if suffixes_to_use_as_prefix is None:
            suffixes_to_use_as_prefix = ["R", "S"]

        if parts_to_delete is None:
            parts_to_delete = ["±"]

        if maps_to_replace is None:
            maps_to_replace = [
                [r"([a-zA-Z]{2,})-([a-zA-Z]{2,})", r"\1\2"],
                [r"trans(-\d+-[a-z]{3,}ene)\b", r"(E)\1"],
                [r"trans(-\d+-[a-z]{3,}ene)\b", r"E\1"],
                [r"cis(-\d+-[a-z]{3,}ene)\b", r"(Z)\1"],
                [r"cis(-\d+-[a-z]{3,}ene)\b", r"Z\1"],
                [r"(.*)(\((?:Z|E)\)-)(.*)", r"\2\1\3"],
                [r"«|»", r""],
                [r"- -", "-"],
                ["flouro", "fluoro"],
                ["-[Nn]-", "-"],
                ["([A-Za-z]{2,}),([A-Za-z]{2,})", r"\1, \2"],
                ["′", "'"],
                ["″", "''"],
                [r"[\[\]]", lambda match: "(" if match.group(0) == "[" else ")"],
            ]

        new_name = name
        for prefix in prefixes_to_delete:
            m = regex.match(rf"\s*\(?\s*{prefix}\s*\)?\s*-?(.*)$", new_name)
            if m:
                new_name = m.group(1)

        for suffix in suffixes_to_delete:
            m = regex.match(rf"(.*)[;,]+\s*\(?\s*{suffix}\s*\)?\s*-?\s*$", new_name)
            if m:
                new_name = m.group(1)

        if new_name not in names:
            names.append(new_name)

        for suffix in suffixes_to_use_as_prefix:
            m = regex.match(rf"(.*)[;,]*\s+\(?\s*{suffix}\s*\)?\s*-?\s*$", new_name)
            if m:
                new_name = f"({suffix})-{m.group(1)}"
                if new_name not in names:
                    names.append(new_name)

        for part in parts_to_delete:
            for temporary_name in names:
                new_name = regex.sub(
                    rf"([;, ]*\(?\s*{part}\s*\)?\s*-?)", "", temporary_name
                )
                if new_name not in names:
                    names.append(new_name)

        for temporary_name in names:
            for pattern, replacement in maps_to_replace:
                ms = regex.findall(pattern, temporary_name)
                if ms:
                    if all(
                        [
                            v not in ms[0]
                            for v in [
                                "trans",
                                "cis",
                                "alpha",
                                "beta",
                                "tert",
                                "sec",
                                "gamma",
                                "erythro",
                                "threo",
                            ]
                        ]
                    ):
                        new_name = regex.sub(pattern, replacement, temporary_name)
                        if new_name not in names:
                            names.append(new_name)

        if names[0] != original_name:
            if original_name in names:
                names.remove(original_name)
            return names.insert(0, original_name)
        return names

    def filter_and_sort_CAS(self, synonyms: list[str]) -> list[str]:
        """
        Filter and sort CAS (Chemical Abstracts Service) registry numbers from a list of synonyms.

        Processes a list of synonyms to extract valid CAS registry numbers,
        remove duplicates, and sort them based on frequency and heuristic approach.

        Args:
            synonyms (list[str]): A list of strings that may contain CAS registry numbers.

        Returns:
            list[str]: A list of unique, valid CAS registry numbers, sorted by frequency
            and heuristic rules. If no valid CAS numbers are found, returns an empty list.

        Notes:
            - CAS numbers are validated using the `is_valid_CAS` method.
            - Duplicate CAS numbers are removed, keeping only unique entries.
            - If all CAS numbers in the input are identical, only one is returned.
            - If multiple unique CAS numbers are found with equal frequency:
              They are sorted based on the numeric value of their first segment.
              The CAS number with the lowest first segment is returned first.
            - If CAS numbers have different frequencies, they are sorted by frequency
              (most common first).
            - Whitespace is stripped from each synonym before processing.
        """
        CAS_rns = []

        if synonyms is not None:
            for synonym in synonyms:
                if synonym:
                    synonym = synonym.strip()
                    if self.is_valid_CAS(synonym):
                        CAS_rns.append(synonym)

        if not CAS_rns:
            return CAS_rns

        counter = collections.Counter(CAS_rns)

        # if all CAS provided are the same, use the first one
        if len(counter) == 1:
            return [CAS_rns[0]]

        # if all CAS provided are as likely as the other use heuristic approach:
        # in my experience the CAS with the lowest first number is mostly the correct one.
        if all([v == 1 for v in counter.values()]):
            CAS_rns = sorted(CAS_rns, key=lambda CAS: int(CAS.split("-")[0]))
            return [CAS_rns[0]]

        return [counter.most_common()[0][0]]

    def combine_molecules(
        self,
        grouped_SMILES: str,
        molecules: list[Molecule],
    ) -> Optional[Molecule]:
        """
        Combine multiple Molecule objects into a single Molecule.

        Merges information from multiple Molecule objects that represent
        the same chemical structure.

        Args:
            grouped_SMILES (str): The SMILES representation of the combined molecule.

            molecules (list[Molecule]): A list of Molecule objects to combine.

        Returns:
            Optional[Molecule]: A new Molecule object combining information from all input molecules,
            or None if combination fails.

        Raises:
            ValueError: If molecules have different structures or identifiers when from the same service.

        Notes:
            - Combines synonyms, CAS numbers, and other information from all input molecules.
            - Prioritizes CAS numbers from official registry if available.
            - Handles cases where molecules are from the same or different services.
        """
        if not molecules:
            return None

        if len(molecules) == 1:
            return molecules[0]

        all_SMILES_are_equal = all(
            [
                self.are_equal(
                    self.get_from_SMILES(cmp.SMILES),
                    self.get_from_SMILES(molecules[0].SMILES),
                )
                for cmp in molecules
            ]
        )
        all_identifiers_are_equal = all(
            [cmp.identifier == molecules[0].identifier for cmp in molecules]
        )
        all_services_are_equal = all(
            [cmp.service == molecules[0].service for cmp in molecules]
        )

        if not all_SMILES_are_equal or (
            all_services_are_equal and not all_identifiers_are_equal
        ):
            raise ValueError(
                "Combining molecules is only allowed if the structures are the same and found by the same identifier in the case of the same service."
            )

        all_molecules_used_same_service = all(
            [cmp.service == molecules[0].service for cmp in molecules]
        )

        merged_synonyms = []
        merged_CAS = []
        merged_modes = []
        merged_services = []
        merged_additional_information = []

        for molecule in molecules:
            merged_additional_information.append(molecule.additional_information)
            merged_CAS.extend(molecule.CAS)
            merged_synonyms.extend(molecule.synonyms)
            merged_modes.append(molecule.mode)
            merged_services.append(molecule.service)

        merged_synonyms = self.filter_and_sort_synonyms(merged_synonyms)

        # prefer CAS from official registry
        if "cas_registry" in merged_services:
            index_cas_registry = merged_services.index("cas_registry")
            merged_CAS = molecules[index_cas_registry].CAS
        else:
            merged_CAS = self.filter_and_sort_CAS(merged_CAS)

        if all_molecules_used_same_service:
            merged_additional_information = [
                v for v in merged_additional_information if v is not None
            ]
            if merged_additional_information:
                merged_additional_information = str(
                    sorted(merged_additional_information)
                )
            else:
                merged_additional_information = ""
            merged_modes = str(sorted(merged_modes))
            merged_services = merged_services[0]
            number_of_crosschecks = 1
        else:
            merged_modes = "; ".join(merged_modes)
            merged_additional_information = "; ".join(merged_additional_information)
            number_of_crosschecks = len(merged_services)
            merged_services = "; ".join(merged_services)

        return Molecule(
            grouped_SMILES,
            merged_synonyms,
            merged_CAS,
            merged_additional_information,
            merged_modes,
            merged_services,
            number_of_crosschecks,
            molecules[0].identifier,
        )

    def group_molecules_by_structure(
        self, molecules: list[Molecule], group_also_by_services: Optional[bool] = True
    ) -> dict[str, list[Molecule]]:
        """
        Group a list of Molecule objects by their structural similarity.

        Organizes molecules into groups based on their structural equivalence,
        optionally considering the services they come from.

        Args:
            molecules (list[Molecule]): A list of Molecule objects to group.

            group_also_by_services (Optional[bool]): If True, groups molecules by both
            structure and service. Defaults to True.

        Returns:
            dict[str, list[Molecule]]: A dictionary where keys are SMILES strings and
            values are lists of structurally equivalent Molecule objects.

        Raises:
            ValueError: If input contains molecules from different sources that were
            previously combined.

        Notes:
            - Uses structural comparison to group molecules.
            - Can handle molecules from primary sources (not previously combined).
            - When all molecules are from the same service and mode, it may select
              the most common structure as the representative.
        """
        if len(molecules) == 1:
            return {molecules[0].SMILES: molecules}

        # if all molecules given are from a primary service, combine them for each service if needed
        all_molecules_from_primary_source = all(
            [
                cmp.service in self._available_services
                and cmp.number_of_crosschecks == 1
                for cmp in molecules
            ]
        )
        if all_molecules_from_primary_source:
            grouped_molecules = {}
            grouped_molecules_SMILES = {}
            grouped_molecules_molecules = {}
            all_keys = []
            for cmp in molecules:
                cmp_mol = self.get_from_SMILES(cmp.SMILES)
                equal_SMILES = None
                old_key = None
                for key, grouped_molecule_mol in grouped_molecules_molecules.items():
                    if self.are_equal(cmp_mol, grouped_molecule_mol):
                        if equal_SMILES is None:
                            equal_SMILES = grouped_molecules_SMILES[key]
                        equal_SMILES = sorted(
                            [equal_SMILES, grouped_molecules_SMILES[key]]
                        )[0]
                        old_key = key
                        break

                if equal_SMILES is None:
                    equal_SMILES = cmp.SMILES

                if group_also_by_services:
                    key = f"{cmp.service}_{equal_SMILES}"
                else:
                    key = equal_SMILES

                if old_key:
                    if key != old_key:
                        grouped_molecules[key] = grouped_molecules[old_key]
                        del grouped_molecules[old_key]

                if key not in grouped_molecules:
                    grouped_molecules[key] = []
                    grouped_molecules_SMILES[key] = equal_SMILES
                    grouped_molecules_molecules[key] = self.get_from_SMILES(
                        equal_SMILES
                    )

                grouped_molecules[key].append(cmp)
                all_keys.append(key)

            # if all where searched with same service, mode and are a primary source
            # take the most common structure
            if len(grouped_molecules) > 1:
                all_molecules_used_same_service_and_mode = all(
                    [
                        cmp.mode == molecules[0].mode
                        and cmp.service == molecules[0].service
                        for cmp in molecules
                    ]
                )
                if all_molecules_used_same_service_and_mode:
                    smiles_count = collections.Counter(all_keys).most_common()
                    if smiles_count[0][1] > smiles_count[1][1]:
                        key = smiles_count[0][0]
                        return {grouped_molecules_SMILES[key]: grouped_molecules[key]}

            grouped_molecules = {
                grouped_molecules_SMILES[key]: grouped_molecules[key]
                for key in all_keys
            }
        else:
            raise ValueError(
                "This function only takes molecules that were not previously combined from different sources."
            )

        return grouped_molecules

    def filter_molecules(
        self,
        molecules: list[Molecule],
        required_formula: Optional[str] = None,
        required_charge: Optional[Union[int, str]] = None,
        required_structure_type: Optional[str] = None,
    ) -> list[Molecule]:
        """
        Filter a list of Molecule objects based on specified criteria.

        Filters molecules based on their chemical formula, charge,
        and structure type.

        Args:
            molecules (list[Molecule]): A list of Molecule objects to filter.

            required_formula (Optional[str]): The required chemical formula.

            required_charge (Optional[Union[int, str]]): The required molecular charge.

            required_structure_type (Optional[str]): The required structure type.

        Returns:
            list[Molecule]: A list of Molecule objects that meet the specified criteria.

        Notes:
            - Filters out molecules with invalid or None SMILES.
            - Applies standardization to SMILES strings of valid molecules.
            - For 'salt' structure type, attempts to disconnect more metals if initial check fails.
            - Uses the `check_SMILES` method to validate molecules against criteria.
        """
        filtered_molecules = []

        if not isinstance(molecules, list):
            molecules = [molecules]
        if len(molecules) == 0:
            return None

        for cmp in molecules:
            if cmp is None:
                continue
            if cmp.SMILES is None:
                continue

            mol = self.get_from_SMILES(cmp.SMILES)
            if mol is None:
                continue

            if self.check_SMILES(
                cmp.SMILES, required_formula, required_charge, required_structure_type
            ):
                cmp.SMILES = self.standardize_SMILES(cmp.SMILES)
                filtered_molecules.append(cmp)
            elif required_structure_type == "salt":
                if self._standardization_options.disconnect_more_metals_for_salts:
                    new_SMILES = self.try_disconnect_more_metals(cmp.SMILES)
                    if self.check_SMILES(
                        new_SMILES,
                        required_formula,
                        required_charge,
                        required_structure_type,
                    ):
                        cmp.SMILES = new_SMILES
                        filtered_molecules.append(cmp)

        return filtered_molecules

    def filter_and_combine_molecules(
        self,
        molecules: Union[list[Molecule], Molecule],
        required_formula: Optional[str] = None,
        required_charge: Optional[Union[int, str]] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Filter and combine a list of molecules based on specified criteria.

        Processes a list of molecules (or a single molecule) by applying
        filters and then combining the filtered results into a single molecule.

        Args:
            molecules (Union[list[Molecule], Molecule]): A list of Molecule objects or a single Molecule object to process.

            required_formula (Optional[str]): The required chemical formula to filter by. Defaults to None.

            required_charge (Optional[Union[int, str]]): The required molecular charge to filter by. Defaults to None.

            required_structure_type (Optional[str]): The required structure type to filter by. Defaults to None.

        Returns:
            Optional[Molecule]: A single Molecule object that results from filtering and combining the input molecules.
            Returns None if no molecules pass the filtering process or if the combination process fails.

        Notes:
            - If a single Molecule is provided, it's converted to a list for processing.
            - First filters the molecules using `filter_molecules`.
            - Filtered molecules are then grouped by structure using `group_molecules_by_structure`.
            - For each group of structurally similar molecules, `combine_molecules` is called.
            - If only one final molecule remains after combining, it is returned.
            - If no molecules pass the filters or if multiple distinct molecules remain after combining,
              the method returns None.
        """
        filtered_molecules = self.filter_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

        if not filtered_molecules:
            return None

        grouped_molecules = self.group_molecules_by_structure(filtered_molecules)

        final_molecules = []
        for grouped_SMILES, cmps_to_combine in grouped_molecules.items():
            final_molecules.append(
                self.combine_molecules(grouped_SMILES, cmps_to_combine)
            )

        # until I have a better algorithm, take first molecule whenever
        # one service returns more than one molecules
        all_molecules_from_same_service = all(
            [m.service == filtered_molecules[0].service for m in filtered_molecules]
        )

        if len(final_molecules) == 1 or all_molecules_from_same_service:
            return final_molecules[0]
        else:
            return None

    @cache
    def formula_to_dictionary(
        self, formula: str, allow_duplicate_atoms: Optional[bool] = True
    ) -> dict[str, Union[int, float]]:
        """
        Convert a chemical formula string to a dictionary of element counts.

        Parses a chemical formula and returns a dictionary where keys are
        element symbols and values are their respective counts in the formula.

        Args:
            formula (str): The chemical formula to parse.

            allow_duplicate_atoms (Optional[bool]): If True, allows multiple occurrences of the same
            element in the formula. If False, raises an error for duplicate elements.
            Defaults to True.

        Returns:
            dict[str, Union[int, float]]: A dictionary where keys are element symbols and
            values are their counts. Values are integers if all counts are whole numbers,
            otherwise floats.

        Raises:
            ValueError: If `allow_duplicate_atoms` is False and an element appears more than once.

        Notes:
            - Supports nested parentheses, square brackets, and curly braces in formulas.
            - Handles hydrates and solvates (formulas connected with '*').
            - Processes fractional counts (e.g., 0.5) and converts to integers when possible.
            - Uses recursion to handle nested structures in the formula.
            - Caches results for improved performance on repeated calls with the same formula.
        """

        def merge_dictionaries(
            d1: dict[str, Union[int, float]],
            d2: dict[str, Union[int, float]],
            d2_multiplier: float = 1.0,
        ):
            d1 = d1.copy()
            for element, count in d2.items():
                if element not in d1:
                    d1[element] = 0
                d1[element] += count * d2_multiplier
            return d1

        def cast_multiplier(multiplier: str) -> float:
            if len(multiplier) > 0:
                return float(multiplier)
            else:
                return 1.0

        d = {}
        if "*" in formula:
            for formula_part in formula.split("*"):  # for hydrates and solvates
                temp_d = self.formula_to_dictionary(formula_part)
                d = merge_dictionaries(d, temp_d)
                return d

        formula = formula.replace("[", "(").replace("{", "(")
        formula = formula.replace("]", ")").replace("}", ")")
        brackets_match = self.formula_bracket_group_regex_compiled.search(formula)
        if brackets_match:
            rest_of_formula = (
                formula[0 : brackets_match.span()[0]]
                + formula[brackets_match.span()[1] :]
            )
            bracket_content = brackets_match.groups()[0][1:-1]
            bracket_multiplier = cast_multiplier(brackets_match.groups()[1])
            d = merge_dictionaries(
                self.formula_to_dictionary(rest_of_formula, allow_duplicate_atoms),
                self.formula_to_dictionary(bracket_content, allow_duplicate_atoms),
                bracket_multiplier,
            )
        else:
            matches = self.empirical_formula_regex_compiled.findall(formula)
            for match in matches:
                element = match[0]
                multiplier = cast_multiplier(match[1])
                if allow_duplicate_atoms:
                    if element not in d:
                        d[element] = 0
                    d[element] += multiplier
                else:
                    if element in d:
                        raise ValueError("element twice in formula")
                    d[element] = multiplier

        only_integers_available = all([float(v).is_integer() for v in d.values()])
        conversion_function = int if only_integers_available else float
        for k, v in d.items():
            d[k] = conversion_function(v)

        return d

    def to_hill_formula(
        self, mol_or_dictionary: Union[Chem.rdchem.Mol, dict[str, Union[int, float]]]
    ) -> str:
        """
        Convert a molecule or element dictionary to a Hill formula string.

        Generates a Hill formula representation of a molecule, either from
        an RDKit molecule object or a dictionary of element counts.

        Args:
            mol_or_dictionary (Union[Chem.rdchem.Mol, dict[str, Union[int, float]]]):
            Either an RDKit molecule object or a dictionary where keys are element
            symbols and values are their counts.

        Returns:
            str: The Hill formula string representation of the molecule.

        Notes:
            - If input is an RDKit molecule, hydrogens are added explicitly before processing.
            - The Hill system orders elements as follows:

              1. Carbon (C) always comes first if present.
              2. Hydrogen (H) comes second if carbon is present.
              3. All other elements follow in alphabetical order.
            - For elements with a count of 1, the number is omitted in the formula.
            - Handles both integer and float counts, though typically counts are integers.
            - If a count is a float but effectively an integer (e.g., 2.0), it's treated as an integer.

        Example:
            >>> mol = Chem.MolFromSmiles("CCO")
            >>> to_hill_formula(mol)
            'C2H6O'
            >>> to_hill_formula({"C": 2, "H": 6, "O": 1})
            'C2H6O'
        """

        if isinstance(mol_or_dictionary, dict):
            temp = mol_or_dictionary
        else:
            mol = Chem.AddHs(mol_or_dictionary)
            temp = {}
            for atom in mol.GetAtoms():
                symbol = atom.GetSymbol()
                if symbol in temp:
                    temp[symbol] += 1
                else:
                    temp[symbol] = 1

        hill_formula = ""

        if "C" in temp:
            hill_formula = hill_formula + "C"

            if temp["C"] > 1:
                hill_formula = hill_formula + str(temp["C"])

            temp.pop("C")

            if "H" in temp:
                hill_formula = hill_formula + "H"

                if temp["H"] > 1:
                    hill_formula = hill_formula + str(temp["H"])

                temp.pop("H")

        for key in sorted(temp.keys()):
            hill_formula = hill_formula + key

            if temp[key] > 1:
                hill_formula = hill_formula + str(temp[key])

        return hill_formula

    def check_formula(
        self,
        mol_or_formula: Union[Chem.rdchem.Mol, str],
        required_formula: Optional[str],
    ) -> bool:
        """
        Check if a molecule or formula matches a required chemical formula.

        Compares the chemical formula of a given molecule or formula string
        against a required formula.

        Args:
            mol_or_formula (Union[Chem.rdchem.Mol, str]): The molecule (as an RDKit Mol object)
            or a chemical formula string to check.

            required_formula (Optional[str]): The required chemical formula to match against.
            If None, the check always returns True.

        Returns:
            bool: True if the molecule or formula matches the required formula, False otherwise.

        Notes:
            - If `required_formula` is None, the method always returns True.
            - For RDKit Mol objects, the molecule is standardized before calculating its formula.
            - The comparison is done by converting both the input and required formulas to
              dictionaries using `formula_to_dictionary`.
        """
        if required_formula is None:
            return True

        if isinstance(mol_or_formula, str):
            mol_formula = mol_or_formula
        else:
            temp_mol = self.standardize_mol(mol_or_formula)
            mol_formula = rdMolDescriptors.CalcMolFormula(temp_mol)

        formula1 = self.formula_to_dictionary(mol_formula)
        formula2 = self.formula_to_dictionary(required_formula)

        return formula1 == formula2

    def check_charge(
        self, mol: Chem.rdchem.Mol, required_charge: Optional[Union[int, str]]
    ) -> bool:
        """
        Check if a molecule's charge matches a required charge.

        Compares the formal charge of a given molecule against a required charge.

        Args:
            mol (Chem.rdchem.Mol): The molecule (as an RDKit Mol object) to check.

            required_charge (Optional[Union[int, str]]): The required charge to match against.
            Can be an integer or one of the strings: "zero", "non_zero", "positive", "negative".
            If None, the check always returns True.

        Returns:
            bool: True if the molecule's charge matches the required charge, False otherwise.

        Notes:
            - If `required_charge` is None, the method always returns True.
            - Uses RDKit's `GetFormalCharge` to calculate the molecule's charge.
            - String inputs for `required_charge` allow for more flexible charge requirements:
              * "zero": Charge must be exactly 0.
              * "non_zero": Charge must not be 0.
              * "positive": Charge must be greater than 0.
              * "negative": Charge must be less than 0.
        """
        if required_charge is None:
            return True

        self._check_parameters(required_charges=required_charge)
        calculated_charge = Chem.rdmolops.GetFormalCharge(mol)

        charge_correct = False
        if isinstance(required_charge, str):
            if required_charge == "zero":
                charge_correct = calculated_charge == 0
            elif required_charge == "non_zero":
                charge_correct = calculated_charge != 0
            elif required_charge == "positive":
                charge_correct = calculated_charge > 0
            elif required_charge == "negative":
                charge_correct = calculated_charge < 0
        else:
            charge_correct = int(required_charge) == calculated_charge

        return charge_correct

    def check_molecular_mass(
        self,
        mol: Chem.rdchem.Mol,
        required_molecular_mass: Union[float, int],
        percentage_deviation_allowed: Optional[float] = 0.001,
    ) -> bool:
        """
        Check if a molecule's mass matches a required molecular mass within a specified tolerance.

        Compares the calculated molecular mass of a given molecule against a
        required molecular mass, allowing for a specified percentage of deviation.

        Args:
            mol (Chem.rdchem.Mol): The molecule to check.

            required_molecular_mass (Union[float, int]): The required molecular mass to match.

            percentage_deviation_allowed (Optional[float]): Allowed percentage deviation between the
            calculated and required molecular mass. Defaults to 0.001 (0.1%).

        Returns:
            bool: True if the molecule's mass is within the allowed deviation of the required mass,
            False otherwise.

        Raises:
            TypeError: If `required_molecular_mass` is not a float or an integer.

        Notes:
            - Uses RDKit's `Descriptors.MolWt` to calculate the molecule's mass.
            - The comparison uses the absolute relative difference between the calculated and
              required mass, allowing for the specified percentage of deviation.
            - The minimum value between the required and calculated mass is used as the
              denominator when calculating the relative difference to handle cases where
              one of the values might be zero or very small.
        """
        if not isinstance(required_molecular_mass, float) and not isinstance(
            required_molecular_mass, int
        ):
            raise TypeError("required_molecular_mass must be a float or an integer.")

        molecular_mass_from_mol = Descriptors.MolWt(mol)
        minimum_value = min(required_molecular_mass, molecular_mass_from_mol)
        absolute_relative_difference = abs(
            (required_molecular_mass - molecular_mass_from_mol) / minimum_value
        )
        return absolute_relative_difference < percentage_deviation_allowed

    @cache
    def check_SMILES(
        self,
        SMILES: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[Union[int, str]] = None,
        required_structure_type: Optional[str] = None,
    ) -> bool:
        """
        Validate a SMILES string against specified requirements.

        Checks if a given SMILES string represents a valid molecule and
        meets the optional requirements for formula, charge, and structure type.

        Args:
            SMILES (str): The SMILES string to validate.

            required_formula (Optional[str]): The expected molecular formula. Defaults to None.

            required_charge (Optional[Union[int, str]]): The expected molecular charge. Defaults to None.

            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            bool: True if the SMILES string is valid and meets all specified requirements,
            False otherwise.

        Notes:
            - Uses caching to improve performance for repeated calls with the same arguments.
            - Performs a quick validity check on the SMILES string.
            - If the SMILES string is valid, creates a molecule object for further checks.
            - Verifies the formula, charge, and structure type if respective requirements are provided.
        """
        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
        )

        if not self.is_valid_SMILES_fast(SMILES):
            return False

        mol = self.get_from_SMILES(SMILES)

        if mol is None:
            return False

        if not self.check_formula(mol, required_formula):
            return False

        if not self.check_charge(mol, required_charge):
            return False

        if not self.check_structure_type(SMILES, required_structure_type):
            return False

        return True

    def get_structure_type_from_SMILES(self, SMILES: str) -> Optional[str]:
        """
        Determine the structure type of a molecule from its SMILES representation.

        Analyzes the SMILES string to classify the molecule into different
        structure types based on its composition and charge distribution.

        Args:
            SMILES (str): The SMILES representation of the molecule.

        Returns:
            Optional[str]: A string representing the structure type of the molecule. Possible values are:

            - "mixture_neutrals": A mixture of neutral molecules.
            - "neutral": A single neutral molecule.
            - "salt": A salt (mixture of oppositely charged ions with total charge of 0).
            - "mixture_ions": A mixture of ions with non-zero total charge.
            - "ion": A single charged molecule.
            - "mixture_neutrals_salts": A mixture of neutral molecules and salts.
            - "mixture_neutrals_ions": A mixture of neutral molecules and ions.
            - None: If the SMILES cannot be parsed.

        Notes:
            - Splits the SMILES on '.' to handle mixtures.
            - Uses RDKit to calculate formal charges for each part of the SMILES.
            - Classification is based on the number of parts and their charges.
            - Returns None if any part of the SMILES fails to parse.
        """
        SMILES_parts = SMILES.split(".")
        SMILES_parts_charges = []
        for SMILES_part in SMILES_parts:
            mol = self.get_from_SMILES(SMILES_part)
            if not mol:
                return "None"
            SMILES_parts_charges.append(Chem.rdmolops.GetFormalCharge(mol))

        total_charge = sum(SMILES_parts_charges)

        if all([charge == 0 for charge in SMILES_parts_charges]):
            if len(SMILES_parts) > 1:
                structure_type = "mixture_neutrals"
            else:
                structure_type = "neutral"
        elif all([charge != 0 for charge in SMILES_parts_charges]):
            if len(SMILES_parts) > 1:
                if total_charge == 0:
                    structure_type = "salt"
                else:
                    structure_type = "mixture_ions"
            else:
                structure_type = "ion"
        else:
            if total_charge == 0:
                structure_type = "mixture_neutrals_salts"
            else:
                structure_type = "mixture_neutrals_ions"

        return structure_type

    def check_structure_type(self, SMILES: str, required_structure_type: str) -> bool:
        """
        Check if a molecule's structure type matches a required structure type.

        Compares the structure type of a molecule (determined from its SMILES)
        against a required structure type.

        Args:
            SMILES (str): The SMILES representation of the molecule to check.

            required_structure_type (str): The required structure type to match against.
            Must be one of the following:
            "mixture_neutrals", "mixture_ions", "neutral", "salt", "ion",
            "mixture_neutrals_salts", "mixture_neutrals_ions".

        Returns:
            bool: True if the molecule's structure type matches the required type, False otherwise.

        Raises:
            ValueError: If `required_structure_type` is not one of the allowed values.

        Notes:
            - If `required_structure_type` is None, the method always returns True.
            - Uses `get_structure_type_from_SMILES` to determine the actual structure type.
            - The comparison is case-sensitive and must match exactly.
        """
        if required_structure_type is None:
            return True

        available_structure_types = [
            "mixture_neutrals",
            "mixture_ions",
            "neutral",
            "salt",
            "ion",
            "mixture_neutrals_salts",
            "mixture_neutrals_ions",
        ]
        if required_structure_type not in available_structure_types:
            raise ValueError(
                "required_structure_type must be one of 'mixture_neutrals','mixture_ions', 'neutral', 'salt', 'ion','mixture_neutrals_salts' or 'mixture_neutrals_ions'."
            )

        return self.get_structure_type_from_SMILES(SMILES) == required_structure_type

    @cache
    def InChI_to_SMILES(self, inchi: str) -> Optional[str]:
        """
        Convert an InChI string to a standardized SMILES string.

        Converts an InChI representation of a molecule to its standardized SMILES form.

        Args:
            inchi (str): The InChI string to convert.

        Returns:
            Optional[str]: The standardized SMILES string, or None if conversion fails.

        Notes:
            - This method is cached to improve performance for repeated calls with the same InChI.
            - Uses `get_from_InChI` to convert InChI to a molecule object.
            - Returns None if the InChI cannot be converted to a molecule.
        """
        mol = self.get_from_InChI(inchi)

        if mol is None:
            return None

        return Chem.MolToSmiles(mol)

    @cache
    def SMILES_to_InChI(self, smiles: str) -> Optional[str]:
        """
        Convert a SMILES string to an InChI string.

        Converts a SMILES (Simplified Molecular Input Line Entry System) string
        to its InChI (International Chemical Identifier) representation.

        Args:
            smiles (str): The SMILES string to convert.

        Returns:
            Optional[str]: The InChI string if conversion is successful, None otherwise.

        Notes:
            - Uses `get_from_SMILES` to create an RDKit molecule object from the SMILES.
            - This method is cached for performance optimization.
        """
        mol = self.get_from_SMILES(smiles)

        if mol is None:
            return None

        return Chem.MolToInchi(mol)

    @cache
    def get_from_InChI(
        self, inchi: str, addHs: Optional[bool] = False
    ) -> Optional[Chem.rdchem.Mol]:
        """
        Create an RDKit molecule object from an InChI string.

        Attempts to create an RDKit molecule object from a given InChI string,
        with options for handling hydrogens.

        Args:
            inchi (str): The InChI string to convert.
            addHs (Optional[bool]): If True, explicitly adds hydrogen atoms to the molecule. Defaults to False.

        Returns:
            Optional[Chem.rdchem.Mol]: An RDKit molecule object if conversion is successful, None otherwise.

        Notes:
            - Returns None if the input InChI is None or an empty string.
            - First attempts to create the molecule with sanitization.
            - If that fails, attempts to create the molecule without sanitization.
            - Updates the property cache of the molecule, with fallback to non-strict update.
            - Calculates the Smallest Set of Smallest Rings (SSSR) for the molecule.
            - Optionally adds explicit hydrogens based on the `addHs` parameter.
            - This method is cached for performance optimization.
        """
        if inchi is None:
            return None
        if inchi == "":
            return None

        mol = Chem.MolFromInchi(inchi)
        if mol is None:
            mol = Chem.MolFromInchi(inchi, sanitize=False)

        if mol is None:
            return None

        try:
            mol.UpdatePropertyCache()
        except Exception:
            mol.UpdatePropertyCache(strict=False)

        Chem.GetSymmSSSR(mol)

        if mol is None:
            return None

        if addHs:
            mol = Chem.AddHs(mol)

        return mol

    @cache
    def get_from_SMILES(
        self, SMILES: str, addHs: Optional[bool] = False
    ) -> Optional[Chem.rdchem.Mol]:
        """
        Create an RDKit molecule object from a SMILES string.

        Attempts to create an RDKit molecule object from a given SMILES string,
        with options for handling multiple parts and hydrogens.

        Args:
            SMILES (str): The SMILES string to convert.
            addHs (bool): If True, explicitly adds hydrogen atoms to the molecule. Defaults to False.

        Returns:
            Optional[Chem.rdchem.Mol]: An RDKit molecule object if conversion is successful, None otherwise.

        Notes:
            - Returns None if the input SMILES is None.
            - Handles multi-part SMILES strings (separated by '.') by combining them into a single molecule.
            - For single-part SMILES, attempts to create the molecule with sanitization first.
            - If sanitization fails, attempts creation without sanitization for certain elements.
            - Updates the property cache of the molecule, with fallback to non-strict update.
            - Calculates the Smallest Set of Smallest Rings (SSSR) for the molecule.
            - Optionally adds explicit hydrogens based on the `addHs` parameter.
            - This method is cached for performance optimization.
        """
        if SMILES is None:
            return None
        SMILES_parts = SMILES.split(".")

        if len(SMILES_parts) > 1:
            mol = Chem.Mol()
            for SMILES in SMILES_parts:
                mol2 = self.get_from_SMILES(SMILES, addHs)
                if mol2 is None:
                    return None
                mol = Chem.CombineMols(mol, mol2)
            try:
                mol.UpdatePropertyCache()
            except Exception:
                mol.UpdatePropertyCache(strict=False)

            Chem.GetSymmSSSR(mol)
            return mol

        mol = Chem.MolFromSmiles(SMILES)
        if mol is None:
            # for some correct SMILES the sanitization process of rdkit does not work properly
            if (
                SMILES.count("F") > 0
                or SMILES.count("Cl") > 0
                or SMILES.count("Br") > 0
                or SMILES.count("B") > 0
            ):
                mol = Chem.MolFromSmiles(SMILES, sanitize=False)
                if mol is None:
                    return None
                mol.UpdatePropertyCache(strict=False)
        else:
            mol.UpdatePropertyCache()

        if mol is None:
            return None

        if addHs:
            mol = Chem.AddHs(mol)

        Chem.GetSymmSSSR(mol)
        return mol

    def to_SMILES(
        self,
        mol: Chem.rdchem.Mol,
        isomeric: Optional[bool] = True,
        canonicalize_tautomer: Optional[bool] = False,
        remove_isotopes: Optional[bool] = False,
    ) -> str:
        """
        Convert an RDKit molecule object to a SMILES string.

        Generates a SMILES string from an RDKit molecule object, with options
        for handling isomers, tautomers, and isotopes.

        Args:
            mol (Chem.rdchem.Mol): The RDKit molecule object to convert.
            isomeric (Optional[bool]): If True, includes isomeric information in the SMILES. Defaults to True.
            canonicalize_tautomer (Optional[bool]): If True, canonicalizes the tautomer before generating SMILES. Defaults to False.
            remove_isotopes (Optional[bool]): If True, removes isotope information before generating SMILES. Defaults to False.

        Returns:
            str: The SMILES string representation of the molecule.

        Notes:
            - If `remove_isotopes` is True, uses `remove_isotopes` method to strip isotope information.
            - If `canonicalize_tautomer` is True, standardizes the tautomeric form of the molecule.
            - Uses RDKit's `MolToSmiles` function to generate the final SMILES string.
        """

        if remove_isotopes:
            mol = self.remove_isotopes(mol)

        if canonicalize_tautomer:
            mol = rdMolStandardize.CanonicalTautomer(mol)

        return Chem.MolToSmiles(mol, isomericSmiles=isomeric)

    @cache
    def is_valid_SMILES_fast(self, SMILES: str) -> bool:
        """
        Quickly check if a string is a potentially valid SMILES representation.

        Performs a fast, preliminary check on a string to determine if it
        could be a valid SMILES representation using regular expressions.

        Args:
            SMILES (str): The string to check.

        Returns:
            bool: True if the string passes the preliminary SMILES validity check, False otherwise.

        Notes:
            - Returns False for None or non-string inputs.
            - Uses a regular expression for validation.
            - This is a fast check and may not catch all invalid SMILES strings.
            - The method is cached for performance optimization.
        """
        if not SMILES:
            return False

        if not isinstance(SMILES, str):
            return False

        if not regex.search(self.non_generic_SMILES_regex_compiled, SMILES):
            return False

        return True

    @cache
    def is_valid_SMILES(self, SMILES: str) -> bool:
        """
        Thoroughly check if a string is a valid SMILES representation.

        Performs a comprehensive check to determine if a string is a valid
        SMILES representation by attempting to parse it into a molecule.

        Args:
            SMILES (str): The string to check.

        Returns:
            bool: True if the string is a valid SMILES representation, False otherwise.

        Notes:
            - First calls `is_valid_SMILES_fast` for a quick preliminary check.
            - If the fast check passes, attempts to convert the SMILES to a molecule using `get_from_SMILES`.
            - Returns True only if both the fast check passes and the conversion to a molecule succeeds.
            - This method is more thorough but slower than `is_valid_SMILES_fast`.
            - The method is cached for performance optimization.
        """
        if not self.is_valid_SMILES_fast(SMILES):
            return False

        return self.get_from_SMILES(SMILES) is not None

    @cache
    def is_valid_InChI(self, InChI: str) -> bool:
        """
        Check if a string is a valid InChI representation.

        Determines if a string is a valid InChI representation by performing a format check
        and attempting to convert it to a molecule.

        Args:
            InChI (str): The string to check.

        Returns:
            bool: True if the string is a valid InChI representation, False otherwise.

        Notes:
            - Returns False for None or non-string inputs.
            - Checks if the string matches the expected InChI format using a regular expression.
            - If the format check passes, attempts to convert the InChI to a molecule using RDKit.
            - Returns True only if both the format check passes and the conversion to a molecule succeeds.
            - The method is cached for performance optimization.
        """
        if not InChI:
            return False

        if not isinstance(InChI, str):
            return False
        
        if not regex.search(self.InChI_regex_compiled, InChI):
            return False

        return Chem.MolFromInchi(InChI) is not None

    @cache
    def is_valid_InChIKey(self, InChIKey: str) -> bool:
        """
        Check if a string is a valid InChIKey representation.

        Determines if a string is a valid InChIKey representation by performing a format check.

        Args:
            InChIKey (str): The string to check.

        Returns:
            bool: True if the string is a valid InChIKeyy representation, False otherwise.

        Notes:
            - Returns False for None or non-string inputs.
            - Checks if the string matches the expected InChIKey format using a regular expression.
            - If the format check passes, returns True only if the format check passes.
            - The method is cached for performance optimization.
        """
        if not InChIKey:
            return False

        if not isinstance(InChIKey, str):
            return False
        
        if not regex.search(self.InChIKey_regex_compiled, InChIKey):
            return False

        return True

    @cache
    def _get_close_digits_on_keyboard(self, d: str) -> list[str]:
        """
        Returns digits that are physically adjacent to d
        on a QWERTY keyboard's top number row.

        Args:
            d (str): a digit

        Returns:
            list[str]: A list of close digits on the keyboard.
        """
        KEYBOARD_ADJ = {
            "0": ["9"],
            "1": ["2"],
            "2": ["1", "3"],
            "3": ["2", "4"],
            "4": ["3", "5"],
            "5": ["4", "6"],
            "6": ["5", "7"],
            "7": ["6", "8"],
            "8": ["7", "9"],
            "9": ["8", "0"],
        }
        return KEYBOARD_ADJ.get(d, [])

    @cache
    def is_valid_CAS(self, cas: Union[str, bool, None]) -> bool:
        """
        Check if a string is a valid CAS Registry Number.

        Validates whether a string represents a valid CAS Registry Number by
        checking its format and verifying its check digit.

        Args:
            cas (Union[str, bool, None]): The string to check.

        Returns:
            bool: True if the string is a valid CAS Registry Number, False otherwise.

        Notes:
            - Returns False for None, boolean inputs, or non-string inputs.
            - Checks if the string matches the standard CAS format using a regular expression.
            - Performs the CAS check digit validation algorithm.
            - The method is cached for performance optimization.

        Example:
            >>> is_valid_CAS("7732-18-5")
            True
            >>> is_valid_CAS("7732-18-6")
            False
        """
        if not cas:
            return False

        if not isinstance(cas, str):
            return False

        # Takes into account the standard CAS formatting e.g. 7732-18-5
        cas_match = regex.search(self.CAS_regex_with_groups, cas)
        if cas_match is None:
            return False
        cas_string = cas_match.group(1) + cas_match.group(2) + cas_match.group(3)

        increment = 0
        sum_cas = 0

        # Slices the reversed number string
        for number in reversed(cas_string):
            if increment == 0:
                validate = int(number)
            else:
                sum_cas = sum_cas + (int(number) * increment)

            increment = increment + 1

        # Does the math
        if validate == sum_cas % 10:
            return True
        else:
            return False

    @cache
    def expand_CAS_heuristically(self, CAS: str, max_swaps: int = 2) -> list[str]:
        """
        Perform a BFS over keyboard-adjacent digit swaps to find valid CAS numbers
        with the lowest amount of swaps, hopefully allowing to fix a CAS.

        Each single-digit swap must replace the digit with a QWERTY-adjacent digit
        from the top row (e.g., '3' can swap with '2' or '4'). The search will stop
        once the minimal number of swaps resulting in a valid CAS is found, or
        after exceeding `max_swaps`.

        Args:
            CAS (str): The input CAS string to be modified.
            max_swaps (int, optional): The maximum number of swaps (BFS depth) to
                explore. Defaults to 2.

        Returns:
            list[str]: A list of all valid CAS strings found at the minimal swap
            distance. Returns an empty list if none are found within `max_swaps`.
        """

        if self.is_valid_CAS(CAS):
            return [CAS]

        digit_positions = [i for i, ch in enumerate(CAS) if ch.isdigit()]

        queue = collections.deque([(CAS, 0)])
        visited = {CAS}

        valid_solutions = []
        found_level = None

        while queue:
            current_str, dist = queue.popleft()

            if found_level is not None and dist > found_level:
                break

            if dist >= max_swaps:
                continue

            for pos in digit_positions:
                orig_digit = current_str[pos]
                neighbors = self._get_close_digits_on_keyboard(orig_digit)

                for nd in neighbors:
                    new_str = current_str[:pos] + nd + current_str[pos + 1 :]

                    if new_str in visited:
                        continue
                    visited.add(new_str)

                    new_dist = dist + 1

                    if self.is_valid_CAS(new_str):
                        if found_level is None:
                            found_level = new_dist
                        if new_dist == found_level:
                            valid_solutions.append(new_str)

                    if found_level is None or new_dist <= found_level:
                        queue.append((new_str, new_dist))

        return valid_solutions

    @cache
    def are_InChIs_equal(
        self,
        InChI1: str,
        InChI2: str,
        differentiate_isomers: Optional[bool] = None,
        differentiate_tautomers: Optional[bool] = None,
        differentiate_isotopes: Optional[bool] = None,
        check_for_resonance_structures: Optional[bool] = None,
    ) -> bool:
        """
        Compare two InChI strings for equality based on specified criteria.

        Determines if two InChI strings represent the same chemical entity,
        with options to consider or ignore various structural features.

        Args:
            InChI1 (str): The first InChI string to compare.
            InChI2 (str): The second InChI string to compare.
            differentiate_isomers (Optional[bool]): If True, considers isomers as different. Defaults to None.
            differentiate_tautomers (Optional[bool]): If True, considers tautomers as different. Defaults to None.
            differentiate_isotopes (Optional[bool]): If True, considers isotopes as different. Defaults to None.
            check_for_resonance_structures (Optional[bool]): If True, checks for resonance structures. Defaults to None.

        Returns:
            bool: True if the InChIs are considered equal based on the specified criteria, False otherwise.

        Notes:
            - Returns False if either InChI string is empty or None.
            - If the InChI strings are identical, returns True immediately.
            - Converts InChI strings to RDKit molecule objects for comparison.
            - Uses the `are_equal` method for detailed comparison based on the specified criteria.
            - This method is cached for performance optimization.
        """
        if not InChI1 or not InChI2:
            return False

        if InChI1 == InChI2:
            return True

        mol1 = self.get_from_InChI(InChI1)
        mol2 = self.get_from_InChI(InChI2)
        return self.are_equal(
            mol1,
            mol2,
            differentiate_isomers,
            differentiate_tautomers,
            differentiate_isotopes,
            check_for_resonance_structures,
        )

    @cache
    def are_SMILES_equal(
        self,
        smiles1: str,
        smiles2: str,
        differentiate_isomers: Optional[bool] = None,
        differentiate_tautomers: Optional[bool] = None,
        differentiate_isotopes: Optional[bool] = None,
        check_for_resonance_structures: Optional[bool] = None,
    ) -> bool:
        """
        Compare two SMILES strings for equality based on specified criteria.

        Determines if two SMILES strings represent the same chemical entity,
        with options to consider or ignore various structural features.

        Args:
            smiles1 (str): The first SMILES string to compare.
            smiles2 (str): The second SMILES string to compare.
            differentiate_isomers (Optional[bool]): If True, considers isomers as different. Defaults to None.
            differentiate_tautomers (Optional[bool]): If True, considers tautomers as different. Defaults to None.
            differentiate_isotopes (Optional[bool]): If True, considers isotopes as different. Defaults to None.
            check_for_resonance_structures (Optional[bool]): If True, checks for resonance structures. Defaults to None.

        Returns:
            bool: True if the SMILES are considered equal based on the specified criteria, False otherwise.

        Notes:
            - Converts SMILES strings to RDKit molecule objects for comparison.
            - Uses the `are_equal` method for detailed comparison based on the specified criteria.
            - This method is cached for performance optimization.
        """

        return self.are_equal(
            self.get_from_SMILES(smiles1),
            self.get_from_SMILES(smiles2),
            differentiate_isomers,
            differentiate_tautomers,
            differentiate_isotopes,
            check_for_resonance_structures,
        )

    @cache
    def get_resonance_SMILES(self, SMILES: str) -> list[str]:
        """
        Generate resonance structures for a given molecule represented by a SMILES string.

        Takes a SMILES string and returns a list of SMILES strings representing all possible resonance structures of the molecule.

        Args:
            SMILES (str): The input SMILES string representing the molecule.

        Returns:
            list[str]: A list of SMILES strings representing the resonance structures.
            If no resonance structures are found or if the input is invalid,
            returns a list containing only the input SMILES.

        Notes:
            - Uses RDKit's `ResonanceMolSupplier` to generate resonance structures.
            - Implements a workaround for a known RDKit issue by setting an empty progress callback.
            - Handles corner cases where `ResonanceMolSupplier` might return `None` for all resonance structures.
            - Filters out any `None` results from the resonance structure generation.
            - This method is cached for performance optimization.

        Example:
            >>> resolver = MoleculeResolver()
            >>> resolver.get_resonance_SMILES("C=C-C=C")
            ['C=C-C=C', 'C=C=C-C', 'C-C=C=C']
        """
        mol = self.get_from_SMILES(SMILES)
        rms = Chem.ResonanceMolSupplier(mol)
        # the following line is a workaroung for https://github.com/rdkit/rdkit/issues/6704
        rms.SetProgressCallback(EmptyResonanceMolSupplierCallback())
        resonance_mols = list(rms)

        # This workaround is needed in very few cornercases where the
        # resonance mol suppler returns molecules that are None
        # e.g. tetrabromocobaltate(II) [Co+2]([Br-])([Br-])([Br-])([Br-])
        if all([resonance_mol is None for resonance_mol in resonance_mols]):
            return [SMILES]

        return [Chem.MolToSmiles(mol_) for mol_ in resonance_mols if mol_]

    def are_equal(
        self,
        mol1: Chem.rdchem.Mol,
        mol2: Chem.rdchem.Mol,
        differentiate_isomers: Optional[bool] = None,
        differentiate_tautomers: Optional[bool] = None,
        differentiate_isotopes: Optional[bool] = None,
        check_for_resonance_structures: Optional[bool] = None,
    ) -> bool:
        """
        Compare two RDKit molecule objects for equality based on specified criteria.

        Determines if two molecules are considered equal, with options to consider or ignore various structural features such as isomers, tautomers, isotopes, and resonance structures.

        Args:
            mol1: The first RDKit molecule object to compare.

            mol2: The second RDKit molecule object to compare.

            differentiate_isomers: If `True`, considers isomers as different.
            If `None`, uses the class's default setting. Defaults to `None`.

            differentiate_tautomers: If `True`, considers tautomers as different.
            If `None`, uses the class's default setting. Defaults to `None`.

            differentiate_isotopes: If `True`, considers isotopes as different.
            If `None`, uses the class's default setting. Defaults to `None`.

            check_for_resonance_structures: If `True`, checks for resonance structures.
            If `None`, uses the class's default setting. Defaults to `None`.

        Returns:
            `True` if the molecules are considered equal based on the specified criteria, `False` otherwise.

        Notes:
            - Returns `False` if either molecule is `None`.
            - Uses class default settings for unspecified comparison criteria.
            - Performs a series of checks based on the specified criteria:
              1. Compares InChI strings if all features are differentiated.
              2. Compares canonical SMILES if isotopes are not differentiated.
              3. Compares InChI strings without isotope layers if tautomers and isomers are differentiated.
              4. Compares InChI strings without stereo layers if only tautomers are differentiated.
              5. Compares InChI strings without stereo and tautomer layers if neither are differentiated.
            - Optionally checks for resonance structures if specified.
            - Uses RDKit's `MolToInchi` and `MolToSmiles` functions for conversions.
            - Handles potential RDKit exceptions during InChI generation.

        Example:
            >>> mol1 = Chem.MolFromSmiles("CC(=O)O")
            >>> mol2 = Chem.MolFromSmiles("CC(O)=O")
            >>> resolver = MoleculeResolver()
            >>> resolver.are_equal(mol1, mol2, differentiate_tautomers=False)
            True
            >>> resolver.are_equal(mol1, mol2, differentiate_tautomers=True)
            False
        """

        if mol1 is None or mol2 is None:
            return False

        if differentiate_isomers is None:
            differentiate_isomers = self._differentiate_isomers

        if differentiate_tautomers is None:
            differentiate_tautomers = self._differentiate_tautomers

        if differentiate_isotopes is None:
            differentiate_isotopes = self._differentiate_isotopes

        if check_for_resonance_structures is None:
            check_for_resonance_structures = self._check_for_resonance_structures

        if not differentiate_isotopes:
            mol1 = self.remove_isotopes(mol1)
            mol2 = self.remove_isotopes(mol2)

        mol1 = self.standardize_mol(mol1)
        mol2 = self.standardize_mol(mol2)

        if not differentiate_tautomers:
            mol1 = rdMolStandardize.CanonicalTautomer(mol1)
            mol2 = rdMolStandardize.CanonicalTautomer(mol2)

        SMILES1 = self.to_SMILES(
            mol1,
            differentiate_isomers,
            not differentiate_tautomers,
            not differentiate_tautomers,
        )
        SMILES2 = self.to_SMILES(
            mol2,
            differentiate_isomers,
            not differentiate_tautomers,
            not differentiate_tautomers,
        )

        if SMILES1 == SMILES2:
            return True

        SMILES1_structure_type = self.get_structure_type_from_SMILES(SMILES1)
        SMILES2_structure_type = self.get_structure_type_from_SMILES(SMILES2)
        if SMILES1_structure_type != SMILES2_structure_type:
            return False

        if check_for_resonance_structures is None:
            check_for_resonance_structures = SMILES1_structure_type in [
                "ion",
                "salt",
            ] or SMILES2_structure_type in ["ion", "salt"]

        def do_resonance_structures_overlap(SMILES1, SMILES2):
            unique_partial_SMILES1 = set(SMILES1.split("."))
            unique_partial_SMILES2 = set(SMILES2.split("."))

            if len(unique_partial_SMILES1) != len(unique_partial_SMILES2):
                return False

            matching_unique_partial_SMILES2_found = []
            for unique_partial_SMILES1_ in unique_partial_SMILES1:
                resonance_SMILES1 = self.get_resonance_SMILES(unique_partial_SMILES1_)
                unique_partial_SMILES2 = unique_partial_SMILES2 - set(
                    matching_unique_partial_SMILES2_found
                )
                for unique_partial_SMILES2_ in unique_partial_SMILES2:
                    resonance_SMILES2 = self.get_resonance_SMILES(
                        unique_partial_SMILES2_
                    )
                    # one would expect the resonance structures to be equal
                    # but due to a bug sometimes you do not get all of them: https://github.com/rdkit/rdkit/issues/4491
                    # this will not be a workaround for all situations
                    # but is still better than just looking at equality
                    if set(resonance_SMILES1).intersection(set(resonance_SMILES2)):
                        matching_unique_partial_SMILES2_found.append(
                            unique_partial_SMILES2_
                        )
                        break

            return len(unique_partial_SMILES1) == len(
                matching_unique_partial_SMILES2_found
            )

        if check_for_resonance_structures:
            return do_resonance_structures_overlap(SMILES1, SMILES2)
        return False

        # https://github.com/rdkit/rdkit/discussions/4719
        # return rdMolHash.MolHash(standardized_molecule1, rdMolHash.HashFunction.HetAtomTautomer) == rdMolHash.MolHash(standardized_molecule1, rdMolHash.HashFunction.HetAtomTautomer)

        # if method == 2:  # using the RDKit RegistrationHash

        #     def get_mol_hash(mol, hashscheme):
        #         return RegistrationHash.GetMolHash(
        #             RegistrationHash.GetMolLayers(mol), hash_scheme=hashscheme
        #         )

        #     if differentiate_isomers:
        #         hashscheme = RegistrationHash.HashScheme.ALL_LAYERS
        #     else:
        #         hashscheme = RegistrationHash.HashScheme.STEREO_INSENSITIVE_LAYERS

        #     mol1_hash = get_mol_hash(mol1, hashscheme)
        #     mol2_hash = get_mol_hash(mol2, hashscheme)

        #     if mol1_hash != mol2_hash and check_for_resonance_structures:
        #         return do_resonance_structures_overlap(SMILES1, SMILES2)

        #     return mol1_hash == mol2_hash

    def find_duplicates_in_molecule_dictionary(
        self, molecules: dict[Any, Molecule], clustered_molecules: Optional[dict] = None
    ) -> Generator[tuple[Any, Molecule, Any, Molecule, bool], None, None]:
        """
        Find duplicate molecules within a dictionary of molecules.

        Identifies and yields pairs of duplicate molecules from the input dictionary. It can use a pre-clustered dictionary of molecules for efficiency.

        Args:
            molecules (dict[Any, Molecule]): A dictionary of molecules to search for duplicates.
            The keys can be of any type, and the values are `Molecule` objects.

            clustered_molecules (Optional[dict]): A pre-clustered dictionary of molecules,
            grouped by molecular formula. If `None`, the method will generate this clustering.
            Defaults to `None`.

        Returns:
            tuple[Any, Molecule, Any, Molecule, bool]: Pairs of duplicate molecules found in the input dictionary.

        Notes:
            - If `clustered_molecules` is not provided, it generates the clustering using
              `group_molecule_dictionary_by_formula`.
            - Uses the `intersect_molecule_dictionaries` method to find duplicates.
            - This method is particularly useful for efficiently identifying duplicate molecules in large datasets.

        Example:
            >>> molecules = {
            ...     "mol1": Molecule("CC"),
            ...     "mol2": Molecule("CCC"),
            ...     "mol3": Molecule("CC")  # Duplicate of mol1
            ... }
            >>> resolver = MoleculeResolver()
            >>> duplicates = list(resolver.find_duplicates_in_molecule_dictionary(molecules))
            >>> print(duplicates)
            [('mol1', Molecule(...), 'mol3', Molecule(...), True)]
        """
        if not clustered_molecules:
            clustered_molecules = self.group_molecule_dictionary_by_formula(molecules)

        yield from self.intersect_molecule_dictionaries(
            molecules,
            molecules,
            clustered_molecules,
            clustered_molecules,
            report_same_keys=False,
        )

    def group_molecule_dictionary_by_formula(
        self, molecules: dict[Any, Molecule]
    ) -> dict[str, dict[Any, tuple[Chem.rdchem.Mol, Molecule]]]:
        """
        Group a dictionary of molecules by their molecular formulas.

        Takes a dictionary of molecules and returns a new dictionary where molecules are grouped by their molecular formulas.

        Args:
            molecules (dict[Any, Molecule]): A dictionary of molecules to be grouped.
            The keys can be of any type, and the values are `Molecule` objects.

        Returns:
            dict[str, dict[Any, tuple[Chem.rdchem.Mol, Molecule]]]: A dictionary where:
            - Keys are molecular formulas (`str`).
            - Values are dictionaries where:

              - Keys are the original keys from the input dictionary.
              - Values are tuples containing: (an RDKit `Mol` object, the original `Molecule` object)

        Notes:
            - Uses the `get_from_SMILES` method to convert SMILES to RDKit `Mol` objects.
            - Calculates molecular formulas using RDKit's `CalcMolFormula` function.
            - This grouping is useful for efficient comparison and searching of molecules with the same formula.

        Example:
            >>> molecules = {
            ...     "mol1": Molecule("CC"),
            ...     "mol2": Molecule("CCC"),
            ...     "mol3": Molecule("CCO")
            ... }
            >>> resolver = MoleculeResolver()
            >>> grouped = resolver.group_molecule_dictionary_by_formula(molecules)
            >>> for formula, mol_dict in grouped.items():
            ...     print(f"{formula}: {list(mol_dict.keys())}")
            C2H6: ['mol1']
            C3H8: ['mol2']
            C2H6O: ['mol3']
        """
        temporary = {}
        for key, molecule in molecules.items():
            mol_molecule = self.get_from_SMILES(molecule.SMILES)
            mol_formula = rdMolDescriptors.CalcMolFormula(mol_molecule)
            if mol_formula not in temporary:
                temporary[mol_formula] = {}

            temporary[mol_formula][key] = (mol_molecule, molecule)

        return temporary

    def intersect_molecule_dictionaries(
        self,
        molecules: dict[Any, Molecule],
        other_molecules: dict[Any, Molecule],
        mode: Optional[str] = "SMILES",
        clustered_molecules: Optional[dict] = None,
        clustered_other_molecules: Optional[dict] = None,
        report_same_keys: Optional[bool] = True,
    ) -> Generator[tuple[Any, Molecule, Any, Molecule, bool], None, None]:
        """
        Intersect two dictionaries of molecules based on structural similarity.

        Compares molecules from two dictionaries and yields pairs of molecules that are structurally similar,
        along with their keys and isomer information.

        Args:
            molecules (dict[Any, Molecule]): First dictionary of molecules to compare.

            other_molecules (dict[Any, Molecule]): Second dictionary of molecules to compare.

            mode (Optional[str]): Comparison mode, either `"SMILES"` or `"inchi"`. Defaults to `"SMILES"`.

            clustered_molecules (Optional[dict]): Pre-clustered version of `molecules`.
            If `None`, clustering will be performed. Defaults to `None`.

            clustered_other_molecules (Optional[dict]): Pre-clustered version of `other_molecules`.
            If `None`, clustering will be performed. Defaults to `None`.

            report_same_keys (Optional[bool]): If `True`, report matches even if the keys are identical.
            Defaults to `True`.

        Returns:
            tuple[Any, Molecule, Any, Molecule, bool]: A tuple containing:

                - Key from the first dictionary.
                - Molecule from the first dictionary.
                - Key from the second dictionary.
                - Molecule from the second dictionary.
                - Boolean indicating if the molecules have the same isomer information.

        Raises:
            ValueError: If an unsupported mode is specified.

        Notes:
            - Molecules are first grouped by molecular formula for efficient comparison.
            - Supports comparison using either SMILES or InChI representation.
            - For SMILES mode, molecules are compared with and without considering isomer information.
            - For InChI mode, molecules are compared with and without considering stereochemistry.
            - Uses the `are_equal` and `are_InChIs_equal` methods for comparisons.
        """
        if not clustered_molecules:
            clustered_molecules = self.group_molecule_dictionary_by_formula(molecules)

        if not clustered_other_molecules:
            clustered_other_molecules = self.group_molecule_dictionary_by_formula(
                other_molecules
            )

        for mol_formula, this_formula_molecules in clustered_molecules.items():
            if mol_formula in clustered_other_molecules:
                for key, (mol_molecule, molecule) in this_formula_molecules.items():
                    for other_key, (
                        other_mol_molecule,
                        other_molecule,
                    ) in clustered_other_molecules[mol_formula].items():
                        are_equal = False
                        same_isomer_information = True
                        if mode == "SMILES":
                            if self.are_equal(mol_molecule, other_mol_molecule):
                                are_equal = True
                                same_isomer_information = True
                            elif self.are_equal(
                                mol_molecule,
                                other_mol_molecule,
                                differentiate_isomers=False,
                            ):
                                are_equal = True
                                same_isomer_information = False
                        elif mode == "inchi":
                            if self.are_InChIs_equal(
                                Chem.MolToInchi(mol_molecule),
                                Chem.MolToInchi(other_mol_molecule),
                            ):
                                are_equal = True
                                same_isomer_information = True
                            elif self.are_InChIs_equal(
                                Chem.MolToInchi(mol_molecule),
                                Chem.MolToInchi(other_mol_molecule),
                                isomeric=False,
                            ):
                                are_equal = True
                                same_isomer_information = False
                        else:
                            raise ValueError('Supported modes are ["SMILES", "inchi"].')

                        if are_equal:
                            if report_same_keys or key != other_key:
                                yield (
                                    key,
                                    molecule,
                                    other_key,
                                    other_molecule,
                                    same_isomer_information,
                                )

    def get_SMILES_from_Mol_format(
        self, *, molblock: Optional[str] = None, url: Optional[str] = None
    ) -> Optional[str]:
        """
        Convert a molecule from MOL format to SMILES representation.

        Takes either a MOL block string or a URL pointing to a MOL file, converts it to an RDKit molecule object,
        standardizes it, and returns the SMILES representation.

        Args:
            molblock (Optional[str]): A string containing the molecule information in MOL format.
            Defaults to `None`.

            url (Optional[str]): A URL pointing to a file containing the molecule information
            in MOL format. Defaults to `None`.

        Returns:
            Optional[str]: The SMILES representation of the molecule if conversion is successful,
            `None` otherwise.

        Raises:
            ValueError: If both `molblock` and `url` are `None`, or if both are provided.
            TypeError: If `molblock` or `url` is provided but is not a string.

        Notes:
            - Either `molblock` or `url` must be provided, but not both.
            - If a URL is provided, the method will attempt to fetch the MOL data from it.
            - Includes a fix for potential issues in the MOL block format.
            - If the initial conversion fails, it attempts to create the molecule without sanitization.
            - The resulting molecule is standardized before converting to SMILES.
            - Uses RDKit for molecule manipulation and conversion.
        """
        if molblock is None and url is None:
            raise ValueError("molblock and url cannot both be None.")
        if molblock:
            if url is not None:
                raise
            if not isinstance(molblock, str):
                raise TypeError("molblock must be a string.")

        if url:
            if molblock is not None:
                raise
            if not isinstance(url, str):
                raise TypeError("url must be a string.")
            molblock = self._resilient_request(url)
            if molblock is None:
                return None

        # https://github.com/rdkit/rdkit/issues/1361
        def fix_MolBlock(mol_block):
            parts = mol_block.split("\n")
            count_line = parts[3].replace("v2000", "V2000").replace("v3000", "V3000")

            change_was_made = True

            ending = ""
            if count_line[-1] == "\r":
                ending = "\r"
                count_line = count_line.rstrip()

            while change_was_made:
                change_was_made = False
                three_digit_blocks = [
                    count_line[i : i + 3] for i in range(0, len(count_line), 3)
                ]

                for i_block, three_digit_block in enumerate(three_digit_blocks):
                    if three_digit_block != "   " and three_digit_block[-1] == " ":
                        count_line = (
                            count_line[: i_block * 3] + " " + count_line[i_block * 3 :]
                        )
                        change_was_made = True
                        break

            parts[3] = count_line + ending

            return "\n".join(parts)

        molblock = fix_MolBlock(molblock)
        mol = Chem.MolFromMolBlock(molblock)
        if not mol:
            mol = Chem.MolFromMolBlock(molblock, sanitize=False)
        if not mol:
            return None

        return Chem.MolToSmiles(mol)

    def get_SMILES_from_image_file(
        self,
        image_path: str,
        engines_order: Optional[list[str]] = ["osra", "molvec", "imago"],
        mode: Optional[str] = "single",
    ) -> Union[str, list[str]]:
        """
        Extract SMILES representation from a chemical structure image file.

        Uses multiple optical structure recognition engines to convert a chemical structure image into a SMILES string.

        Args:
            image_path (str): The file path to the image containing the chemical structure.

            engines_order (Optional[list[str]]): The order in which to try different recognition engines.
            Defaults to `["osra", "molvec", "imago"]`.

            mode (Optional[str]): The extraction mode. Can be either `"single"` (return first successful result)
                or `"all"` (return results from all successful engines). Defaults to `"single"`.

        Returns:
            Union[str, list[str]]: The SMILES representation of the chemical structure.
            If mode is `"single"`, returns the first successful SMILES string.
            If mode is `"all"`, returns a list of all successful SMILES strings.

        Raises:
            ValueError: If an invalid mode is specified.

        Notes:
            - **This function will throw an error if the services are offline.**
            - Attempts to use the specified engines in the given order.
            - OSRA is typically the most effective engine and is set as the default first choice.
            - Uses the `molvec.ncats.io` API for structure recognition.
            - If an engine successfully recognizes the structure, the result is converted to SMILES.
            - In `"single"` mode, the method stops after the first successful recognition.
            - Final SMILES strings are standardized before being returned.
            - If no engines successfully recognize the structure, an empty string or list is returned.
        """
        raise ConnectionError(
            "Unfortunately, the service has been offline for some time."
        )
        # usually OSRA works best, this is why it is left as default
        if mode not in ["single", "all"]:
            raise ValueError("The modes single and all are allowed.")

        SMILES = []

        with open(image_path, "rb") as handle:
            try:
                response_text = self._resilient_request(
                    "https://molvec.ncats.io/all",
                    {"data": handle.read()},
                    request_type="post",
                )

                data = json.loads(response_text)
                for engine in engines_order:
                    if data[engine]["status"].lower() in ["success", "cached"]:
                        temp_SMILES = self.get_SMILES_from_Mol_format(
                            molblock=data[engine]["molfile"]
                        )
                        if temp_SMILES is not None:
                            SMILES.append(temp_SMILES)
                            if mode == "single":
                                break
            except Exception:
                pass

        if mode == "single" and len(SMILES) > 0:
            SMILES = SMILES[0]

        SMILES = self.standardize_SMILES(SMILES)

        return SMILES

    def show_mol_and_pause(
        self,
        mol: Chem.rdchem.Mol,
        name: Optional[str] = None,
        size: Optional[tuple[int, int]] = (1000, 1000),
    ) -> None:
        """
        Display a molecule image with additional information and pause execution.

        Generates an image of the molecule, adds title information including the molecule's name (if provided)
        and formal charge, and displays the image.

        Args:
            mol (Chem.rdchem.Mol): The RDKit molecule object to be displayed.

            name (Optional[str]): The name of the molecule to be displayed in the title.
            If `None`, only the charge is shown. Defaults to `None`.

            size (Optional[tuple[int, int]]): The size of the output image in pixels (width, height).
            Defaults to `(1000, 1000)`.

        Notes:
            - Adjusts drawing options based on the image size for optimal visualization.
            - The formal charge of the molecule is always displayed in the title.
            - The image is displayed using the default image viewer of the system.
            - Execution is paused after displaying the image (implicit in `img.show()`).
            - Uses RDKit's `Draw` module for molecule rendering.
            - The title is added to the image using the PIL (Python Imaging Library) module.
            - The font size for atom labels and other drawing options are scaled based on the image size.
        """
        scaling_size = min(size)
        Draw.DrawingOptions.atomLabelFontSize = int(50 * scaling_size / 1000)
        Draw.DrawingOptions.dotsPerAngstrom = int(300 * scaling_size / 1000)
        Draw.DrawingOptions.bondLineWidth = max(float(4.0 * scaling_size / 1000), 1.0)
        img = Draw.MolToImage(mol, size=size)

        title = "charge: " + str(Chem.rdmolops.GetFormalCharge(mol))
        if name is not None:
            title = "name: " + name + "\n" + title

        draw = ImageDraw.Draw(img)
        s = int(size[1] / 30)
        fnt = ImageFont.truetype(font="arial.ttf", size=s)
        draw.multiline_text((10, 10), title, font=fnt, fill=(0, 0, 0, 255))

        img.show()

    def save_mol_to_PNG(
        self,
        mol: Chem.rdchem.Mol,
        filename: str,
        atom_infos: Optional[list] = None,
        atom_infos_format_string: Optional[str] = "%.3f",
        size: Optional[tuple[int, int]] = (1000, 1000),
    ) -> None:
        """
        Save a molecule image to a PNG file with optional atom information.

        Generates an image of the molecule, optionally adds atom-specific information, and saves it as a PNG file.

        Args:
            mol (Chem.rdchem.Mol): The RDKit molecule object to be saved.

            filename (str): The path and name of the file where the image will be saved.

            atom_infos (Optional[list]): A list of values to be displayed for each atom.
            If provided, must have the same length as the number of atoms in the molecule.
            Defaults to `None`.

            atom_infos_format_string (Optional[str]): The format string to use when converting
            `atom_infos` values to strings. Defaults to `"%.3f"`.

            size (Optional[tuple[int, int]]): The size of the output image in pixels (width, height).
            Defaults to `(1000, 1000)`.

        Notes:
            - If `atom_infos` is provided, each atom in the molecule will be annotated with the corresponding value.
            - Creates a deep copy of the molecule to avoid modifying the original.
            - Atom annotations are added as `'atomNote'` properties to each atom.
            - The image is saved in PNG format using RDKit's `MolToFile` function.
            - Uses RDKit's `Draw` module for molecule rendering.
            - If `atom_infos` is provided but doesn't match the number of atoms, it may lead to unexpected results.
        """
        new_mol = copy.deepcopy(mol)

        for i, atom in enumerate(new_mol.GetAtoms()):
            label = atom_infos_format_string % atom_infos[i]
            atom.SetProp("atomNote", label)

        Draw.MolToFile(new_mol, filename, size=size)

    def normalize_html(self, html_code: str) -> str:
        """
        Normalize HTML content.

        This method is a placeholder for HTML normalization logic.

        Args:
            html_code (str): The HTML content to normalize.

        Returns:
            str: The normalized HTML content.

        Notes:
            - Currently not implemented.
        """
        html_code = regex.sub("\s+", " ", html_code)
        html_code = html_code.replace("&nbsp;", " ")
        html_code = regex.sub("\s+", " ", html_code)
        return html_code

    def parse_items_from_html(
        self,
        html_code: str,
        split_tag: str,
        properties_regex: list[tuple[str, str, list[int]]],
        property_indices_required: Optional[list[int]] = None,
    ) -> list[list]:
        """
        Parse and extract items from HTML content based on specified regex patterns.

        Splits the HTML content, applies regex patterns to extract properties, and returns a list of items that meet the specified requirements.

        Args:
            html_code (str): The HTML content to parse.

            split_tag (str): The HTML tag used to split the content into parts.
            If empty, the entire HTML is treated as one part.

            properties_regex (list[tuple[str, str, list[int]]]): A list of tuples, each containing:

                - `source_type` (str): `'html'` or `'text'` to indicate where to apply the regex.
                - `property_regex` (str): The regex pattern to extract the property.
                - `allowed_match_number` (list[int]): A list of allowed numbers of matches for the regex.

            property_indices_required (Optional[list[int]]): Indices of properties that must be non-`None` for an item to be included in the result. Defaults to `None`.

        Returns:
            list[list]: A list of items, where each item is a list of extracted properties.

        Raises:
            RuntimeError: If the number of regex matches doesn't fall within the allowed range for any property.

        Notes:
            - First normalizes the HTML using the `normalize_html` method.

            - Splits the HTML based on the provided `split_tag`.

            - For each part, applies the regex patterns to extract properties.

            - Items are only included in the result if they meet the requirements specified by
              `property_indices_required`.

            - HTMLparser was thought of here, but IMHO it is way too complicated to match the correct
              opening and closing tags with the python std library. I did not want to depend on
              beautifulsoup at the beginning, but it is a dependency that I am willing to add in the
              future. The following code is a simple and quick solution that works for the time being.

        """
        html_code = self.normalize_html(html_code)

        if split_tag:
            html_parts = html_code.split(split_tag)[1:]
        else:
            html_parts = [html_code]

        all_items = []
        for html_part in html_parts:
            text_part = regex.sub(r"<.*?>", "", html_part)
            text_part = regex.sub(r"\s+", " ", text_part)
            item = []
            for source_type, property_regex, allowed_match_number in properties_regex:
                source = html_part if source_type == "html" else text_part
                property_matches = regex.findall(property_regex, source)
                if allowed_match_number:
                    if len(property_matches) not in allowed_match_number:
                        raise RuntimeError(
                            "The webpage either changed its format or the regex is picking up something unwanted."
                        )
                value = property_matches[0].strip() if property_matches else None
                item.append(value)

            add_item = False
            if property_indices_required:
                add_item = all(
                    [
                        item[property_index] is not None
                        for property_index in property_indices_required
                    ]
                )

            if add_item:
                all_items.append(item)

        return all_items

    def get_java_path(self) -> Optional[str]:
        """
        Get the full path of the Java executable.

        Determines the full path of the Java executable on the system, supporting Windows, Linux, and macOS platforms.

        Returns:
            Optional[str]: The full path to the Java executable if found, `None` otherwise.

        Raises:
            NotImplementedError: If the method is called on an unsupported operating system.

        Notes:
            - On Windows, uses the `'where'` command to locate Java.
            - On Linux and macOS, uses the `'which'` command.
            - Checks if the found Java paths actually exist on the file system.
            - If multiple Java installations are found, the first valid path is returned.
            - Returns `None` if Java is not found or if the `'where'`/`'which'` command fails.
        """
        if platform.system().lower() == "windows":
            search_command = "where"
        elif platform.system().lower() in ["linux", "darwin"]:
            search_command = "which"
        else:
            raise NotImplementedError(
                f"For the following OS, getting the full path of the java executable needs to be programmed: {platform.system()}"
            )

        output = subprocess.run([search_command, "java"], capture_output=True)
        java_paths = output.stdout.decode("utf-8").strip().split(os.linesep)
        java_paths = [
            java_path for java_path in java_paths if os.path.exists(java_path)
        ]

        if output.returncode != 0 or len(java_paths) == 0:
            return None
        else:
            return java_paths[0]

    def get_molecule_from_OPSIN(
        self,
        name: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
        allow_warnings: Optional[bool] = False,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from OPSIN based on its name.

        Queries the OPSIN (Open Parser for Systematic IUPAC Nomenclature) service
        to convert a chemical name into a molecular structure, with optional constraints.

        Args:
            name (str): The chemical name to be converted to a molecular structure.
            required_formula (Optional[str]): The expected molecular formula.
            required_charge (Optional[int]): The expected molecular charge.
            required_structure_type (Optional[str]): The expected structure type.
            allow_warnings (Optional[bool]): If True, accept results with warnings.

        Returns:
            Optional[Molecule]: A Molecule object if successful, None otherwise.

        Notes:
            - Checks the molecule cache before querying OPSIN.
            - Uses a resilient request method to handle potential network issues.
            - The SMILES string returned by OPSIN is checked against the required parameters.
            - If the SMILES check fails, it attempts to convert the InChI to SMILES.
            - Standardizes the SMILES string before creating the Molecule object.
            - Any warnings from OPSIN are included in the Molecule's additional_information.
            - The result is filtered and combined with other results before being returned.
            - If no valid molecule is found, the method returns None.
        """
        SMILES = None
        additional_information = ""
        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            services="opsin",
        )

        with self.query_molecule_cache("opsin", "name", name) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                response_text = self._resilient_request(
                    f"https://opsin.ch.cam.ac.uk/opsin/{urllib.parse.quote(name)}"
                )

                if response_text is not None:
                    temp = json.loads(response_text)
                    if temp["status"] == "SUCCESS" or (
                        allow_warnings and temp["status"] == "WARNING"
                    ):
                        if self.check_SMILES(
                            temp["smiles"],
                            required_formula,
                            required_charge,
                            required_structure_type,
                        ):
                            SMILES = temp["smiles"]
                        else:
                            SMILES_from_InChI = self.InChI_to_SMILES(temp["inchi"])
                            if self.check_SMILES(
                                SMILES_from_InChI,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            ):
                                SMILES = SMILES_from_InChI

                        additional_information = ""
                        if "warnings" in temp:
                            additional_information = "WARNINGS: " + ", ".join(
                                temp["warnings"]
                            )

                        if SMILES:
                            molecules.append(
                                Molecule(
                                    SMILES,
                                    service="opsin",
                                    mode="name",
                                    additional_information=additional_information,
                                )
                            )

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_molecule_from_OPSIN_batchmode(
        self, names: list[str], allow_uninterpretable_stereo: Optional[bool] = False
    ) -> list[Optional[Molecule]]:
        """
        Convert a batch of chemical names to molecules using OPSIN in offline mode.

        Uses OPSIN to convert a list of chemical names to molecular structures in batch mode.

        Args:
            names (list[str]): A list of chemical names to be converted.
            allow_uninterpretable_stereo (Optional[bool]): Allows OPSIN to ignore uninterpretable stereochemistry.

        Returns:
            list[Optional[Molecule]]: A list of Molecule objects corresponding to the input names.

        Raises:
            FileNotFoundError: If the Java installation could not be found.
            RuntimeError: If there was a problem parsing the OPSIN offline output file.

        Notes:
            - Checks a cache for previously processed molecules.
            - Downloads the latest version of OPSIN if not already present.
            - Runs OPSIN in offline mode using Java, processing all names in a single batch.
            - Standardizes the SMILES strings returned by OPSIN.
            - Each successfully converted molecule is stored with metadata.
            - Uses a temporary directory for OPSIN operations, cleaned up after use.
            - If allow_uninterpretable_stereo is True, it's noted in the molecule's additional_information.
        """
        self._check_parameters(
            identifiers=names,
            modes="name",
            services="opsin",
            context="get_molecules_batch",
        )

        if not self._java_path:
            raise FileNotFoundError(
                "The java installation could not be found. Either it is not installed or its location has not been added to the path environment variable."
            )

        def download_OPSIN_offline_and_convert_names(names: list[str], tempfolder: str):
            if not os.path.exists(os.path.join(tempfolder, "opsin.jar")):
                response_text = self._resilient_request(
                    "https://api.github.com/repos/dan2097/opsin/releases/latest"
                )
                temp = json.loads(response_text)
                for asset in temp["assets"]:
                    if asset["name"].count("opsin-cli"):
                        download_url = asset["browser_download_url"]
                        with closing(urllib.request.urlopen(download_url)) as r:
                            with open(os.path.join(tempfolder, "opsin.jar"), "wb") as f:
                                for chunk in r:
                                    f.write(chunk)

            unique_id = str(uuid.uuid4())  # needed for multiple runs in parallel
            input_file = os.path.join(tempfolder, f"input_{unique_id}.txt")
            with open(input_file, "w", encoding="utf8") as f:
                f.write("\n".join(names))

            # in the future adding the parameter -s would allow getting more structures if the stereo information opsin cannot interpret can be ignored
            cmd = [
                self._java_path,
                "-jar",
                "opsin.jar",
                "-osmi",
                f"input_{unique_id}.txt",
                f"output_{unique_id}.txt",
            ]

            if allow_uninterpretable_stereo:
                cmd.insert(4, "--allowUninterpretableStereo")

            _ = subprocess.run(
                cmd,
                capture_output=True,
                cwd=tempfolder,
            )

            with open(os.path.join(tempfolder, f"output_{unique_id}.txt"), "r") as f:
                output = f.read()

            output_values = output.split("\n")
            # I don't know in what situations, but in some OPSIN adds an empty line at the end of the output file.
            if len(output_values) == len(names) + 1 and output_values[-1] == "":
                output_values = output_values[:-1]

            SMILES = [value.strip() for value in output_values]
            if len(SMILES) != len(names):
                raise RuntimeError(
                    "There was a problem parsing the OPSIN offline output file."
                )

            return SMILES

        with self.query_molecule_cache_batchmode("opsin", "name", names) as (
            identifiers_to_search,
            indices_of_identifiers_to_search,
            results,
        ):
            if len(identifiers_to_search) == 0:
                return results

            if self._OPSIN_tempfolder is None or not os.path.exists(
                self._OPSIN_tempfolder.name
            ):
                with tempfile.TemporaryDirectory() as tempfolder:
                    SMILES = download_OPSIN_offline_and_convert_names(
                        identifiers_to_search, tempfolder
                    )
            else:
                SMILES = download_OPSIN_offline_and_convert_names(
                    identifiers_to_search, self._OPSIN_tempfolder.name
                )

            for molecule_index, name, smi in zip(
                indices_of_identifiers_to_search,
                identifiers_to_search,
                SMILES,
                strict=True,
            ):
                if smi:
                    results[molecule_index] = [
                        Molecule(
                            SMILES=smi,
                            synonyms=[name],
                            mode="name",
                            service="opsin",
                            identifier=name,
                            additional_information=(
                                "allow_uninterpretable_stereo"
                                if allow_uninterpretable_stereo
                                else ""
                            ),
                        )
                    ]

            return results

    @cache
    def get_molecule_for_ion_from_partial_pubchem_search(
        self,
        identifier: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
    ) -> Optional[list[tuple[Molecule, int]]]:
        """
        Search PubChem for an ion molecule based on a partial identifier.

        Performs a partial search on PubChem using the given identifier and returns matching ion molecules
        that satisfy the specified formula and charge requirements.

        Args:
            identifier (str): The partial identifier to search for.
            required_formula (Optional[str]): The expected molecular formula.
            required_charge (Optional[int]): The expected molecular charge.

        Returns:
            Optional[list[tuple[Molecule, int]]]: A list of tuples with Molecule objects and their occurrence counts.

        Raises:
            ValueError: If the parameters fail validation in the _check_parameters method.

        Notes:
            - Cached to improve performance for repeated calls.
            - Uses a resilient request method to handle potential network issues.
            - The search is performed on the PubChem Compound database.
            - For each compound found, checks if it's a salt or mixture.
            - Each component is checked against the required formula and charge.
            - Standardizes the SMILES strings of matching molecules.
            - Results are sorted by frequency.
            - Returns None if no matching molecules are found.
        """
        required_structure_type = "ion"
        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
        )

        search_response_text = self._resilient_request(
            f'https://www.ncbi.nlm.nih.gov/pccompound/?term={urllib.parse.quote(identifier, safe="")}'
        )
        found_SMILES = []
        if search_response_text is not None:
            items = self.parse_items_from_html(
                search_response_text,
                '<div class="rprt">',
                [("text", "CID: (\d+)", [1])],
                [0],
            )
            for (cid,) in items:
                molecule = self.get_molecule_from_pubchem(cid, "cid")
                is_salt_or_mixture = molecule.SMILES.count(".") > 0

                if is_salt_or_mixture:
                    SMILES_list = molecule.SMILES.split(".")
                    for temptative_SMILES in SMILES_list:
                        temp_mol = self.get_from_SMILES(temptative_SMILES)
                        if self.check_formula(
                            temp_mol, required_formula
                        ) and self.check_charge(temp_mol, required_charge):
                            SMILES = Chem.MolToSmiles(temp_mol)
                            SMILES = self.standardize_SMILES(SMILES)
                            found_SMILES.append(SMILES)

        if len(found_SMILES) == 0:
            return None

        # sort to most likely candidate
        counts = collections.Counter(found_SMILES)

        return [
            (Molecule(smi, service="pubchem"), y) for smi, y in counts.most_common()
        ]

    def get_molecule_from_ChEBI_old(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from ChEBI based on the provided identifier and mode.

        Queries the ChEBI database to retrieve molecule information based on various identifiers.

        Args:
            identifier (str): The identifier to search for.
            mode (str): The type of identifier (e.g., 'name', 'cas', 'formula', 'smiles', 'inchi', 'inchikey').
            required_formula (Optional[str]): The expected molecular formula.
            required_charge (Optional[int]): The expected molecular charge.
            required_structure_type (Optional[str]): The expected structure type.

        Returns:
            Optional[Molecule]: A Molecule object if found and meets all requirements, None otherwise.

        Raises:
            ValueError: If the input parameters fail validation.

        Notes:
            - Checks a cache for previously retrieved molecules.
            - Uses the ChEBI web services API to query the database.
            - Retrieves up to 5 results for formula searches; otherwise, retrieves 1 result.
            - Extracts SMILES, synonyms, CAS numbers, and ChEBI ID.
            - Filters results based on required formula, charge, and structure type.
            - Combines multiple matching molecules into a single result if necessary.
            - Uses resilient network requests to handle potential connection issues.
        """
        if required_formula is None:
            if mode == "formula":
                required_formula = identifier

        self._check_parameters(
            services="chebi",
            modes=mode,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            required_formulas=required_formula,
        )

        with self.query_molecule_cache("chebi", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                CHEBI_URL = "https://www.ebi.ac.uk/webservices/chebi/2.0/test/"
                mode_mapping = {
                    "name": "ALL+NAMES",
                    "cas": "REGISTRY+NUMBERS",
                    "formula": "FORMULA",
                    "smiles": "SMILES",
                    "inchi": "INCHI%2FINCHI+KEY",
                    "inchikey": "INCHI%2FINCHI+KEY",
                }
                maximumResults = 5 if mode == "formula" else 1
                search_response_text = self._resilient_request(
                    f'{CHEBI_URL}getLiteEntity?search={urllib.parse.quote(identifier, safe="")}&searchCategory={mode_mapping[mode]}&maximumResults={maximumResults}&starsCategory=ALL'
                )

                SMILES = None
                synonyms = []
                CAS = []
                ChEBI_id = None

                if search_response_text is not None:
                    root = xmltodict.parse(search_response_text)
                    if "getLiteEntityResponse" not in root["S:Envelope"]["S:Body"]:
                        return None

                    temp = root["S:Envelope"]["S:Body"]["getLiteEntityResponse"][
                        "return"
                    ]
                    if not temp:
                        return None

                    temp_list = temp["ListElement"]
                    if not isinstance(temp["ListElement"], list):
                        temp_list = [temp["ListElement"]]

                    found_ChEBI_ids = [item["chebiId"] for item in temp_list]
                    for ChEBI_id in found_ChEBI_ids:
                        molecule_response_text = self._resilient_request(
                            f"{CHEBI_URL}getCompleteEntity?chebiId={ChEBI_id}"
                        )

                        if molecule_response_text is not None:
                            root = xmltodict.parse(molecule_response_text)
                            molecule = root["S:Envelope"]["S:Body"][
                                "getCompleteEntityResponse"
                            ]["return"]

                            if "smiles" not in molecule:
                                continue
                            SMILES = molecule["smiles"]

                            if "IupacNames" in molecule:
                                if isinstance(molecule["IupacNames"], list):
                                    synonyms.extend(
                                        [
                                            synonym["data"]
                                            for synonym in molecule["IupacNames"]
                                        ]
                                    )
                                elif isinstance(molecule["IupacNames"], dict):
                                    synonyms.append(molecule["IupacNames"]["data"])
                            if "Synonyms" in molecule:
                                temp = molecule["Synonyms"]
                                if not isinstance(molecule["Synonyms"], list):
                                    temp = [temp]
                                synonyms.extend([synonym["data"] for synonym in temp])

                            synonyms = self.filter_and_sort_synonyms(synonyms)

                            if "RegistryNumbers" in molecule:
                                temp = molecule["RegistryNumbers"]
                                if not isinstance(molecule["RegistryNumbers"], list):
                                    temp = [molecule["RegistryNumbers"]]
                                CAS = self.filter_and_sort_CAS(
                                    [synonym["data"] for synonym in temp]
                                )

                            molecule = Molecule(
                                SMILES,
                                synonyms,
                                CAS,
                                ChEBI_id.split(":")[1],
                                mode,
                                service="chebi",
                            )
                            molecules.append(molecule)

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_molecule_from_ChEBI(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from ChEBI based on the provided identifier and mode.

        Queries the ChEBI database to retrieve molecule information based on various identifiers.

        Args:
            identifier (str): The identifier to search for.
            mode (str): The type of identifier (e.g., 'name', 'cas', 'formula', 'smiles', 'inchi', 'inchikey').
            required_formula (Optional[str]): The expected molecular formula.
            required_charge (Optional[int]): The expected molecular charge.
            required_structure_type (Optional[str]): The expected structure type.

        Returns:
            Optional[Molecule]: A Molecule object if found and meets all requirements, None otherwise.
        """
        if required_formula is None:
            if mode == "formula":
                required_formula = identifier

        self._check_parameters(
            services="chebi",
            modes=mode,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            required_formulas=required_formula,
        )

        with self.query_molecule_cache("chebi", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                SOAP_ENDPOINT = (
                    "http://www.ebi.ac.uk:80/webservices/chebi/2.0/webservice"
                )
                mode_mapping = {
                    "name": "ALL NAMES",
                    "cas": "REGISTRY NUMBERS",
                    "formula": "FORMULA",
                    "smiles": "SMILES",
                    "inchi": "INCHI/INCHI KEY",
                    "inchikey": "INCHI/INCHI KEY",
                }
                maximumResults = 1 if mode == "formula" else 5

                # Construct SOAP request for getLiteEntity
                soap_request = f"""
                <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                                xmlns:chebi="https://www.ebi.ac.uk/webservices/chebi">
                <soapenv:Header/>
                <soapenv:Body>
                    <chebi:getLiteEntity>
                        <chebi:search>{identifier}</chebi:search>
                        <chebi:searchCategory>{mode_mapping[mode]}</chebi:searchCategory>
                        <chebi:maximumResults>{maximumResults}</chebi:maximumResults>
                        <chebi:stars>ALL</chebi:stars>
                    </chebi:getLiteEntity>
                </soapenv:Body>
                </soapenv:Envelope>
                """
                # chebi does not allow differentiating between inchi and inchicode
                # so I am adding the functionality here
                if mode == "inchi" and not self.InChI_regex_compiled.match(identifier):
                    return None

                if mode == "inchikey" and not self.InChIKey_regex_compiled.match(
                    identifier
                ):
                    return None

                search_response_text = self._resilient_request(
                    SOAP_ENDPOINT,
                    {
                        "data": soap_request,
                        "headers": {"Content-Type": "text/xml; charset=utf-8"},
                    },
                    request_type="post",
                    rejected_status_codes=[404, 500],
                )
                if search_response_text is None:
                    return None

                root = xmltodict.parse(search_response_text)
                if (
                    root["S:Envelope"]["S:Body"]["getLiteEntityResponse"]["return"]
                    is None
                ):
                    return None
                temp_list = root["S:Envelope"]["S:Body"]["getLiteEntityResponse"][
                    "return"
                ]["ListElement"]
                if not isinstance(temp_list, list):
                    temp_list = [temp_list]
                found_ChEBI_ids = [item["chebiId"] for item in temp_list]

                # Construct SOAP request for getCompleteEntityByList
                ids_xml = "".join(
                    f"<chebi:ListOfChEBIIds>{chebi_id}</chebi:ListOfChEBIIds>"
                    for chebi_id in found_ChEBI_ids
                )
                complete_entity_request = f"""
                <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                                xmlns:chebi="https://www.ebi.ac.uk/webservices/chebi">
                <soapenv:Header/>
                <soapenv:Body>
                    <chebi:getCompleteEntityByList>
                        {ids_xml}
                    </chebi:getCompleteEntityByList>
                </soapenv:Body>
                </soapenv:Envelope>
                """
                complete_response_text = self._resilient_request(
                    SOAP_ENDPOINT,
                    {"data": complete_entity_request},
                    request_type="post",
                )
                if complete_response_text is None:
                    return None

                root = xmltodict.parse(complete_response_text)
                try:
                    entities = root["S:Envelope"]["S:Body"][
                        "getCompleteEntityByListResponse"
                    ]["return"]
                    if not isinstance(entities, list):
                        entities = [entities]
                except KeyError:
                    return None

                SMILES = None
                for entity in entities:
                    if "smiles" in entity:
                        SMILES = entity["smiles"]
                    else:
                        if "ChemicalStructures" in entity:
                            if not isinstance(entity["ChemicalStructures"], list):
                                entity["ChemicalStructures"] = [
                                    entity["ChemicalStructures"]
                                ]
                            for structure in entity["ChemicalStructures"]:
                                if structure["type"] == "mol":
                                    mol = Chem.MolFromMolBlock(structure["structure"])
                                    if mol:
                                        SMILES = Chem.MolToSmiles(mol)
                                        break

                    if not SMILES:
                        continue

                    synonyms = []
                    CAS = []

                    if "chebiAsciiName" in entity:
                        synonyms.append(entity["chebiAsciiName"])

                    if "IupacNames" in entity:
                        if not isinstance(entity["IupacNames"], list):
                            entity["IupacNames"] = [entity["IupacNames"]]
                        synonyms.extend(
                            [synonym["data"] for synonym in entity["IupacNames"]]
                        )

                    if "Synonyms" in entity:
                        temp = entity["Synonyms"]
                        if not isinstance(temp, list):
                            temp = [temp]
                        synonyms.extend([synonym["data"] for synonym in temp])

                    # Check if the lowercase identifier matches any synonym
                    # chebi gives back lots of bad results with partial name matches
                    if mode == "name":
                        if identifier.lower() not in [
                            synonym.lower() for synonym in synonyms
                        ]:
                            continue

                    synonyms = self.filter_and_sort_synonyms(synonyms)

                    if "RegistryNumbers" in entity:
                        temp = entity["RegistryNumbers"]
                        if not isinstance(temp, list):
                            temp = [temp]
                        CAS = self.filter_and_sort_CAS(
                            [registry["data"] for registry in temp]
                        )

                    molecule = Molecule(
                        SMILES,
                        synonyms,
                        CAS,
                        entity["chebiId"].split(":")[1],
                        mode,
                        service="chebi",
                    )
                    molecules.append(molecule)

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_molecules_using_batchmode_from(
        self,
        identifiers: list[list[str]],
        modes: list[list[str]],
        service: str,
        batch_size: Optional[int] = 1000,
        progressbar: Optional[bool] = False,
    ) -> tuple[dict[str, list[Optional[Molecule]]], list[str]]:
        """
        Retrieve molecules in batch mode from a specified service.

        Performs batch retrieval of molecules from a supported service using provided identifiers and modes.

        Args:
            identifiers (list[list[str]]): A list of lists containing identifiers for molecules.
            modes (list[list[str]]): A list of lists containing modes corresponding to the identifiers.
            service (str): The name of the service to use for retrieval.
            batch_size (Optional[int]): Number of identifiers to process in each batch.
            progressbar (Optional[bool]): Display a progress bar during processing.

        Returns:
            tuple[dict[str, list[Optional[Molecule]]], list[str]]: A dictionary of results and a list of supported modes.

        Raises:
            ValueError: If the specified service does not have batch capabilities or is not supported.
            ValueError: If the list of identifiers and modes for a molecule are not of the same length.
            NotImplementedError: If batch functionality for the specified service is not implemented.
            RuntimeError: If the number of returned values doesn't match the number of unique values requested.

        Notes:
            - Checks if the service supports batch capabilities.
            - Groups identifiers by mode for efficient batch processing.
            - Uses different batch retrieval functions based on the service.
            - Supports services like PubChem, SRS, CompTox, and OPSIN.
            - Organizes results by mode.
        """
        if service not in self._available_services_with_batch_capabilities:
            raise ValueError(
                f"The service {service} does not have batch capabilities or is not supported."
            )

        all_supported_modes = []
        new_modes = []
        for _modes in modes:
            _modes = [mode.lower() for mode in _modes]
            new_modes.append(_modes)
            [
                all_supported_modes.append(mode)
                for mode in _modes
                if mode not in all_supported_modes
                and mode in self.supported_modes_by_services[service]
            ]

        modes = new_modes

        self._check_parameters(identifiers=identifiers, modes=modes, context="batch")

        identifier_sets = {}
        identifier_indices_sets = {}
        for mode in all_supported_modes:
            identifier_sets[mode] = []
            identifier_indices_sets[mode] = []

        for i_molecule, (i_molecule_identifiers, i_molecule_modes) in enumerate(
            zip(identifiers, modes, strict=True)
        ):
            if len(i_molecule_identifiers) != len(i_molecule_modes):
                raise ValueError(
                    "The list of identifiers and modes of a molecule must be of the same length."
                )
            (
                flattened_i_molecule_identifiers,
                flattened_i_molecule_modes,
                _,
                _,
                _,
            ) = self._check_and_flatten_identifiers_and_modes(
                i_molecule_identifiers, i_molecule_modes
            )
            for identifier, mode in zip(
                flattened_i_molecule_identifiers,
                flattened_i_molecule_modes,
                strict=True,
            ):
                if mode in all_supported_modes:
                    identifier_sets[mode].append(identifier)
                    identifier_indices_sets[mode].append(i_molecule)

        results = {}
        for mode in all_supported_modes:
            this_mode_results = {}

            _identifiers = identifier_sets[mode]
            _identifier_indices = identifier_indices_sets[mode]

            chunks = list(
                self.chunker(
                    list(zip(_identifiers, _identifier_indices, strict=True)),
                    batch_size,
                )
            )

            if chunks:
                for chunk in tqdm(chunks, disable=not progressbar):
                    identifier_indices_chunk = [val[1] for val in chunk]
                    identifier_chunk = [val[0] for val in chunk]

                    if service == "pubchem":
                        results_for_this_chunk = (
                            self.get_molecules_from_pubchem_batchmode(
                                identifier_chunk, mode
                            )
                        )
                    elif service == "srs":
                        results_for_this_chunk = self.get_molecules_from_SRS_batchmode(
                            identifier_chunk, mode
                        )
                    elif service == "comptox":
                        results_for_this_chunk = (
                            self.get_molecules_from_CompTox_batchmode(
                                identifier_chunk, mode
                            )
                        )
                    elif service == "opsin":
                        results_for_this_chunk = self.get_molecule_from_OPSIN_batchmode(
                            identifier_chunk, True
                        )
                    else:
                        raise NotImplementedError(
                            "The batch functionality for this service is either not implemented or the service does not support batch searches."
                        )

                    for index, result in zip(
                        identifier_indices_chunk, results_for_this_chunk, strict=True
                    ):
                        if index not in this_mode_results:
                            this_mode_results[index] = None

                        if result and this_mode_results[index] is None:
                            this_mode_results[index] = result

                    if (
                        len(
                            set(identifier_indices_chunk)
                            - set(this_mode_results.keys())
                        )
                        != 0
                    ):
                        raise RuntimeError(
                            "The number of returned values should be equal to the unique values asked for."
                        )

            results[mode] = []
            for index in range(len(identifiers)):
                if index in _identifier_indices:
                    if index in this_mode_results:
                        results[mode].append(this_mode_results[index])
                    else:
                        results[mode].append(None)

        return results, all_supported_modes

    def get_molecule_from_CompTox(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from CompTox based on the provided identifier and mode.

        Queries the CompTox database to retrieve molecule information based on various identifiers.

        Args:
            identifier (str): The identifier to search for.
            mode (str): The type of identifier.
            required_formula (Optional[str]): The expected molecular formula.
            required_charge (Optional[int]): The expected molecular charge.
            required_structure_type (Optional[str]): The expected structure type.

        Returns:
            Optional[Molecule]: A Molecule object if found and meets all requirements, None otherwise.

        Raises:
            ValueError: If the input parameters fail validation.

        Notes:
            - Checks a cache for previously retrieved molecules.
            - Uses the CompTox API to query the database.
            - Extracts SMILES, synonyms, CAS numbers, and DTXSID.
            - Filters results based on required formula, charge, and structure type.
            - Includes CompTox QC level in the additional information.
            - # new API, at some point we need to change to https://api-ccte.epa.gov/docs/chemical.html#/

        """

        self._check_parameters(
            modes=mode,
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            services="comptox",
        )

        with self.query_molecule_cache("comptox", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                COMPTOX_URL = "https://comptox.epa.gov/dashboard-api/"

                response_text = self._resilient_request(
                    f'{COMPTOX_URL}ccdapp1/search/chemical/equal/{urllib.parse.quote(identifier, safe="")}',
                    rejected_status_codes=[400, 404],
                )

                if response_text is not None:
                    results = json.loads(response_text)
                    # sort and filter best results
                    results = sorted(results, key=lambda x: x["rank"])
                    results = list(
                        filter(
                            lambda x: x["rank"] == results[0]["rank"],
                            results,
                        )
                    )

                    def process_dtxsid(temp_dtxsid):
                        substance_response_text = self._resilient_request(
                            f"{COMPTOX_URL}ccdapp2/chemical-detail/search/by-dsstoxsid/?id={urllib.parse.quote(temp_dtxsid)}"
                        )

                        if substance_response_text is not None:
                            temp_substance = json.loads(substance_response_text)

                            temp_SMILES = temp_substance["smiles"]
                            if not temp_SMILES:
                                return
                            temp_synonyms = []
                            temp_CAS = []

                            if "casrn" in temp_substance:
                                temp_CAS.append(temp_substance["casrn"])

                            if "preferredName" in temp_substance:
                                if temp_substance["preferredName"] is not None:
                                    temp_synonyms.append(
                                        temp_substance["preferredName"]
                                    )
                            if "acdIupacName" in temp_substance:
                                if temp_substance["acdIupacName"] is not None:
                                    temp_synonyms.append(temp_substance["acdIupacName"])
                            if "wikipediaName" in temp_substance:
                                if temp_substance["wikipediaName"] is not None:
                                    temp_synonyms.append(
                                        temp_substance["wikipediaName"]
                                    )

                            if mode == "name":
                                temp_synonyms.append(identifier)

                            QC_LEVEL_str = ""
                            if "qcLevel" in temp_substance:
                                if temp_substance["qcLevel"]:
                                    QC_LEVEL_str = (
                                        f'|QC_LEVEL:{float(temp_substance["qcLevel"])}'
                                    )
                            temp_synonyms = self.filter_and_sort_synonyms(temp_synonyms)
                            molecules.append(
                                Molecule(
                                    temp_SMILES,
                                    temp_synonyms,
                                    temp_CAS,
                                    temp_dtxsid + QC_LEVEL_str,
                                    mode,
                                    service="comptox",
                                )
                            )

                    if isinstance(results, list):
                        for i in range(len(results)):
                            temp = results[i]
                            temp_dtxsid = temp["dtxsid"]
                            process_dtxsid(temp_dtxsid)

                    else:
                        temp_dtxsid = results["dtxsid"]
                        process_dtxsid(temp_dtxsid)

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_molecule_from_CTS(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from CTS (Chemical Translation Service) based on the provided identifier and mode.

        Queries CTS to convert various chemical identifiers into molecular structures and associated information.

        Args:
            identifier (str): The chemical identifier to search for.
            mode (str): The type of identifier (e.g., 'name', 'cas', 'inchi', 'smiles').
            required_formula (Optional[str]): The expected molecular formula.
            required_charge (Optional[int]): The expected molecular charge.
            required_structure_type (Optional[str]): The expected structure type.

        Returns:
            Optional[Molecule]: A Molecule object if found and meets all requirements, None otherwise.

        Raises:
            ValueError: If the input parameters fail validation.

        Notes:
            - Checks a cache for previously retrieved molecules.
            - Uses the CTS REST API to query the service.
            - Retrieves SMILES, CAS numbers, and synonyms.
            - Filters out radicals and mixtures based on the required_structure_type.
            - Filters results based on required formula, charge, and structure type.
            - Issues a warning and returns None if CTS is down.
            - Uses resilient network requests to handle potential connection issues.
        """
        self._check_parameters(
            identifiers=identifier,
            modes=mode,
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            services="cts",
        )

        with self.query_molecule_cache("cts", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                if "CTS_is_down" in self._message_slugs_shown:
                    return None

                cts_modes = {
                    "name": "Chemical Name",
                    "cas": "CAS",
                    "inchi": "smiles",
                    "smiles": "smiles",
                }

                SMILES = None
                CAS = []
                synonyms = []

                mode_used = mode
                if mode == "smiles":
                    SMILES = identifier

                elif mode == "inchi":
                    SMILES = self.InChI_to_SMILES(identifier)
                    identifier = SMILES
                    mode_used = "smiles calculated from inchi"

                elif mode == "cas":
                    CAS.append(identifier)

                # unfortunatedly sometimes the server returns a status code 500 for valid names
                try:
                    CTS_URL = "https://cts.fiehnlab.ucdavis.edu/rest/convert/"
                    response_text = self._resilient_request(
                        f'{CTS_URL}{urllib.parse.quote(cts_modes[mode])}/Chemical%20Name/{urllib.parse.quote(identifier, safe="")}',
                        kwargs={"timeout": 5},
                        rejected_status_codes=[400, 404, 500],
                        max_retries=3,
                    )
                except requests.exceptions.ConnectionError:
                    # I don't know why, but somtimes CTS is offline. This would make the module much slower as
                    # it tries to connect multiple times anyway. Instead we give a warning and skip CTS.
                    if "CTS_is_down" not in self._message_slugs_shown:
                        self._message_slugs_shown.append("CTS_is_down")
                        warnings.warn(
                            "CTS seems to be down, to continue working this instance of MoleculeResolver will skip CTS."
                        )

                if "CTS_is_down" not in self._message_slugs_shown:
                    # the search for synonyms returns a lot of unusable data
                    # only use the three most sensible ones
                    if response_text is not None:
                        temp = json.loads(response_text)[0]
                        synonyms.extend(
                            self.filter_and_sort_synonyms(temp["results"], 3)
                        )

                    if mode != "cas":
                        response_text = self._resilient_request(
                            f'{CTS_URL}{urllib.parse.quote(cts_modes[mode])}/CAS/{urllib.parse.quote(identifier, safe="")}',
                            kwargs={"timeout": 10},
                            rejected_status_codes=[404, 500],
                        )
                        if response_text is not None:
                            CAS_rns = json.loads(response_text)[0]["results"]
                            CAS = self.filter_and_sort_CAS(CAS_rns)

                    if mode not in ["inchi", "smiles"]:
                        response_text = self._resilient_request(
                            f'{CTS_URL}{urllib.parse.quote(cts_modes[mode])}/InChI%20Code/{urllib.parse.quote(identifier, safe="")}',
                            kwargs={"timeout": 10},
                            rejected_status_codes=[404, 500],
                        )
                        if response_text is not None:
                            temp = json.loads(response_text)[0]
                            found_InChIs = temp["results"]
                            accepted_SMILES = []
                            for InChI in found_InChIs:
                                this_SMILES = self.InChI_to_SMILES(InChI)
                                if not this_SMILES:
                                    continue
                                if (
                                    not required_structure_type
                                    or "mixture" not in required_structure_type
                                ):
                                    if "." in this_SMILES:
                                        continue
                                # try filtering radicals as CTS has them in some cases
                                try:
                                    if (
                                        Descriptors.NumRadicalElectrons(
                                            self.get_from_SMILES(this_SMILES)
                                        )
                                        == 0
                                    ):
                                        accepted_SMILES.append(this_SMILES)
                                except Exception:
                                    continue

                            if len(accepted_SMILES) == 1:
                                if mode == "name":
                                    synonyms.insert(0, identifier)
                                    synonyms = self.filter_and_sort_synonyms(synonyms)
                                SMILES = accepted_SMILES[0]
                    if SMILES is not None:
                        molecules.append(
                            Molecule(SMILES, synonyms, CAS, None, mode_used, "cts")
                        )

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_CompTox_request_unique_id(self) -> str:
        """
        Generate a unique identifier for CompTox requests based on the current timestamp.

        Returns:
            str: A unique base36-encoded string derived from the current timestamp.

        Notes:
            - Uses the current time in milliseconds for increased uniqueness.
            - Base36 encoding results in a shorter string compared to decimal representation.
            - Useful for generating unique, short, and time-based request IDs.
        """

        def base36encode(
            number: int,
            alphabet: Optional[str] = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ".lower(),
        ):
            """Converts an integer to a base36 string.

            Args:
                number (int): The integer to convert.
                alphabet (Optional[str]): The alphabet to use for base36 encoding.
                Defaults to lowercase alphanumeric characters.

            Returns:
                str: The base36 encoded string.

            Raises:
                TypeError: If the input is not an integer.
            """
            """Converts an integer to a base36 string."""
            if not isinstance(number, (int)):
                raise TypeError("number must be an integer")

            base36 = ""
            sign = ""

            if number < 0:
                sign = "-"
                number = -number

            if 0 <= number < len(alphabet):
                return sign + alphabet[number]

            while number != 0:
                number, i = divmod(number, len(alphabet))
                base36 = alphabet[i] + base36

            return sign + base36

        return base36encode(int(time.time() * 1000))

    def get_molecules_from_CompTox_batchmode(
        self, identifiers: list[str], mode: str
    ) -> list[Optional[Molecule]]:
        """
        Retrieve molecules from CompTox in batch mode.

        This method queries the CompTox (Computational Toxicology) database to retrieve
        molecule information for multiple identifiers in a single batch request.

        Args:
            identifiers (list[str]): A list of chemical identifiers to search for in CompTox.
            mode (str): The type of identifier. Supported modes are 'name', 'cas', and 'inchikey'.

        Returns:
            list[Optional[Molecule]]: A list of Molecule objects corresponding to the input identifiers.
            Each element is either a Molecule object if found, or None if not found or invalid.

        Raises:
            TypeError: If any of the identifiers is not a string.
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks a cache for previously retrieved molecules.
            - It uses the CompTox batch search API to query the database.
            - The search is performed based on the specified mode for all identifiers.
            - The method retrieves SMILES, synonyms, CAS numbers, IUPAC names, and QC levels.
            - It creates Molecule objects for valid results, including metadata.
            - The results are filtered based on the quality of the match and data.
            - The method uses resilient network requests to handle potential connection issues.
            - It includes a polling mechanism to wait for the CompTox job to complete.
            - The results are processed from an Excel file returned by the CompTox API.
            - Synonyms are extracted from a separate sheet in the Excel file.

        Example:
            >>> resolver = MoleculeResolver()
            >>> identifiers = ["50-00-0", "64-17-5", "71-43-2"]
            >>> molecules = resolver.get_molecules_from_CompTox_batchmode(identifiers, mode="cas")
            >>> for identifier, molecule in zip(identifiers, molecules):
            ...     if molecule:
            ...         print(f"Found molecule for {identifier}: {molecule.get_SMILES()}")
            ...     else:
            ...         print(f"No molecule found for {identifier}")
        """
        self._check_parameters(
            identifiers=identifiers,
            modes=mode,
            services="comptox",
            context="get_molecules_batch",
        )
        if not all([isinstance(identifier, str) for identifier in identifiers]):
            raise TypeError("All identifiers must be strings.")

        with self.query_molecule_cache_batchmode(
            "comptox", mode, identifiers, save_not_found=False
        ) as (identifiers_to_search, indices_of_identifiers_to_search, results):
            if len(identifiers_to_search) == 0:
                return results

            COMPTOX_URL = "https://comptox.epa.gov/dashboard-api/batchsearch/export/"
            comptox_modes = {
                "name": "chemical_name",
                "cas": "CASRN",
                "inchikey": "INCHIKEY",
            }

            postdata = {
                "downloadItems": [
                    "SYNONYM_IDENTIFIER",
                    "QC_LEVEL",
                    "CASRN",
                    "IUPAC_NAME",
                    "SMILES",
                ],
                "downloadType": "EXCEL",
                "identifierTypes": [comptox_modes[mode]],
                "inputType": "IDENTIFIER",
                "massError": 0,
                "searchItems": "\n".join(identifiers_to_search),
                # 'qc_level', 'expocast', 'data_sources', 'toxvaldata', 'assays', 'number_of_pubmed_articles',
                # 'number_of_pubchem_data_sources', 'number_of_cpdat_sources', 'in_iris_list', 'in_pprtv_list',
                # 'in_wikipedia_list', 'qsar_ready_smiles', 'ms_ready_smiles', 'synonym_identifier'
            }

            def poll_request(job_id):
                download_url = None
                n_try = 0
                while True:
                    poll_response_text = self._resilient_request(
                        f"{COMPTOX_URL}/status/{job_id}/?{self.get_CompTox_request_unique_id()}"
                    )

                    if poll_response_text.strip().lower() == "true":
                        download_url = f"{COMPTOX_URL}content/{job_id}"
                        break

                    if n_try > 15:
                        break

                    n_try += 1

                    time.sleep(2)

                return download_url

            post_request_text = self._resilient_request(
                f"{COMPTOX_URL}/?{self.get_CompTox_request_unique_id()}",
                json=postdata,
                request_type="post",
                accepted_status_codes=[202],
            )  # allow_redirects=True)

            download_url = poll_request(post_request_text.strip())

            # QC_Level : https://jcheminf.biomedcentral.com/articles/10.1186/s13321-017-0247-6/tables/1
            # 1 Expert curated: highest confidence in accuracy and consistency of unique chemical identifiers
            # 2 Expert curated: unique chemical identifiers confirmed using multiple public sources
            # 3 Programmatically curated from high quality EPA source(s) and unique chemical identifiers have no conflicts in ChemIDPlus and PubChem
            # 4 Programmatically curated from ChemIDPlus. Unique chemical identifiers have no conflicts in PubChem
            # 5 Programmatically curated from ACToR or PubChem. Unique chemical identifiers have low confidence and have a single public source

            # in some rare occasions, this times out and returns None
            if download_url:
                request_response = self._resilient_request(
                    download_url, return_response=True
                )
                with (
                    tempfile.TemporaryDirectory() as temp_dir,
                    warnings.catch_warnings(),
                ):
                    warnings.simplefilter("ignore")

                    temp_path = os.path.join(
                        temp_dir, "moleculeresolver_comptox_postdata.xlsx"
                    )
                    with open(temp_path, "wb") as f:
                        f.write(request_response.content)

                    wb = openpyxl.load_workbook(temp_path)

                    synonym_sheet = wb.worksheets[2]
                    synonyms_by_identifier_lower = {}
                    for i_row, row in enumerate(synonym_sheet.iter_rows()):
                        if i_row == 0:
                            continue

                        row_values = [cell.value for cell in row]
                        synonyms = []
                        identifier = row_values[0].strip()
                        if row_values[1]:
                            synonyms = [v.strip() for v in row_values[1].split("|")]

                        synonyms_by_identifier_lower[identifier.lower()] = synonyms

                    temp_results_by_identifier = {}

                    def parse_CompTox_result(row):
                        identifier = str(row[0]).strip()
                        found_by = row[1].lower().strip()

                        SMILES = None
                        CAS = []
                        synonyms = []

                        if found_by.count("found 0 results") == 0:
                            SMILES = row[7].strip()
                            if not SMILES:
                                return
                            dtxsid = row[2].strip()
                            preferred_name = row[3].strip()
                            qc_level = row[4]
                            CAS = str(row[5]).strip()
                            iupac_name = row[6].strip()

                            if iupac_name and iupac_name != "N/A":
                                if iupac_name not in synonyms:
                                    synonyms.insert(0, iupac_name)

                            if preferred_name and preferred_name != "N/A":
                                if preferred_name not in synonyms:
                                    synonyms.insert(0, preferred_name)

                            if not self.is_valid_CAS(CAS):
                                CAS = []
                            else:
                                CAS = [CAS]

                            if "N/A" in SMILES or not self.is_valid_SMILES(SMILES):
                                return

                            # the rating is needed because for some substances it
                            # returns more than one row. It is used to get the best result later.
                            if "expert" in found_by:
                                own_rating = 10
                            elif "approved" in found_by:
                                own_rating = 8
                            elif "valid source" in found_by:
                                own_rating = 6
                            elif "systematic name" in found_by:
                                own_rating = 4
                            else:
                                own_rating = 2

                            if CAS:
                                own_rating += 1

                            if identifier not in temp_results_by_identifier:
                                temp_results_by_identifier[identifier] = []

                            if SMILES:
                                temp_results_by_identifier[identifier].append(
                                    (
                                        qc_level,
                                        own_rating,
                                        SMILES,
                                        synonyms,
                                        CAS,
                                        dtxsid + "|QC_LEVEL:" + str(qc_level),
                                    )
                                )

                        else:
                            temp_results_by_identifier[identifier] = None

                    results_sheet = wb.worksheets[1]
                    for i_row, row in enumerate(results_sheet.iter_rows()):
                        if i_row == 0:
                            continue

                        row_values = [cell.value for cell in row]
                        parse_CompTox_result(row_values)

                    for molecule_index, identifier in zip(
                        indices_of_identifiers_to_search,
                        identifiers_to_search,
                        strict=True,
                    ):
                        if identifier in temp_results_by_identifier:
                            temp_result = temp_results_by_identifier[identifier]
                            if temp_result:
                                if len(temp_result) > 1:
                                    best_results = temp_result
                                    try:
                                        best_qc_level = min(
                                            temp_result, key=lambda x: x[0]
                                        )[0]
                                        best_results = list(
                                            filter(
                                                lambda x: x[0] == best_qc_level,
                                                temp_result,
                                            )
                                        )
                                    except Exception:
                                        pass

                                    if len(best_results) > 1:
                                        best_own_rating = max(
                                            best_results, key=lambda x: x[1]
                                        )[1]
                                        best_results = list(
                                            filter(
                                                lambda x: x[1] == best_own_rating,
                                                best_results,
                                            )
                                        )

                                    temp_results_by_identifier[identifier] = (
                                        best_results
                                    )
                                results[molecule_index] = []

                                for (
                                    _,
                                    _,
                                    SMILES,
                                    synonyms,
                                    CAS,
                                    additional_information,
                                ) in temp_results_by_identifier[identifier]:
                                    this_synonyms = synonyms.copy()
                                    for synonym in synonyms:
                                        if (
                                            synonym.lower()
                                            in synonyms_by_identifier_lower
                                        ):
                                            for (
                                                new_synonym
                                            ) in synonyms_by_identifier_lower[
                                                synonym.lower()
                                            ]:
                                                if new_synonym not in this_synonyms:
                                                    this_synonyms.append(new_synonym)
                                    synonyms = self.filter_and_sort_synonyms(
                                        this_synonyms
                                    )
                                    results[molecule_index].append(
                                        Molecule(
                                            SMILES,
                                            synonyms,
                                            CAS,
                                            additional_information,
                                            mode,
                                            "comptox",
                                            identifier=identifier,
                                        )
                                    )

            return results

    def get_molecule_from_Chemeo(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from Chemeo based on the provided identifier and mode.

        This method queries the Chemeo database to retrieve molecule information using
        various types of chemical identifiers.

        Args:
            identifier (str): The chemical identifier to search for in Chemeo.
            mode (str): The type of identifier. Supported modes include 'name', 'cas', and 'smiles'.
            required_formula (Optional[str]): The expected molecular formula. Defaults to None.
            required_charge (Optional[int]): The expected molecular charge. Defaults to None.
            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and meets all requirements,
            None otherwise.

        Raises:
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks a cache for previously retrieved molecules.
            - It requires a valid Chemeo API token to perform the search.
            - The search is performed in two steps: first converting the identifier to a Chemeo CID,
              then searching for the compound details using the CID.
            - It extracts SMILES, synonyms, CAS numbers, and other relevant information from the API response.
            - The method prioritizes 3D mol block over 2D mol block, then InChI, and finally SMILES for structure representation.
            - It filters results to ensure the returned molecule matches the input SMILES (if mode is 'smiles').
            - The results are further filtered based on the required formula, charge, and structure type.
            - The method uses resilient network requests to handle potential connection issues.

        Example:
            >>> resolver = MoleculeResolver()
            >>> molecule = resolver.get_molecule_from_Chemeo("ethanol", mode="name")
            >>> if molecule:
            ...     print(f"Found molecule: {molecule.get_SMILES()}")
            ... else:
            ...     print("No matching molecule found")
        """
        self._check_parameters(
            identifiers=identifier,
            modes=mode,
            required_charges=required_charge,
            required_formulas=required_formula,
            required_structure_types=required_structure_type,
            services="chemeo",
        )

        with self.query_molecule_cache("chemeo", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                valid_token_found = (
                    self.available_service_API_keys["chemeo"] is not None
                )
                if valid_token_found:
                    valid_token_found = isinstance(
                        self.available_service_API_keys["chemeo"], str
                    )
                    valid_token_found = self.chemeo_API_token_regex_compiled.match(
                        self.available_service_API_keys["chemeo"]
                    )

                if not valid_token_found:
                    if "chemeo_API_token_missing" not in self._message_slugs_shown:
                        self._message_slugs_shown.append("chemeo_API_token_missing")
                        warnings.warn(
                            "chemeo requires a valid API token to allow search. After registering please insert the token in the corresponding variable when initializing the class instance."
                        )

                    return None

                CHEMEO_URL = "https://www.chemeo.com/api/v1/"
                API_bearer_headers = {}
                API_bearer_headers["Accept"] = "application/json"
                API_bearer_headers["Authorization"] = (
                    f'Bearer {self.available_service_API_keys["chemeo"]}'
                )
                temp_mode = mode if mode != "cas" else "name"
                request_text = self._resilient_request(
                    f'{CHEMEO_URL}convert/{temp_mode}/{urllib.parse.quote(identifier, safe="")}',
                    kwargs={"headers": API_bearer_headers},
                    rejected_status_codes=[403, 404],
                )

                if (
                    request_text is not None
                    and request_text != '{"message":"Compound not found."}'
                ):
                    temp = json.loads(request_text)
                    request_text = self._resilient_request(
                        f'{CHEMEO_URL}search?q={temp["cid"]}',
                        kwargs={"headers": API_bearer_headers},
                        rejected_status_codes=[403, 404],
                    )

                    temp = json.loads(request_text)

                    for i in range(temp["total"]):
                        result = temp["comps"][i]

                        temp_synonyms = []
                        temp_SMILES = None

                        if "other_names" in result:
                            temp_synonyms.extend(result["other_names"])
                        temp_synonyms = self.filter_and_sort_synonyms(temp_synonyms)

                        temp_CAS = []
                        if "cas" in result:
                            temp_CAS = result["cas"]
                            if not self.is_valid_CAS(temp_CAS):
                                temp_CAS = []
                            else:
                                temp_CAS = [temp_CAS]

                        # using the mol block first instead of the readily available SMILES,
                        # because the later one is not an isomeric SMILES
                        if "mol3d" in result:
                            temp_SMILES = self.get_SMILES_from_Mol_format(
                                molblock=result["mol3d"]
                            )

                        if temp_SMILES is None and "mol2d" in result:
                            temp_SMILES = self.get_SMILES_from_Mol_format(
                                molblock=result["mol2d"]
                            )

                        if temp_SMILES is None and "inchi" in result:
                            temp_SMILES = self.InChI_to_SMILES(result["inchi"])

                        if temp_SMILES is None and "smiles" in result:
                            temp_SMILES = result["smiles"]

                        if temp_SMILES is None:
                            raise ValueError("Structure could not be found.")

                        # sometimes chemeo gives back substructure results when searching by SMILES,
                        # this is to filter out most of them ensuring that at least
                        # non-isomeric SMILES are the same:
                        if mode == "smiles":
                            if not self.are_SMILES_equal(
                                temp_SMILES, identifier, differentiate_isomers=False
                            ):
                                continue

                        molecules.append(
                            Molecule(
                                temp_SMILES,
                                temp_synonyms,
                                temp_CAS,
                                result["id"],
                                mode,
                                "chemeo",
                            )
                        )

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_molecules_from_pubchem_batchmode(
        self, original_identifiers: list[str], mode: str
    ) -> list[Optional[list[Molecule]]]:
        """
        Retrieve molecules from PubChem in batch mode.

        This method queries the PubChem database to retrieve molecule information for multiple
        identifiers in a single batch request using the PubChem Power User Gateway (PUG).

        Args:
            original_identifiers (list[str]): A list of chemical identifiers to search for in PubChem.

            mode (str): The type of identifier. Supported modes include 'name', 'cas', 'formula',
            'smiles', 'inchi', and 'inchikey'.

        Returns:
            list[Optional[list[Molecule]]]: A list where each element corresponds to an input identifier.
            Each element is either a list of Molecule objects (if found) or None (if not found or invalid).
            Multiple Molecule objects may be returned for a single identifier if multiple matches are found.

        Raises:
            TypeError: If any of the identifiers is not a string.
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks parameters using the _check_parameters method.
            - It cleans identifiers, especially for 'name' mode, removing certain prefixes and suffixes.
            - The method uses PubChem's PUG XML API for batch requests.
            - It performs two main steps: CID search and information retrieval for found CIDs.
            - The search and retrieval are done in batches to handle large numbers of identifiers efficiently.
            - It extracts SMILES, synonyms, CAS numbers, and other relevant information from the API response.
            - The method uses resilient network requests to handle potential connection issues.
            - It includes a polling mechanism to wait for PubChem jobs to complete.
            - Results are processed from gzipped files returned by the PubChem API.
            - The method standardizes SMILES strings and creates Molecule objects with metadata.
            - API: https://pubchem.ncbi.nlm.nih.gov/docs/power-user-gateway
            - API: https://pubchem.ncbi.nlm.nih.gov/docs/identifier-exchange-service

        Example:
            >>> resolver = MoleculeResolver()
            >>> identifiers = ["ethanol", "acetone", "benzene"]
            >>> results = resolver.get_molecules_from_pubchem_batchmode(identifiers, mode="name")
            >>> for identifier, molecules in zip(identifiers, results):
            ...     if molecules:
            ...         print(f"Found {len(molecules)} molecule(s) for {identifier}")
            ...     else:
            ...         print(f"No molecules found for {identifier}")
        """

        self._check_parameters(
            identifiers=original_identifiers,
            modes=mode,
            services="pubchem",
            context="get_molecules_batch",
        )
        if not all([type(identifier) is str for identifier in original_identifiers]):
            raise TypeError("All identifiers must be strings.")

        def clean_identifier(identifier):
            if mode == "name":
                identifier = self.clean_chemical_name(
                    identifier, spell_out_greek_characters=True
                )

                if identifier.endswith(", (+-)-"):
                    identifier = "".join(identifier.split(", (+-)-")[:-1])
                if identifier.startswith("(+-)-"):
                    identifier = "".join(identifier.split("(+-)-")[1:])

                return identifier
            else:
                return identifier

        PUBCHEM_URL = "https://pubchem.ncbi.nlm.nih.gov/pug/pug.cgi"
        pubchem_xml_mode = {
            "name": "synonyms",
            "cas": "synonyms",
            "formula": "synonyms",
            "smiles": "smiles",
            "inchi": "inchis",
            "inchikey": "inchi-keys",
        }

        def parse_request_id(request_text):
            matches = regex.findall(
                "PCT-Waiting_reqid>(.*)</PCT-Waiting_reqid", request_text
            )
            if len(matches) != 1:
                return None
            request_id = matches[0]
            return request_id

        def cid_search_request():
            query_Uids = []
            for cleaned_unique_identifier in cleaned_unique_identifiers:
                query_Uids.append(
                    f"<PCT-QueryUids_{pubchem_xml_mode[mode]}_E>{cleaned_unique_identifier}</PCT-QueryUids_{pubchem_xml_mode[mode]}_E>"
                )
            temp = "".join(query_Uids)
            root_query_Uids = f"<PCT-QueryUids_{pubchem_xml_mode[mode]}>{temp}</PCT-QueryUids_{pubchem_xml_mode[mode]}>"

            xml_template = f"""<?xml version="1.0"?>
                                <!DOCTYPE PCT-Data PUBLIC "-//NCBI//NCBI PCTools/EN" "NCBI_PCTools.dtd">
                                <PCT-Data>
                                <PCT-Data_input>
                                    <PCT-InputData>
                                    <PCT-InputData_query>
                                        <PCT-Query>
                                        <PCT-Query_type>
                                            <PCT-QueryType>
                                            <PCT-QueryType_id-exchange>
                                                <PCT-QueryIDExchange>
                                                <PCT-QueryIDExchange_input>
                                                    <PCT-QueryUids>
                                                    {root_query_Uids}
                                                    </PCT-QueryUids>
                                                </PCT-QueryIDExchange_input>
                                                <PCT-QueryIDExchange_operation-type value="same"/>
                                                <PCT-QueryIDExchange_output-type value="cid"/>
                                                <PCT-QueryIDExchange_output-method value="file-pair"/>
                                                <PCT-QueryIDExchange_compression value="gzip"/>
                                                </PCT-QueryIDExchange>
                                            </PCT-QueryType_id-exchange>
                                            </PCT-QueryType>
                                        </PCT-Query_type>
                                        </PCT-Query>
                                    </PCT-InputData_query>
                                    </PCT-InputData>
                                </PCT-Data_input>
                                </PCT-Data>"""

            request_id = None
            n_try = 0
            while request_id is None and n_try < 5:
                request_text = self._resilient_request(
                    PUBCHEM_URL,
                    {
                        "data": xml_template.encode("utf-8"),
                        "headers": {"Content-type": "application/xml; charset=utf-8"},
                    },
                    request_type="post",
                )

                if (
                    request_text is not None
                    and request_text.count("Result set is empty.") == 0
                    and request_text.count("server-error") == 0
                ):
                    request_id = parse_request_id(request_text)

                n_try += 1
                time.sleep(3)

            return request_id

        def info_request(cids_to_request, info_name):
            xml_for_cids_to_request = [
                f"<PCT-ID-List_uids_E>{cid}</PCT-ID-List_uids_E>"
                for cid in cids_to_request
            ]
            xml_template = f"""<?xml version="1.0"?>
                            <!DOCTYPE PCT-Data PUBLIC "-//NCBI//NCBI PCTools/EN" "NCBI_PCTools.dtd">
                            <PCT-Data>
                            <PCT-Data_input>
                                <PCT-InputData>
                                <PCT-InputData_query>
                                    <PCT-Query>
                                    <PCT-Query_type>
                                        <PCT-QueryType>
                                        <PCT-QueryType_id-exchange>
                                            <PCT-QueryIDExchange>
                                            <PCT-QueryIDExchange_input>
                                                <PCT-QueryUids>
                                                <PCT-QueryUids_ids>
                                                    <PCT-ID-List>
                                                    <PCT-ID-List_db>pccompound</PCT-ID-List_db>
                                                    <PCT-ID-List_uids>
                                                        {''.join(xml_for_cids_to_request)}
                                                    </PCT-ID-List_uids>
                                                    </PCT-ID-List>
                                                </PCT-QueryUids_ids>
                                                </PCT-QueryUids>
                                            </PCT-QueryIDExchange_input>
                                            <PCT-QueryIDExchange_operation-type value="same"/>
                                            <PCT-QueryIDExchange_output-type value="{info_name}"/>
                                            <PCT-QueryIDExchange_output-method value="file-pair"/>
                                            <PCT-QueryIDExchange_compression value="gzip"/>
                                            </PCT-QueryIDExchange>
                                        </PCT-QueryType_id-exchange>
                                        </PCT-QueryType>
                                    </PCT-Query_type>
                                    </PCT-Query>
                                </PCT-InputData_query>
                                </PCT-InputData>
                            </PCT-Data_input>
                            </PCT-Data>"""

            request_id = None
            n_try = 0
            while request_id is None or n_try < 5:
                request_text = self._resilient_request(
                    PUBCHEM_URL,
                    {
                        "data": xml_template.encode("utf-8"),
                        "headers": {"Content-type": "application/xml; charset=utf-8"},
                    },
                    request_type="post",
                )

                if (
                    request_text is not None
                    and request_text.count("Result set is empty.") == 0
                    and request_text.count("server-error") == 0
                ):
                    request_id = parse_request_id(request_text)

                n_try += 1
                time.sleep(3)

            return request_id

        def poll_request(_request_id):
            download_url = None
            poll_request_xml = f"""<?xml version="1.0"?>
                        <!DOCTYPE PCT-Data PUBLIC "-//NCBI//NCBI PCTools/EN" "http://pubchem.ncbi.nlm.nih.gov/pug/pug.dtd">
                        <PCT-Data>
                        <PCT-Data_input>
                            <PCT-InputData>
                            <PCT-InputData_request>
                                <PCT-Request>
                                <PCT-Request_reqid>{_request_id}</PCT-Request_reqid>
                                <PCT-Request_type value="status"/>
                                </PCT-Request>
                            </PCT-InputData_request>
                            </PCT-InputData>
                        </PCT-Data_input>
                        </PCT-Data>"""

            n_try = 0
            while True:
                poll_response = self._resilient_request(
                    PUBCHEM_URL,
                    {
                        "data": poll_request_xml.encode("utf-8"),
                        "headers": {
                            "Content-type": "application/xml; charset=utf-8",
                            "timeout": str(30),
                        },
                    },
                    request_type="post",
                )
                if poll_response:
                    matches = regex.findall(
                        "PCT-Download-URL_url>(.*)</PCT-Download-URL_url", poll_response
                    )

                    if matches:
                        if len(matches) != 1:
                            raise
                        download_url = matches[0]
                        if download_url.startswith("ftp://"):
                            download_url = "https" + download_url[3:]
                            break

                if n_try > 15:
                    break

                time.sleep(2)

            return download_url

        def get_results(download_url):
            found_results = {}
            n_try = 5
            while True:
                try:
                    with closing(urllib.request.urlopen(download_url)) as r:
                        zipped_content = gzip.GzipFile(fileobj=r)
                        content = zipped_content.read()
                        content_str = content.decode("utf-8")
                        results = content_str.split("\n")
                        found_results = [
                            result
                            for result in results
                            if result != "" and result != "Result set is empty."
                        ]
                        break
                except Exception as e:
                    if n_try > 5:
                        raise e
                    n_try += 1
                    time.sleep(3)

            return found_results

        with self.query_molecule_cache_batchmode(
            "pubchem", mode, original_identifiers, save_not_found=False
        ) as (
            original_identifiers_to_search,
            indices_of_identifiers_to_search,
            results,
        ):
            if len(original_identifiers_to_search) == 0:
                return results

            cleaned_identifiers = [
                clean_identifier(identifier)
                for identifier in original_identifiers_to_search
            ]  # cleaned identifiers
            cleaned_unique_identifiers = set(cleaned_identifiers)

            request_id = cid_search_request()

            if request_id is not None:
                download_url = poll_request(request_id)

                found_results_by_identifier = {}
                found_results_by_cid_identifier = {}
                found_results_by_cid = {}
                SMILES = {}
                CAS = {}

                results_with_name_errors = []
                for result in get_results(download_url):
                    cleaned_unique_identifier, cid = result.split("\t")

                    if cleaned_unique_identifier not in found_results_by_cid_identifier:
                        found_results_by_cid_identifier[cleaned_unique_identifier] = []

                    if cid != "":
                        found_results_by_cid_identifier[
                            cleaned_unique_identifier
                        ].append(int(cid))
                    else:
                        found_results_by_cid_identifier[
                            cleaned_unique_identifier
                        ].append(None)

                    if mode == "name":
                        if "#" in cleaned_unique_identifier:
                            results_with_name_errors.append(cleaned_unique_identifier)

                if results_with_name_errors:
                    temp = {}
                    for name_with_error in results_with_name_errors:
                        parts = name_with_error.split("#")
                        longest_part = max(parts, key=len)
                        candidates = [
                            v for v in cleaned_identifiers if longest_part in v
                        ]
                        temp[name_with_error] = candidates
                    print(
                        "The following identifiers have errors: {}".format(
                            ", ".join(temp)
                        )
                    )

                if len(found_results_by_cid_identifier) > 0:
                    for original_identifier, cleaned_identifier in zip(
                        original_identifiers_to_search, cleaned_identifiers, strict=True
                    ):
                        if cleaned_identifier in found_results_by_cid_identifier:
                            cids = found_results_by_cid_identifier[cleaned_identifier]
                            cids = [cid for cid in cids if cid is not None]
                            if len(cids) == 0:
                                found_results_by_identifier[original_identifier] = None
                            else:
                                cid = cids[
                                    0
                                ]  # first cid seems always to be the best match
                                found_results_by_identifier[original_identifier] = cid
                                found_results_by_cid[cid] = original_identifier

                    cids_to_request = found_results_by_cid.keys()

                    synonyms = {}
                    for cid in cids_to_request:
                        if cid not in synonyms:
                            synonyms[cid] = []
                            CAS[cid] = []

                        if mode == "name":
                            this_cid_requested_name = found_results_by_cid[cid]
                            synonyms[cid].append(this_cid_requested_name)

                    request_id = info_request(cids_to_request, "iupac")
                    download_url = poll_request(request_id)
                    iupac_results = get_results(download_url)
                    if not len(cids_to_request) <= len(iupac_results):
                        raise

                    for result in iupac_results:
                        cid, iupac_name = result.split("\t")

                        if len(iupac_name) > 1:
                            synonyms[int(cid)].append(iupac_name)

                    request_id = info_request(cids_to_request, "synonyms")
                    download_url = poll_request(request_id)
                    synonym_results = get_results(download_url)

                    for result in synonym_results:
                        cid, synonym = result.split("\t")
                        synonyms[int(cid)].append(synonym)

                    request_id = info_request(cids_to_request, "smiles")
                    download_url = poll_request(request_id)
                    SMILES_results = get_results(download_url)
                    if not len(cids_to_request) <= len(SMILES_results):
                        raise

                    for result in SMILES_results:
                        cid, this_SMILES = result.split("\t")
                        SMILES[int(cid)] = this_SMILES

                for molecule_index, original_identifier in zip(
                    indices_of_identifiers_to_search,
                    original_identifiers_to_search,
                    strict=True,
                ):
                    cid = (
                        found_results_by_identifier[original_identifier]
                        if original_identifier in found_results_by_identifier
                        else None
                    )
                    if cid is not None:
                        results[molecule_index] = [
                            Molecule(
                                SMILES[cid],
                                self.filter_and_sort_synonyms(synonyms[cid]),
                                self.filter_and_sort_CAS(synonyms[cid]),
                                cid,
                                mode,
                                "pubchem",
                                identifier=original_identifier,
                            )
                        ]

            return results

    def get_molecule_from_pubchem(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from PubChem based on the provided identifier and mode.

        This method queries the PubChem database to retrieve molecule information using
        various types of chemical identifiers.

        Args:
            identifier (str): The chemical identifier to search for in PubChem.

            mode (str): The type of identifier. Supported modes include 'name', 'cas', 'formula',
            'smiles', 'inchi', and 'inchikey'.

            required_formula (Optional[str]): The expected molecular formula. Defaults to None.

            required_charge (Optional[int]): The expected molecular charge. Defaults to None.

            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and meets all requirements,
                                None otherwise.

        Raises:
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks a cache for previously retrieved molecules.
            - It uses the PubChem REST API to query the database.
            - For 'formula' mode, it retrieves multiple results and processes them individually.
            - The method extracts SMILES, synonyms, CAS numbers, and PubChem CID from the API response.
            - It standardizes SMILES strings and creates Molecule objects with metadata.
            - The results are filtered based on the required formula, charge, and structure type.
            - For 'formula' mode, if no results are found, it tries searching with the Hill formula.
            - The method uses resilient network requests to handle potential connection issues.
            - It includes special handling for different PubChem response formats and data structures.

        Example:
            >>> resolver = MoleculeResolver()
            >>> molecule = resolver.get_molecule_from_pubchem("ethanol", mode="name")
            >>> if molecule:
            ...     print(f"Found molecule: {molecule.get_SMILES()}")
            ... else:
            ...     print("No matching molecule found")
        """
        if required_formula is None:
            if mode == "formula":
                required_formula = identifier

        self._check_parameters(
            identifiers=identifier,
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            modes=mode,
            services="pubchem",
        )

        with self.query_molecule_cache("pubchem", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                SMILES = None
                synonyms = []
                CAS = []
                cid = None

                PUBCHEM_URL = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/"
                pubchem_mode = mode
                if mode == "cas":
                    pubchem_mode = "name"

                # here we use rejected_status_codes = [400, 404] 400 means the request was not correctly formatted, but
                # we can rule this out and it is probably to the identifier causing some URL issues. Because of this it is rejected.
                accepted_status_codes = (
                    [200, 202] if pubchem_mode == "formula" else [200]
                )
                request_text = self._resilient_request(
                    f'{PUBCHEM_URL}{pubchem_mode}/json?{pubchem_mode}={urllib.parse.quote(identifier, safe="")}',
                    rejected_status_codes=[400, 404],
                    accepted_status_codes=accepted_status_codes,
                )

                accepted_results = []

                if request_text is not None:
                    temp = json.loads(request_text)

                    if pubchem_mode == "formula":
                        listkey = temp["Waiting"]["ListKey"]
                        cids_response_text = self._resilient_request(
                            f"{PUBCHEM_URL}listkey/{listkey}/cids/json",
                            rejected_status_codes=[400, 404],
                            accepted_status_codes=[200],
                            sleep_time=7,
                        )

                        if cids_response_text is not None:
                            cid_list = json.loads(cids_response_text)["IdentifierList"][
                                "CID"
                            ]

                            for cid_temp in cid_list:
                                cmp_temp = self.get_molecule_from_pubchem(
                                    str(cid_temp),
                                    "cid",
                                    required_formula,
                                    required_charge,
                                    required_structure_type,
                                )
                                if cmp_temp:
                                    cmp_temp.mode = mode
                                    accepted_results.append(cmp_temp)

                        if not accepted_results:
                            # if nothing was foung, try with the hill formula, as pubchem sometimes finds results using it
                            hill_formula = self.to_hill_formula(
                                self.formula_to_dictionary(identifier)
                            )
                            if identifier != hill_formula:
                                return self.get_molecule_from_pubchem(
                                    hill_formula,
                                    mode,
                                    required_formula,
                                    required_charge,
                                    required_structure_type,
                                )

                    else:

                        def get_prop_value(
                            _compound, label, name, conversion_funtion, all_names=False
                        ):
                            props_found = [
                                prop["value"]
                                for prop in _compound["props"]
                                if prop["urn"]["label"] == label
                                and prop["urn"]["name"] == name
                            ]

                            if all_names:
                                return {
                                    prop["urn"]["name"]: conversion_funtion(
                                        prop["value"]["sval"]
                                    )
                                    for prop in _compound["props"]
                                    if prop["urn"]["label"] == label
                                }
                            else:
                                if len(props_found) != 1:
                                    raise
                                prop_vals = list(props_found[0].values())

                                return conversion_funtion(prop_vals[0])

                        # for i in range(len(temp['PC_Compounds'])):
                        # first cid seems always to be the best match
                        compound = temp["PC_Compounds"][0]

                        if len(compound["id"]) == 0:
                            molecules.clear()
                            return

                        cid = compound["id"]["id"]["cid"]
                        SMILES = get_prop_value(compound, "SMILES", "Isomeric", str)
                        SMILES_from_InChI = self.InChI_to_SMILES(
                            get_prop_value(compound, "InChI", "Standard", str)
                        )

                        if self.check_SMILES(
                            SMILES,
                            required_formula,
                            required_charge,
                            required_structure_type,
                        ):
                            pass
                        elif self.check_SMILES(
                            SMILES_from_InChI,
                            required_formula,
                            required_charge,
                            required_structure_type,
                        ):
                            SMILES = SMILES_from_InChI
                        else:
                            molecules.clear()
                            return

                        synonym_response_text = self._resilient_request(
                            f"{PUBCHEM_URL}cid/{cid}/synonyms/TXT"
                        )
                        if synonym_response_text is None:
                            temp_synonyms = []
                        else:
                            temp_synonyms = synonym_response_text.split("\n")
                        IUPAC_names = get_prop_value(
                            compound, "IUPAC Name", "Preferred", str, all_names=True
                        )

                        if "Preferred" in IUPAC_names:
                            temp_synonyms.insert(0, IUPAC_names["Preferred"])
                        elif "CAS-like Style" in IUPAC_names:
                            temp_synonyms.insert(0, IUPAC_names["CAS-like Style"])
                        elif "Systematic" in IUPAC_names:
                            temp_synonyms.insert(0, IUPAC_names["Systematic"])
                        elif "Traditional" in IUPAC_names:
                            temp_synonyms.insert(0, IUPAC_names["Traditional"])

                        CAS = self.filter_and_sort_CAS(temp_synonyms)
                        synonyms = self.filter_and_sort_synonyms(temp_synonyms)

                        molecules.append(
                            Molecule(SMILES, synonyms, CAS, cid, mode, "pubchem")
                        )

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def get_molecule_from_CAS_registry(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from CAS Registry based on the provided identifier and mode.

        This method queries the CAS (Chemical Abstracts Service) Registry to retrieve
        molecule information using various types of chemical identifiers.

        Args:
            identifier (str): The chemical identifier to search for in CAS Registry.

            mode (str): The type of identifier. Supported modes include 'name', 'cas', 'formula',
            'smiles', 'inchi', and 'inchikey'.

            required_formula (Optional[str]): The expected molecular formula. Defaults to None.
            required_charge (Optional[int]): The expected molecular charge. Defaults to None.

            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and meets all requirements,
            None otherwise.

        Raises:
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks a cache for previously retrieved molecules.
            - It uses the CAS Common Chemistry API to query the database.
            - The search is performed in two steps: first a general search, then a detailed lookup.
            - It extracts SMILES, synonyms, CAS numbers, and InChI from the API response.
            - The method standardizes SMILES strings and creates Molecule objects with metadata.
            - Results are filtered based on the required formula, charge, and structure type.
            - For 'smiles' mode, it performs additional checks to ensure the returned structure matches the input.
            - If no results are found for 'smiles' mode, it attempts to search using the InChI representation.
            - The method uses resilient network requests to handle potential connection issues.
            - Rejected status codes (403, 404) are handled gracefully.

        Example:
            >>> resolver = MoleculeResolver()
            >>> molecule = resolver.get_molecule_from_CAS_registry("64-17-5", mode="cas")
            >>> if molecule:
            ...     print(f"Found molecule: {molecule.get_SMILES()}")
            ... else:
            ...     print("No matching molecule found")

        """
        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            modes=mode,
            services="cas_registry",
        )

        with self.query_molecule_cache("cas_registry", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                CAS_URL = "https://commonchemistry.cas.org/api/"
                SMILES = None
                synonyms = []
                CAS = []
                mode_used = mode

                search_response_text = self._resilient_request(
                    f'{CAS_URL}search?q={urllib.parse.quote(identifier, safe="")}',
                    {"headers": {"accept": "application/json"}},
                    rejected_status_codes=[403, 404],
                )

                if search_response_text is not None:
                    results = json.loads(search_response_text)
                    results = results["results"]

                    for result in results:
                        CAS = result["rn"]
                        detail_response_text = self._resilient_request(
                            f"{CAS_URL}detail?cas_rn={urllib.parse.quote(CAS)}",
                            {"headers": {"accept": "application/json"}},
                        )
                        if detail_response_text is not None:
                            details = json.loads(detail_response_text)

                            synonyms = []
                            if "name" in details:
                                synonyms.append(details["name"])
                            synonyms.extend(details["synonyms"])

                            SMILES = details["smile"]

                            # sometimes the cas registry gives back wrong aromaticity results
                            # when searching by SMILES, this is to filter out most of them
                            # ensuring that at least non-isomeric SMILES are the same:
                            if mode == "smiles":
                                if not self.are_SMILES_equal(
                                    SMILES, identifier, differentiate_isomers=False
                                ):
                                    continue

                            inchi = details["inchi"]

                            if self.check_SMILES(
                                SMILES,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            ):
                                molecules.append(
                                    Molecule(
                                        SMILES,
                                        self.filter_and_sort_synonyms(synonyms),
                                        [CAS],
                                        None,
                                        mode_used,
                                        "cas_registry",
                                    )
                                )
                            elif inchi != "":
                                SMILES_from_InChI = self.InChI_to_SMILES(inchi)
                                if self.check_SMILES(
                                    SMILES_from_InChI,
                                    required_formula,
                                    required_charge,
                                    required_structure_type,
                                ):
                                    molecules.append(
                                        Molecule(
                                            SMILES_from_InChI,
                                            self.filter_and_sort_synonyms(synonyms),
                                            [CAS],
                                            None,
                                            mode_used,
                                            "cas_registry",
                                        )
                                    )

                # the search by SMILES on the CAS common chemistry service is not doing any kind of canonicalization
                # prior to searching, so in a lot of cases it does not deliver a result even though the molecule is
                # in the database, the service however does not suffer from this issue when searching by InChI
                if mode == "smiles" and len(molecules) == 0:
                    inchi = self.SMILES_to_InChI(self.standardize_SMILES(identifier))
                    cmp = self.get_molecule_from_CAS_registry(
                        inchi,
                        "inchi",
                        required_formula,
                        required_charge,
                        required_structure_type,
                    )
                    if cmp is not None:
                        if self.are_SMILES_equal(
                            cmp.SMILES, identifier, differentiate_isomers=False
                        ):
                            cmp.mode = mode
                            molecules.append(cmp)

        return self.filter_and_combine_molecules(
            molecules,
            required_formula,
            required_charge,
            required_structure_type,
        )

    def _match_SRS_results_to_identifiers(
        self, identifiers: list[str], mode: str, results: list[dict]
    ) -> dict[str, list[tuple]]:
        """
        Match SRS (Substance Registry Services) results to input identifiers.

        This method processes the results from an SRS query and matches them to the
        original input identifiers, organizing the data for easy retrieval.

        Args:
            identifiers (list[str]): A list of chemical identifiers used in the original query.
            mode (str): The type of identifier used. Supported modes are 'cas' and 'name'.
            results (list[dict]): A list of dictionaries containing the SRS query results.

        Returns:
            dict[str, list[tuple]]: A dictionary where keys are the original identifiers and values
            are lists of tuples. Each tuple contains information about a matched molecule:
            (SMILES, primary_names, synonyms, CAS numbers, ITN, all_synonyms_lower).

        Raises:
            RuntimeError: If an ITN (Internal Tracking Number) is encountered more than once in the results.

        Notes:
            - The method processes each result, extracting relevant information such as names, CAS numbers, and SMILES.
            - It organizes the data into several dictionaries for efficient lookup:

                - infos_by_ITN: Stores all information for each ITN.
                - ITNs_by_primary_name: Maps primary names to their corresponding ITNs.
                - ITNs_by_synonym: Maps synonyms to their corresponding ITNs.
                - ITNs_by_all_names: Maps all names (primary and synonyms) to their corresponding ITNs.
                - ITNs_by_CAS: Maps CAS numbers to their corresponding ITNs.
            - The method attempts to standardize SMILES notations and convert InChI to SMILES if necessary.
            - It implements a matching algorithm that tries to find the best match for each input identifier.
            - The matching process is iterative and attempts to resolve ambiguities when multiple matches are found.
            - It uses various strategies to determine the best match, including checking for unique matches,
              comparing SMILES, and counting synonym occurrences.

        Example:
            >>> resolver = MoleculeResolver()
            >>> identifiers = ["50-00-0", "ethanol"]
            >>> mode = "name"
            >>> results = [...]  # SRS query results
            >>> matched = resolver._match_SRS_results_to_identifiers(identifiers, mode, results)
            >>> for identifier, matches in matched.items():
            ...     print(f"Matches for {identifier}: {len(matches)}")
        """
        infos_by_ITN = {}
        ITNs_by_primary_name = {}
        ITNs_by_synonym = {}
        ITNs_by_all_names = {}
        ITNs_by_CAS = {}
        for result in results:
            primary_names = []
            epaName = result["epaName"]
            if epaName:
                primary_names.append(epaName.strip())
            iupacName = result["iupacName"]
            if iupacName:
                primary_names.append(iupacName.strip())
            systematicName = result["systematicName"]
            if systematicName:
                primary_names.append(systematicName.strip())

            CAS = []
            if "currentCasNumber" in result:
                if result["currentCasNumber"]:
                    CAS.append(result["currentCasNumber"].strip())

            synonyms = []
            if "synonyms" in result:
                synonyms.extend(
                    [
                        synonym["synonymName"].strip()
                        for synonym in result["synonyms"]
                        if synonym["synonymName"]
                    ]
                )

            ITN = result["internalTrackingNumber"]
            SMILES = result["smilesNotation"]

            if SMILES is None:
                if isinstance(result["inchiNotation"], str):
                    inchi = result["inchiNotation"]
                    if not inchi.lower().startswith("inchi="):
                        inchi = "InChI=" + inchi
                    SMILES = self.InChI_to_SMILES(inchi)

            if not SMILES:
                continue

            if ITN in infos_by_ITN:
                raise RuntimeError(
                    "ITN already exists in infos_by_ITN, ITNs are expected to be returned only once."
                )

            all_synonyms_lower = []
            for name in primary_names:
                nl = name.lower()
                all_synonyms_lower.append(nl)
                if nl not in ITNs_by_primary_name:
                    ITNs_by_primary_name[nl] = []

                if ITN not in ITNs_by_primary_name[nl]:
                    ITNs_by_primary_name[nl].append(ITN)

                if nl not in ITNs_by_all_names:
                    ITNs_by_all_names[nl] = []

                if ITN not in ITNs_by_all_names[nl]:
                    ITNs_by_all_names[nl].append(ITN)

            for name in synonyms:
                nl = name.lower()
                all_synonyms_lower.append(nl)
                if nl not in ITNs_by_synonym:
                    ITNs_by_synonym[nl] = []

                if ITN not in ITNs_by_synonym[nl]:
                    ITNs_by_synonym[nl].append(ITN)

                if nl not in ITNs_by_all_names:
                    ITNs_by_all_names[nl] = []

                if ITN not in ITNs_by_all_names[nl]:
                    ITNs_by_all_names[nl].append(ITN)

            for CAS_ in CAS:
                if CAS_:
                    if CAS_ not in ITNs_by_CAS:
                        ITNs_by_CAS[CAS_] = []

                    if ITN not in ITNs_by_CAS[CAS_]:
                        ITNs_by_CAS[CAS_].append(ITN)

            infos_by_ITN[ITN] = (
                SMILES,
                primary_names,
                synonyms,
                CAS,
                ITN,
                all_synonyms_lower,
            )

        SRS_results_by_identifier = {}
        uniquely_matched_ITNs = {}

        last_SRS_results_by_identifier_length = -1
        i_iteration = 0
        max_iterations = len(results) * 2
        while len(SRS_results_by_identifier) != last_SRS_results_by_identifier_length:
            i_iteration += 1
            if i_iteration > max_iterations:
                break
            last_SRS_results_by_identifier_length = len(SRS_results_by_identifier)
            for identifier in identifiers:
                il = identifier.lower()
                if identifier not in SRS_results_by_identifier:
                    ITNs_for_this_identifier = []
                    if mode == "cas":
                        if identifier in ITNs_by_CAS:
                            ITNs_for_this_identifier = ITNs_by_CAS[identifier]

                    elif mode == "name":
                        if il in ITNs_by_primary_name:
                            ITNs_for_this_identifier = ITNs_by_primary_name[il]
                        elif il in ITNs_by_synonym:
                            ITNs_for_this_identifier = ITNs_by_synonym[il]

                    best_ITNs = None

                    if ITNs_for_this_identifier:
                        if len(ITNs_for_this_identifier) == 1:
                            best_ITNs = ITNs_for_this_identifier
                        else:
                            temptative_ITNs = set(ITNs_for_this_identifier) - set(
                                uniquely_matched_ITNs.values()
                            )
                            if len(temptative_ITNs) == 1:
                                best_ITNs = list(temptative_ITNs)
                            else:
                                unique_SMILES = set(
                                    [
                                        self.standardize_SMILES(infos_by_ITN[ITN][0])
                                        for ITN in temptative_ITNs
                                    ]
                                )
                                if len(unique_SMILES) == 1:
                                    # if same structure, use the one with the smallest ITN
                                    best_ITNs = [
                                        str(min([int(ITN) for ITN in temptative_ITNs]))
                                    ]
                                else:
                                    if len(temptative_ITNs) > 1:
                                        number_of_synonyms_found_in_ITN_info = []
                                        for ITN in temptative_ITNs:
                                            number_of_synonyms_found_in_ITN_info.extend(
                                                infos_by_ITN[ITN][5].count(il) * [ITN]
                                            )

                                        most_common = collections.Counter(
                                            number_of_synonyms_found_in_ITN_info
                                        ).most_common(2)
                                        if most_common[0][1] > most_common[1][1]:
                                            best_ITNs = [most_common[0][0]]

                    if best_ITNs:
                        if len(best_ITNs) == 1:
                            uniquely_matched_ITNs[identifier] = best_ITNs[0]

                        SRS_results_by_identifier[identifier] = [
                            infos_by_ITN[ITN] for ITN in best_ITNs
                        ]

        return SRS_results_by_identifier

    def get_molecules_from_SRS_batchmode(
        self, identifiers: list[str], mode: str
    ) -> list[Optional[list[Molecule]]]:
        """
        Retrieve molecules from SRS (Substance Registry Services) in batch mode.

        This method queries the EPA's Substance Registry Services to retrieve molecule
        information for multiple identifiers in a single batch request.

        Args:
            identifiers (list[str]): A list of chemical identifiers to search for in SRS.

            mode (str): The type of identifier. Supported modes are determined by the
            _check_parameters method, typically including 'name' and 'cas'.

        Returns:
            list[Optional[list[Molecule]]]: A list where each element corresponds to an input
            identifier. Each element is either a list of Molecule objects (if found) or
            None (if not found or invalid).

        Raises:
            TypeError: If any of the identifiers is not a string.
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks a cache for previously retrieved molecules.
            - It uses the SRS REST API to query the database.
            - The search is performed in batches to handle large numbers of identifiers efficiently.
            - It extracts SMILES, synonyms, CAS numbers, and other relevant information from the API response.
            - The method uses resilient network requests to handle potential connection issues.
            - Results are processed and matched to the original identifiers using the _match_SRS_results_to_identifiers method.
            - The SRS API has a limit on URL length, so the method splits large batches into smaller chunks.
            - Synonyms are filtered and sorted before being added to the Molecule objects.
            - API: https://www.postman.com/api-evangelist/workspace/environmental-protection-agency-epa/collection/35240-6b84cc71-ce77-48b8-babd-323eb8d670bd
            - new API; https://cdxappstest.epacdx.net/oms-substance-registry-services/swagger-ui/

        Example:
            >>> resolver = MoleculeResolver()
            >>> identifiers = ["50-00-0", "64-17-5", "71-43-2"]
            >>> molecules = resolver.get_molecules_from_SRS_batchmode(identifiers, mode="cas")
            >>> for identifier, molecule in zip(identifiers, molecules):
            ...     if molecule:
            ...         print(f"Found molecule for {identifier}: {molecule[0].get_SMILES()}")
            ...     else:
            ...         print(f"No molecule found for {identifier}")
        """

        self._check_parameters(
            identifiers=identifiers,
            modes=mode,
            services="srs",
            context="get_molecules_batch",
        )
        if not all([type(identifier) is str for identifier in identifiers]):
            raise TypeError("All identifiers must be strings.")

        with self.query_molecule_cache_batchmode(
            "srs", mode, identifiers, save_not_found=False
        ) as (identifiers_to_search, indices_of_identifiers_to_search, results):
            if len(identifiers_to_search) == 0:
                return results

            SRS_URL = "https://cdxappstest.epacdx.net/oms-substance-registry-services/rest-api/substances"
            chunks_identifiers = []
            chunks_identifer_indices = []
            this_chunk_identifiers = []
            this_chunk_identifier_indices = []
            for identifier_index, identifier in zip(
                indices_of_identifiers_to_search, identifiers_to_search, strict=True
            ):
                this_chunk_identifiers.append(identifier)
                this_chunk_identifier_indices.append(identifier_index)
                if (
                    len(
                        f'{SRS_URL}/{mode}?{mode}List={urllib.parse.quote("|".join(this_chunk_identifiers))}&qualifier=exact'
                    )
                    >= 2000
                ):
                    chunks_identifiers.append(this_chunk_identifiers)
                    chunks_identifer_indices.append(this_chunk_identifier_indices)
                    this_chunk_identifiers = []
                    this_chunk_identifier_indices = []

            if this_chunk_identifiers:
                chunks_identifiers.append(this_chunk_identifiers)
                chunks_identifer_indices.append(this_chunk_identifier_indices)

            for chunk_identifier_indices, chunk_identifiers in zip(
                chunks_identifer_indices, chunks_identifiers, strict=True
            ):
                search_response_text = self._resilient_request(
                    f'{SRS_URL}/{mode}?{mode}List={urllib.parse.quote("|".join(chunk_identifiers))}&qualifier=exact',
                    rejected_status_codes=[404, 500],
                    kwargs={"timeout": 60},
                )

                if search_response_text is not None:
                    infos_by_identifier = self._match_SRS_results_to_identifiers(
                        chunk_identifiers,
                        mode,
                        json.loads(search_response_text),
                    )

                    for molecule_index, identifier in zip(
                        chunk_identifier_indices, chunk_identifiers, strict=True
                    ):
                        if identifier in infos_by_identifier:
                            this_molecules = []
                            for (
                                SMILES,
                                primary_names,
                                synonyms,
                                CAS,
                                ITN,
                                _,
                            ) in infos_by_identifier[identifier]:
                                temp_synonyms = self.filter_and_sort_synonyms(
                                    primary_names + synonyms
                                )
                                this_molecules.append(
                                    Molecule(
                                        SMILES,
                                        temp_synonyms,
                                        CAS,
                                        ITN,
                                        mode,
                                        "srs",
                                        identifier=identifier,
                                    )
                                )

                            results[molecule_index] = this_molecules

            return results

    def get_molecule_from_SRS(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from SRS (Substance Registry Services) based on the provided identifier and mode.

        This method queries the EPA's Substance Registry Services to retrieve molecule
        information using various types of chemical identifiers.

        Args:
            identifier (str): The chemical identifier to search for in SRS.

            mode (str): The type of identifier. Supported modes are determined by the
            _check_parameters method, typically including 'name' and 'cas'.

            required_formula (Optional[str]): The expected molecular formula. Defaults to None.

            required_charge (Optional[int]): The expected molecular charge. Defaults to None.

            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and meets all requirements,
            None otherwise.

        Raises:
            ValueError: If the input parameters fail validation in the _check_parameters method.
            RuntimeError: If more than one molecule is found for the given identifier.

        Notes:
            - The method first checks parameters using the _check_parameters method.
            - It then checks a cache for previously retrieved molecules.
            - The method uses the SRS REST API to query the database.
            - It constructs the API URL based on the mode and identifier.
            - The query uses resilient network requests to handle potential connection issues.
            - If a result is found, it processes the JSON response to extract relevant information.
            - The method extracts SMILES, primary names, synonyms, CAS numbers, and ITN (Internal Tracking Number).
            - It filters and sorts synonyms before creating the Molecule object.
            - The SMILES is checked against required formula, charge, and structure type if specified.
            - Only one molecule is returned; if multiple are found, it raises a RuntimeError.

        Example:
            >>> resolver = MoleculeResolver()
            >>> molecule = resolver.get_molecule_from_SRS("50-00-0", mode="cas")
            >>> if molecule:
            ...     print(f"Found molecule: {molecule.get_SMILES()}")
            ... else:
            ...     print("No matching molecule found")
        """
        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            services="srs",
            modes=mode,
        )

        with self.query_molecule_cache("srs", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                SRS_URL = "https://cdxappstest.epacdx.net/oms-substance-registry-services/rest-api/substance"
                search_response_text = self._resilient_request(
                    f'{SRS_URL}/{mode}/{urllib.parse.quote(identifier, safe="")}',
                    rejected_status_codes=[400, 404, 500],
                    kwargs={"timeout": 10},
                )

                if search_response_text is not None:
                    results = json.loads(search_response_text)
                    infos_by_identifier = self._match_SRS_results_to_identifiers(
                        [identifier], mode, results
                    )

                    if identifier in infos_by_identifier:
                        if len(infos_by_identifier[identifier]) != 1:
                            raise RuntimeError(
                                "More than one molecule found for the given identifier."
                            )
                        (
                            SMILES,
                            primary_names,
                            synonyms,
                            CAS,
                            ITN,
                            _,
                        ) = infos_by_identifier[identifier][0]
                        synonyms = self.filter_and_sort_synonyms(
                            primary_names + synonyms
                        )
                        if self.check_SMILES(
                            SMILES,
                            required_formula,
                            required_charge,
                            required_structure_type,
                        ):
                            molecules.append(
                                Molecule(
                                    SMILES,
                                    synonyms,
                                    CAS,
                                    ITN,
                                    mode,
                                    "srs",
                                    identifier=identifier,
                                )
                            )

        return self.filter_and_combine_molecules(
            molecules, required_formula, required_charge, required_structure_type
        )

    @cache
    def _get_info_from_CIR(
        self,
        structure_identifier: str,
        representation: str,
        resolvers_to_use: tuple[str],
        expected_number_of_results: Optional[int] = None,
    ) -> Optional[list[str]]:
        """
        Retrieve chemical information from the Chemical Identifier Resolver (CIR).

        This method queries the CIR API to obtain various representations of a chemical structure.

        Args:
            structure_identifier (str): The chemical identifier to search for.
            representation (str): The desired representation of the chemical (e.g., 'smiles', 'iupac_name').
            resolvers_to_use (tuple[str]): A tuple of resolver names to use in the query.
            expected_number_of_results (Optional[int]): The expected number of results.
            Defaults to None.

        Returns:
            Optional[list[str]]: A list of strings containing the requested chemical information,
            or None if the query fails or no results are found.

        Raises:
            RuntimeError: If the number of results doesn't match the expected number.
            requests.exceptions.ConnectionError: If there's a connection error during the API request.

        Notes:
            - This method is cached to improve performance for repeated queries.
            - It uses a resilient request mechanism to handle potential network issues.
            - If CIR is down, it sets a flag to avoid further attempts in the same session.
            - There's a 1-second delay between requests to avoid overwhelming the CIR server.
            - The method can handle connection reset errors by reinitializing the session.
            - API: https://cactus.nci.nih.gov/chemical/structure_documentation
            - API: https://search.r-project.org/CRAN/refm<ans/webchem/html/cir_query.html
            - API: https://github.com/mcs07/CIRpy

        Example:
            >>> resolver = MoleculeResolver()
            >>> smiles = resolver._get_info_from_CIR("ethanol", "smiles", ("name",))
            >>> print(smiles)
        """

        if "CIR_is_down" in self._message_slugs_shown:
            return None

        CIR_URL = "https://cactus.nci.nih.gov/chemical/structure/"
        # info can be e.g. smiles, iupac_name
        try:
            resolver_info = ""
            if resolvers_to_use:
                resolver_info = f'?resolver={",".join(resolvers_to_use)}'
            # although the documentation says otherwise it returns a 500 response even if it should return 404
            response_text = self._resilient_request(
                f"{CIR_URL}{urllib.parse.quote(structure_identifier)}/{representation}{resolver_info}",
                rejected_status_codes=[404, 500],
            )
            time.sleep(1)
            if not response_text:
                return None
            response_values = response_text.split("\n")
            if response_values:
                if expected_number_of_results is not None:
                    if len(response_values) != expected_number_of_results:
                        raise RuntimeError(
                            "More than one iupac_name was found. This was unexpected."
                        )
                return response_values
        except requests.exceptions.ConnectionError as e:
            # CIR closes the connection sometimes, so we reset the session and continue
            if "ConnectionResetError" in str(e):
                self._init_session()
                time.sleep(5)
                return self._get_info_from_CIR(
                    structure_identifier,
                    representation,
                    resolvers_to_use,
                    expected_number_of_results,
                )

            # I don't know why, but somtimes CIR is offline. This would make the module much slower as
            # it tries to connect multiple times anyway. Instead we give a warning and skip CIR.
            if "CIR_is_down" not in self._message_slugs_shown:
                self._message_slugs_shown.append("CIR_is_down")
                warnings.warn(
                    "CIR seems to be down, to continue working this instance of MoleculeResolver will skip CIR."
                )

        return None

    @cache
    def get_iupac_name_from_CIR(self, SMILES: str) -> Optional[str]:
        """
        Retrieve the IUPAC name for a given SMILES string using the Chemical Identifier Resolver (CIR).

        This method uses the CIR API to convert a SMILES representation of a molecule to its IUPAC name.

        Args:
            SMILES (str): The SMILES string representation of the molecule.

        Returns:
            Optional[str]: The IUPAC name of the molecule if found, or None if not found or in case of an error.

        Notes:
            - This method is cached to improve performance for repeated queries.
            - It internally uses the _get_info_from_CIR method to perform the API request.
            - The method specifically requests the 'iupac_name' representation and uses 'smiles' as the resolver.

        Example:
            >>> resolver = MoleculeResolver()
            >>> iupac_name = resolver.get_iupac_name_from_CIR("CCO")
            >>> print(iupac_name)
        """
        result = self._get_info_from_CIR(SMILES, "iupac_name", ("smiles",))
        return result[0] if result else None

    def get_molecule_from_CIR(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from Chemical Identifier Resolver (CIR) based on the provided identifier and mode.

        This method queries the CIR to retrieve molecule information using various types of chemical identifiers.

        Args:
            identifier (str): The chemical identifier to search for in CIR.
            mode (str): The type of identifier. Supported modes include 'formula', 'name', 'smiles', 'inchi', 'inchikey', and 'cas'.
            required_formula (Optional[str]): The expected molecular formula. Defaults to None.
            required_charge (Optional[int]): The expected molecular charge. Defaults to None.
            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and meets all requirements,
            None otherwise.

        Raises:
            ValueError: If the input parameters fail validation in the _check_parameters method.

        Notes:
            - The method first checks parameters using the _check_parameters method.
            - It then checks a cache for previously retrieved molecules.
            - The method uses different resolvers based on the input mode.
            - It retrieves SMILES representation from CIR using the _get_info_from_CIR method.
            - If a SMILES is found, it retrieves additional information like names and CAS numbers.
            - Synonyms and CAS numbers are filtered and sorted before being added to the Molecule object.
            - The resulting molecule is filtered based on the required formula, charge, and structure type if specified.

        Example:
            >>> resolver = MoleculeResolver()
            >>> molecule = resolver.get_molecule_from_CIR("ethanol", mode="name")
            >>> if molecule:
            ...     print(f"Found molecule: {molecule.get_SMILES()}")
            ... else:
            ...     print("No matching molecule found")
        """
        if required_formula is None:
            if mode == "formula":
                required_formula = identifier

        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            services="cir",
            modes=mode,
        )

        with self.query_molecule_cache("cir", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                resolvers_by_mode = {
                    "formula": None,
                    "name": ("name_by_cir",),
                    "smiles": ("smiles",),
                    "inchi": ("stdinchi",),
                    "inchikey": ("stdinchikey",),
                    "cas": ("cas_number", "name_by_cir"),
                }

                SMILES = self._get_info_from_CIR(
                    identifier, "smiles", resolvers_by_mode[mode], 1
                )
                if not SMILES:
                    return None
                else:
                    SMILES = SMILES[0]

                if SMILES:
                    CIR_names = self._get_info_from_CIR(
                        identifier, "names", resolvers_by_mode[mode]
                    )
                    synonyms = self.filter_and_sort_synonyms(
                        CIR_names if CIR_names else []
                    )
                    CAS = self.filter_and_sort_CAS(CIR_names if CIR_names else [])
                    molecules.append(
                        Molecule(SMILES, synonyms, CAS, mode=mode, service="cir")
                    )

        return self.filter_and_combine_molecules(
            molecules, required_formula, required_charge, required_structure_type
        )

    def get_molecule_from_NIST(
        self,
        identifier: str,
        mode: str,
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """
        Retrieve a molecule from NIST (National Institute of Standards and Technology) based on the provided identifier and mode.

        This method queries the NIST Chemistry WebBook to retrieve molecule information using various types of chemical identifiers.

        Args:
            identifier (str): The chemical identifier to search for in NIST.
            mode (str): The type of identifier. Supported modes include 'formula', 'name', 'cas', 'inchi', and 'smiles'.
            required_formula (Optional[str]): The expected molecular formula. Defaults to None.
            required_charge (Optional[int]): The expected molecular charge. Defaults to None.
            required_structure_type (Optional[str]): The expected structure type. Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and meets all requirements,
            None otherwise.

        Raises:
            ValueError: If the input parameters fail validation in the _check_parameters method.
            RuntimeError: If the webpage format has changed and cannot be parsed.

        Notes:
            - The method first checks parameters using the _check_parameters method.
            - It then checks a cache for previously retrieved molecules.
            - For 'smiles' mode, it converts the SMILES to InChI before querying NIST.
            - The method uses web scraping techniques to extract information from the NIST Chemistry WebBook.
            - It handles different response formats, including search results and direct molecule pages.
            - The method extracts InChI, CAS numbers, names, and synonyms from the HTML content.
            - It converts InChI to SMILES using the InChI_to_SMILES method.
            - Synonyms and CAS numbers are filtered and sorted before being added to the Molecule object.
            - The resulting molecule is filtered based on the required formula, charge, and structure type if specified.
            - For search result pages, it recursively queries each result until a matching molecule is found.

        Example:
            >>> resolver = MoleculeResolver()
            >>> molecule = resolver.get_molecule_from_NIST("64-17-5", mode="cas")
            >>> if molecule:
            ...     print(f"Found molecule: {molecule.get_SMILES()}")
            ... else:
            ...     print("No matching molecule found")
        """
        if required_formula is None:
            if mode == "formula":
                required_formula = identifier

        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            services="nist",
            modes=mode,
        )

        with self.query_molecule_cache("nist", mode, identifier) as (
            entry_available,
            molecules,
        ):
            if not entry_available:
                nist_modes = {
                    "formula": "Formula",
                    "name": "Name",
                    "cas": "ID",
                    "inchi": "InChI",
                    "smiles": "InChI",
                }
                NIST_Webbook_ID_regex = (
                    r'<a\s+href\s*=\s*"/cgi/cbook\.cgi\?ID=(.\d+).*">'
                )

                mode_used = mode
                if mode == "smiles":
                    identifier = self.SMILES_to_InChI(
                        self.standardize_SMILES(identifier)
                    )

                response_text = self._resilient_request(
                    f'https://webbook.nist.gov/cgi/cbook.cgi?{urllib.parse.quote(nist_modes[mode])}={urllib.parse.quote(identifier, safe="")}'
                )

                def parse_molecule(temp_content):
                    items = self.parse_items_from_html(
                        temp_content,
                        None,
                        [
                            ("html", '<h1 id="Top">(.*?)</h1>', [1]),
                            ("html", NIST_Webbook_ID_regex, [0, 1]),
                            (
                                "text",
                                rf"InChI: ({self.InChI_regex_compiled.pattern[1:-1]})",
                                [0, 1],
                            ),
                            (
                                "text",
                                f"CAS Registry Number: {self.CAS_regex}",
                                [0, 1],
                            ),
                            (
                                "html",
                                r"<li>\s*<strong>\s*Other\s+names:\s*</strong>(.*?)<\/li>",
                                [0, 1],
                            ),
                        ],
                        [0, 1],
                    )
                    if len(items) == 0:
                        return
                    if len(items) != 1:
                        raise RuntimeError("The webpage probably changed the format.")
                    name, additional_informationm, inchi, CAS, synonyms = items[0]

                    if not inchi:
                        return

                    if not synonyms:
                        synonyms = ""

                    if CAS:
                        CAS = [CAS]
                    else:
                        CAS = []

                    synonyms = [
                        name.strip()
                        for name in html.unescape(synonyms).split(";")
                        if name.strip()
                    ]
                    synonyms.insert(0, name)
                    synonyms = self.filter_and_sort_synonyms(synonyms)
                    CAS = self.filter_and_sort_CAS(CAS)
                    SMILES = self.InChI_to_SMILES(inchi)

                    if self.check_SMILES(
                        SMILES,
                        required_formula,
                        required_charge,
                        required_structure_type,
                    ):
                        molecules.append(
                            Molecule(
                                SMILES,
                                synonyms,
                                CAS,
                                additional_informationm,
                                mode_used,
                                "nist",
                                1,
                                identifier,
                            )
                        )

                if response_text is not None:
                    relevant_response_text = self.normalize_html(response_text)
                    relevant_response_text = regex.findall(
                        '<main id="main">(.*)</main>', relevant_response_text
                    )
                    if len(relevant_response_text) != 1:
                        raise RuntimeError("The format of the page might have changed.")
                    relevant_response_text = relevant_response_text[0].strip()

                    if (
                        not regex.search(
                            "<h1>.*(Not Found|no)\s*</h1>", relevant_response_text
                        )
                        and not regex.search(
                            "<h1>.*No Matching Names Found\s*</h1>",
                            relevant_response_text,
                        )
                        and not regex.search(
                            "<h1>.*No Matching Species Found\s*</h1>",
                            relevant_response_text,
                        )
                        and "no matching entries" not in relevant_response_text
                    ):
                        if "Search Results" not in relevant_response_text:
                            parse_molecule(relevant_response_text)
                        else:
                            # check all search results using check_SMILES
                            ids = regex.findall(NIST_Webbook_ID_regex, response_text)

                            for id in ids:
                                temp_result = self.get_molecule_from_NIST(
                                    id,
                                    "cas",
                                    required_formula,
                                    required_charge,
                                    required_structure_type,
                                )
                                if temp_result is not None:
                                    temp_result.mode = mode_used
                                    molecules.append(temp_result)

        return self.filter_and_combine_molecules(
            molecules, required_formula, required_charge, required_structure_type
        )

    def find_salt_molecules(
        self,
        identifiers: list[str],
        modes: Optional[list[str]] = ["name"],
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
        services_to_use: Optional[list[str]] = None,
        search_iupac_name: Optional[bool] = False,
        interactive: Optional[bool] = False,
        minimum_number_of_cross_checks: Optional[int] = 1,
        ignore_exceptions: Optional[bool] = False,
    ) -> tuple[list, list[int]]:
        """Finds salt molecules based on the provided identifiers and criteria.

        This method searches for salt molecules using the specified identifiers and optional parameters.
        It checks the structure type, charge, and other attributes to ensure accurate identification
        of salt compounds. If the search is unsuccessful, it attempts to derive information from
        synonyms or provided SMILES strings.

        Args:
            identifiers (list[str]): A list of identifiers for the molecules to search for.

            modes (Optional[list[str]]): The modes of identification for the identifiers.
            Defaults to ['name'].

            required_formula (Optional[str]): A chemical formula that the molecules must match.
            Defaults to None, meaning no specific formula is required.

            required_charge (Optional[int]): The charge that the molecules must possess.
            Defaults to None, which sets the charge to 0.

            required_structure_type (Optional[str]): The required structure type
            (e.g., "salt"). Defaults to None, which defaults to "salt".

            services_to_use (Optional[list[str]]): Specific services to be used for
            retrieving data. Defaults to None, meaning all services are available.

            search_iupac_name (Optional[bool]): If True, attempts to search using the IUPAC name.
            Defaults to False.

            interactive (Optional[bool]): If True, allows for interactive user input if necessary.
            Defaults to False.

            minimum_number_of_cross_checks (Optional[int]): Minimum number of services
            to cross-check for validity. Defaults to 1.

            ignore_exceptions (Optional[bool]): If True, ignores exceptions that may occur during
            the search process. Defaults to False.

        Returns:
            tuple[list, list[int]]: A tuple containing:

                - A list of found molecules, where each molecule is represented as a tuple of
                  relevant information.
                - A list of stoichiometric coefficients corresponding to the found molecules.

        Raises:
            ValueError: If the provided SMILES does not represent a salt.
            NotImplementedError: If the functionality for salts with more than 2 ions is invoked.

        Example:
            >>> find_salt_molecules(["NaCl", "K2SO4"], modes=['name'], required_charge=0)
            ([(...molecule data...)], [...stoichiometric coefficients...])
        """
        (
            flattened_identifiers,
            flattened_modes,
            synonyms,
            CAS,
            given_SMILES,
        ) = self._check_and_flatten_identifiers_and_modes(identifiers, modes)
        CAS = list(CAS)

        if given_SMILES:
            if self.get_structure_type_from_SMILES(given_SMILES) != "salt":
                raise ValueError("The given SMILES is not a salt.")

        if required_charge is None:
            required_charge = 0

        if required_structure_type is None:
            required_structure_type = "salt"

        if minimum_number_of_cross_checks is None:
            minimum_number_of_cross_checks = 1

        salt_info = self.find_single_molecule_cross_checked(
            identifiers,
            modes=modes,
            required_formula=required_formula,
            required_charge=required_charge,
            required_structure_type=required_structure_type,
            services_to_use=services_to_use,
            search_iupac_name=search_iupac_name,
            minimum_number_of_cross_checks=minimum_number_of_cross_checks,
            ignore_exceptions=ignore_exceptions,
        )

        if salt_info[0] is None:
            salt_info = (
                given_SMILES,
                synonyms,
                CAS,
                "given identifiers: " + ", ".join(flattened_identifiers),
                ", ".join(flattened_modes),
                "manual",
                1,
            )
            if interactive:
                salt_info = self.find_single_molecule(
                    identifiers,
                    modes=modes,
                    required_formula=required_formula,
                    required_charge=required_charge,
                    required_structure_type=required_structure_type,
                    services_to_use=[],
                    search_iupac_name=search_iupac_name,
                    interactive=interactive,
                    ignore_exceptions=ignore_exceptions,
                )

        all_molecules = []
        stoichometric_coefficients = []
        if salt_info[0] is not None:
            all_molecules.append(salt_info)
            stoichometric_coefficients.append(-1)
            SMILES = salt_info[0]

            if self.is_valid_SMILES(SMILES):
                if SMILES.count(".") > 0:
                    ionic_SMILES_list = SMILES.split(".")
                    ionic_SMILES_list = [
                        self.standardize_SMILES(smi) for smi in ionic_SMILES_list
                    ]
                    SMILES = ".".join(ionic_SMILES_list)
                    ionic_SMILES_set = set(ionic_SMILES_list)
                    if not len(ionic_SMILES_set) > 1:
                        raise ValueError("SMILES does not represent a salt")

                charge_sum = 0
                for ionic_SMILES in ionic_SMILES_set:
                    stoichometric_coefficients.append(SMILES.count(ionic_SMILES))

                    ionic_mol = self.get_from_SMILES(ionic_SMILES)
                    ionic_charge = Chem.rdmolops.GetFormalCharge(ionic_mol)

                    ionic_info = self.find_single_molecule_cross_checked(
                        ionic_SMILES,
                        modes="SMILES",
                        required_charge=ionic_charge,
                        required_structure_type="ion",
                        services_to_use=services_to_use,
                        search_iupac_name=search_iupac_name,
                        minimum_number_of_cross_checks=minimum_number_of_cross_checks,
                        ignore_exceptions=ignore_exceptions,
                    )

                    if ionic_info[0] is None:
                        ionic_info = (
                            ionic_SMILES,
                            [],
                            [],
                            f"given identifiers: {ionic_SMILES}",
                            "smiles",
                            "manual",
                            1,
                        )
                        if interactive:
                            ionic_info = self.find_single_molecule(
                                ionic_SMILES,
                                modes="SMILES",
                                required_charge=ionic_charge,
                                required_structure_type="ion",
                                services_to_use=[],
                                search_iupac_name=search_iupac_name,
                                interactive=interactive,
                                ignore_exceptions=ignore_exceptions,
                            )

                    if ionic_info[0] is not None:
                        all_molecules.append(ionic_info)

                        charge_sum += stoichometric_coefficients[-1] * ionic_charge

                if charge_sum != 0:
                    all_molecules = [salt_info]

        # not all molecules have been found
        if len(all_molecules) < 3:
            cation_info = None
            anion_info = None
            all_molecules = []
            if salt_info[0] is not None:
                all_molecules.append(salt_info)
                stoichometric_coefficients = [-1]

            for synonym in synonyms:
                synonym_parts = synonym.split(" ")
                possible_cation_name = synonym_parts[0]

                cation_info = self.find_single_molecule_cross_checked(
                    possible_cation_name,
                    modes=["name"],
                    required_charge="positive",
                    required_structure_type="ion",
                    services_to_use=services_to_use,
                    search_iupac_name=search_iupac_name,
                    minimum_number_of_cross_checks=minimum_number_of_cross_checks,
                    ignore_exceptions=ignore_exceptions,
                )

                if cation_info[0] is None:
                    if interactive:
                        cation_info = self.find_single_molecule(
                            possible_cation_name,
                            modes=["name"],
                            required_charge="positive",
                            required_structure_type="ion",
                            services_to_use=[],
                            search_iupac_name=search_iupac_name,
                            interactive=interactive,
                            ignore_exceptions=ignore_exceptions,
                        )

                possible_anion_name = synonym_parts[1:]

                anion_info = self.find_single_molecule_cross_checked(
                    possible_anion_name,
                    modes=["name"],
                    required_charge="negative",
                    required_structure_type="ion",
                    services_to_use=services_to_use,
                    search_iupac_name=search_iupac_name,
                    minimum_number_of_cross_checks=minimum_number_of_cross_checks,
                    ignore_exceptions=ignore_exceptions,
                )

                if anion_info[0] is None:
                    if interactive:
                        anion_info = self.find_single_molecule(
                            possible_anion_name,
                            modes=["name"],
                            required_charge="negative",
                            required_structure_type="ion",
                            services_to_use=[],
                            search_iupac_name=search_iupac_name,
                            interactive=interactive,
                            ignore_exceptions=ignore_exceptions,
                        )

                if cation_info[0] is not None and anion_info[0] is not None:
                    all_molecules.append(cation_info)
                    all_molecules.append(anion_info)

                if len(all_molecules) > 2:
                    break

            if len(all_molecules) > 2:
                if len(all_molecules) != 3:
                    raise NotImplementedError(
                        "Functionality for salts with more than 2 ions has not been implemented."
                    )

                cation_charge = Chem.rdmolops.GetFormalCharge(
                    self.get_from_SMILES(cation_info[0])
                )
                anion_charge = Chem.rdmolops.GetFormalCharge(
                    self.get_from_SMILES(anion_info[0])
                )

                if cation_charge >= abs(anion_charge):
                    if cation_charge % abs(anion_charge) == 0:
                        stoichometric_coefficients.append(1)
                        stoichometric_coefficients.append(
                            abs(int(cation_charge / anion_charge))
                        )
                    else:
                        stoichometric_coefficients.append(abs(anion_charge))
                        stoichometric_coefficients.append(abs(cation_charge))
                else:
                    if abs(anion_charge) % cation_charge == 0:
                        stoichometric_coefficients.append(
                            abs(int(anion_charge / cation_charge))
                        )
                        stoichometric_coefficients.append(1)
                    else:
                        stoichometric_coefficients.append(abs(anion_charge))
                        stoichometric_coefficients.append(abs(cation_charge))

                if stoichometric_coefficients[0] != -1:
                    temp_smiles = stoichometric_coefficients[1] * [cation_info[0]]
                    temp_smiles.extend(stoichometric_coefficients[2] * [anion_info[0]])
                    salt_info = (
                        self.standardize_SMILES(".".join(temp_smiles)),
                        synonyms,
                        list(CAS),
                        "from consisting ions",
                        "smiles",
                        "automatic",
                        1,
                    )
                    all_molecules.insert(0, salt_info)
                    stoichometric_coefficients.insert(0, -1)

        if len(all_molecules) < 3:
            stoichometric_coefficients = []

        return all_molecules, stoichometric_coefficients

    def find_single_molecule(
        self,
        identifiers: list[str],
        modes: Optional[list[str]] = ["name"],
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
        services_to_use: Optional[list[str]] = None,
        search_iupac_name: Optional[bool] = False,
        interactive: Optional[bool] = False,
        ignore_exceptions: Optional[bool] = False,
    ) -> Optional[Molecule]:
        """Searches for a single molecule across multiple chemical databases and services.

        This method attempts to find a molecule based on the provided identifiers and modes,
        using various chemical databases and services. It returns the first matching molecule
        that satisfies all the specified requirements.

        Args:
            identifiers (list[str]): A list of identifiers for the molecule (e.g., names, formulas, CAS numbers).

            modes (Optional[list[str]]): A list of modes corresponding to each identifier. Defaults to ['name']

            required_formula (Optional[str]): The required molecular formula. Defaults to None.

            required_charge (Optional[int]): The required molecular charge. Defaults to None.

            required_structure_type (Optional[str]): The required structure type. Defaults to None.

            services_to_use (Optional[list[str]]): A list of services to use for the search.
            If None, all available services will be used. Defaults to None.

            search_iupac_name (Optional[bool]): Whether to search for IUPAC names. Defaults to False.

            interactive (Optional[bool]): Whether to run in interactive mode. Defaults to False.

            ignore_exceptions (Optional[bool]): Whether to ignore exceptions during the search. Defaults to False.

        Returns:
            Optional[Molecule]: A Molecule object if found, None otherwise.

        Raises:
            Various exceptions depending on the services used and error conditions encountered.

        Notes:
            This method searches through multiple chemical databases and services in a specific order.
            It stops and returns the first matching molecule that satisfies all the specified requirements.
            The search order and the exact behavior may depend on the available services and the provided parameters.
        """
        if services_to_use is None:
            services_to_use = self._available_services

        (
            flattened_identifiers,
            flattened_modes,
            synonyms,
            CAS,
            given_SMILES,
        ) = self._check_and_flatten_identifiers_and_modes(identifiers, modes)
        self._check_parameters(
            services=services_to_use,
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            context="find_single",
        )

        if required_formula is None and modes.count("formula") == 1:
            required_formula = identifiers[modes.index("formula")]

        SMILES = None
        additional_information = None
        mode_used = None
        identifier_used = None
        current_service = None
        try:
            for service in services_to_use:
                current_service = service
                if service == "cas_registry":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_CAS_registry(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS = set(
                                    cmp.CAS
                                )  # overwrite CAS with data from the CAS registry
                                additional_information = cmp.service
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "pubchem":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_pubchem(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = (
                                    f"{cmp.service} id: {cmp.additional_information}"
                                )
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "cir":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_CIR(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                additional_information = cmp.service
                                mode_used = mode
                                identifier_used = cmp.identifier
                                break
                elif service == "opsin":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_OPSIN(
                                identifier,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                additional_information = cmp.additional_information
                                mode_used = mode
                                identifier_used = cmp.identifier
                                break
                elif service == "chebi":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_ChEBI(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = (
                                    f"{cmp.service} id: {cmp.additional_information}"
                                )
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "srs":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_SRS(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = (
                                    f"{cmp.service} id: {cmp.additional_information}"
                                )
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "comptox":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_CompTox(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = (
                                    f"{cmp.service} id: {cmp.additional_information}"
                                )
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "chemeo":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_Chemeo(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = (
                                    f"{cmp.service} id: {cmp.additional_information}"
                                )
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "cts":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_CTS(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = "cts"
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break
                elif service == "nist":
                    for identifier, mode in zip(
                        flattened_identifiers, flattened_modes, strict=True
                    ):
                        if mode in self.supported_modes_by_services[service]:
                            cmp = self.get_molecule_from_NIST(
                                identifier,
                                mode,
                                required_formula,
                                required_charge,
                                required_structure_type,
                            )
                            if cmp is not None:
                                SMILES = cmp.SMILES
                                synonyms.extend(cmp.synonyms)
                                CAS.update(cmp.CAS)
                                additional_information = (
                                    f"{cmp.service} id: {cmp.additional_information}"
                                )
                                mode_used = cmp.mode
                                identifier_used = cmp.identifier
                                break

                if SMILES is not None:
                    break

            if SMILES is None:
                if given_SMILES is not None:
                    if self.check_SMILES(
                        given_SMILES,
                        required_formula,
                        required_charge,
                        required_structure_type,
                    ):
                        SMILES = given_SMILES
                        additional_information = "given SMILES"
                        mode_used = "smiles"
                        if synonyms is None or len(synonyms) == 0:
                            search_iupac_name = True

            if SMILES is None:
                if len(synonyms) > 0 and required_charge is not None:
                    # if searching for an ion
                    # search for salts in pubchem and extract the single ions, take the one found most often
                    if required_charge != "zero" and required_charge != 0:
                        for identifier, mode in zip(
                            flattened_identifiers, flattened_modes, strict=True
                        ):
                            molecules = (
                                self.get_molecule_for_ion_from_partial_pubchem_search(
                                    identifier, required_formula, required_charge
                                )
                            )
                            if molecules is not None:
                                if len(molecules) > 0:
                                    mode_used = mode
                                    identifier_used = identifier
                                    current_service = "pubchem"
                                    additional_information = (
                                        "get_SMILES_for_ion_from_partial_pubchem_search"
                                    )
                                    SMILES = molecules[0][
                                        0
                                    ].SMILES  # get most likely candidate
                                    break

            if interactive and SMILES is None:
                return self.find_single_molecule_interactively(
                    identifiers,
                    modes,
                    required_formula=required_formula,
                    required_charge=required_charge,
                )

            if SMILES is not None:
                # keep first synonym
                kept_synonyms = []
                if len(synonyms) > 0:
                    kept_synonyms = [synonyms.pop(0)]

                if search_iupac_name:
                    # add iupac name if available
                    iupac_name = self.get_iupac_name_from_CIR(SMILES)

                    if iupac_name is not None:
                        if not isinstance(iupac_name, list):
                            iupac_name = [iupac_name]
                        kept_synonyms.extend(iupac_name)

                # add rest of names
                if len(synonyms) > 0:
                    kept_synonyms.extend(synonyms)

                kept_synonyms = self.filter_and_sort_synonyms(kept_synonyms)

                if len(kept_synonyms) == 0:
                    if interactive:
                        manual_name = ""
                        print("Could not infere name for following SMILES: " + SMILES)
                        while len(manual_name) < 3:
                            manual_name = input("please input manually:")
                            manual_name = manual_name.strip()

                        kept_synonyms = [manual_name]

                synonyms = kept_synonyms
        except Exception:
            if ignore_exceptions:
                print()
                print("Error with searching for", identifiers, "on", current_service)
                print(traceback.format_exc())
                print()
                return None
            else:
                raise

        if len(synonyms) == 0 or SMILES is None:
            return None

        SMILES = self.standardize_SMILES(SMILES)

        return Molecule(
            SMILES,
            synonyms,
            list(CAS),
            additional_information,
            mode_used,
            current_service,
            1,
            identifier_used,
        )

    def find_single_molecule_interactively(
        self,
        identifiers: list[str],
        modes: Optional[list[str]] = ["name"],
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
    ) -> Optional[Molecule]:
        """Interactively searches for a single molecule based on user input.

        This method prompts the user to input various identifiers for a molecule and
        attempts to find a matching molecule that satisfies all specified requirements.
        It allows for iterative refinement of the search based on user feedback.

        Args:
            identifiers (list[str]): Initial list of identifiers for the molecule.

            modes (Optional[list[str]]): List of modes corresponding to each identifier.
            Defaults to ['name'].

            required_formula (Optional[str]): The required molecular formula.
            Defaults to None.

            required_charge (Optional[int]): The required molecular charge.
            Defaults to None.

            required_structure_type (Optional[str]): The required structure type.
            Defaults to None.

        Returns:
            Optional[Molecule]: A Molecule object if a matching molecule is found and
            confirmed by the user, None if the search is aborted or no matching
            molecule is found.

        Raises:
            Various exceptions may be raised during the molecule search process.

        Notes:
            - The method uses a combination of automated searches and user input.
            - It supports various input formats including PubChem CID, name, CAS number,
              SMILES, and InChI.
            - The user can view and confirm the molecular structure before accepting
              the result.
            - If no synonyms are found, the user is prompted to provide a name for
              the molecule.
            - The method allows the user to input or confirm the CAS number.

        """
        (
            flattened_identifiers,
            flattened_modes,
            synonyms,
            CAS,
            given_SMILES,
        ) = self._check_and_flatten_identifiers_and_modes(identifiers, modes)
        self._check_parameters(
            required_formulas=required_formula,
            required_charges=required_charge,
            required_structure_types=required_structure_type,
            context="find_single",
        )

        SMILES = None
        start_geometry_comment = None

        # print header
        print("------------------------------------")
        formula = ""
        if required_formula is not None:
            formula = " | formula: " + required_formula
        charge = ""
        if required_charge is not None:
            charge = " | charge: " + str(required_charge)
        print(f"Searching for molecule {formula} {charge}")
        print("with the following identifiers:")
        for identifier, mode in zip(
            flattened_identifiers, flattened_modes, strict=True
        ):
            print(mode + ":", identifier)
        print("")

        mode = ""
        return_value = ""
        mode_used = "interactive"
        prompt_session = PromptSession()
        # regex for inchi, inchikey, SMILES, etc.
        # https://gist.github.com/lsauer/1312860/264ae813c2bd2c27a769d261c8c6b38da34e22fb
        while SMILES is None:
            if given_SMILES is not None:
                return_value = given_SMILES
            else:
                return_value = prompt_session.prompt(
                    "Please input the pubchem cid, name, CAS, SMILES, InChI or none to skip:",
                    default=str(return_value),
                )

            if return_value.lower() == "none":
                return None

            identifier = return_value
            try:
                return_value = int(return_value)
                mode = "cid"
                identifier = str(return_value)
            except Exception:
                if self.is_valid_SMILES(return_value):
                    mode = "smiles"
                elif self.is_valid_CAS(return_value):
                    mode = "cas"
                elif self.is_valid_InChI(return_value):
                    mode = "inchi"
                else:
                    mode = "name"

            try:
                if mode == "smiles":
                    (
                        temptative_SMILES,
                        temptative_synonyms,
                        temptative_CAS,
                        temptative_start_geometry_comment,
                        mode_used,
                    ) = (identifier, synonyms, CAS, "manual geometry", "interactive")
                else:
                    cmp = self.find_single_molecule(
                        [identifier], [mode], required_formula, required_charge
                    )
                    (
                        temptative_SMILES,
                        temptative_synonyms,
                        temptative_CAS,
                        temptative_start_geometry_comment,
                        mode_used,
                    ) = (
                        (
                            cmp.SMILES,
                            cmp.synonyms,
                            cmp.CAS,
                            cmp.additional_information,
                            cmp.mode,
                        )
                        if cmp is not None
                        else (None, [], [], "", "")
                    )
            except Exception:
                if mode == "cid":
                    print(
                        "There was a problem retreiving the molecule from pubchem. Does it exist?"
                    )

            if self.check_SMILES(
                temptative_SMILES,
                required_formula,
                required_charge,
                required_structure_type,
            ):
                temptative_synonyms = self.filter_and_sort_synonyms(
                    synonyms + temptative_synonyms
                )
                if len(temptative_synonyms) == 0:
                    possible_name = ""
                    while True:
                        possible_name = prompt_session.prompt(
                            "Input one name for the molecule", default=possible_name
                        )
                        if yes_no_dialog("Is this name correct?").run():
                            break
                    temptative_synonyms = [possible_name]

                temp_mol = self.get_from_SMILES(temptative_SMILES)

                if temp_mol is not None:
                    self.show_mol_and_pause(temp_mol, temptative_synonyms[0])
                else:
                    print(
                        "Error: Image could not be gernerated from SMILES, this however does not mean always that the SMILES ist wrong."
                    )

                if yes_no_dialog(
                    title="Confirmation", text="Is the molecule structure correct?"
                ).run():
                    SMILES = temptative_SMILES
                    synonyms = temptative_synonyms
                    CAS.update(temptative_CAS)
                    temptative_CAS = CAS
                    start_geometry_comment = temptative_start_geometry_comment

                    while True:
                        temptative_CAS = prompt_session.prompt(
                            "Input valid CAS or empty:",
                            default=",".join(temptative_CAS),
                        )

                        if len(temptative_CAS) == 0 or all(
                            [
                                self.is_valid_CAS(x.strip())
                                for x in temptative_CAS.split(",")
                            ]
                        ):
                            CAS = temptative_CAS.split(",")
                            break
        SMILES = self.standardize_SMILES(SMILES)

        return Molecule(
            SMILES,
            synonyms,
            list(CAS),
            start_geometry_comment,
            mode_used,
            "interactive",
            1,
        )

    def find_single_molecule_cross_checked(
        self,
        identifiers: list[str],
        modes: Optional[list[str]] = ["name"],
        required_formula: Optional[str] = None,
        required_charge: Optional[int] = None,
        required_structure_type: Optional[str] = None,
        services_to_use: Optional[list[str]] = None,
        search_iupac_name: Optional[bool] = False,
        minimum_number_of_cross_checks: Optional[int] = 1,
        try_to_choose_best_structure: Optional[bool] = True,
        ignore_exceptions: Optional[bool] = False,
    ) -> Union[Optional[Molecule], list[Optional[Molecule]]]:
        """Finds a single molecule with cross-checking across multiple services.

        This method searches for a molecule using the provided identifiers and modes,
        cross-checking the results across multiple chemical services to ensure accuracy.

        Args:
            identifiers (list[str]): List of identifiers for the molecule.
            modes (Optional[list[str]]): List of search modes. Defaults to ['name'].
            required_formula (Optional[str]): Required molecular formula. Defaults to None.
            required_charge (Optional[int]): Required molecular charge. Defaults to None.
            required_structure_type (Optional[str]): Required structure type. Defaults to None.
            services_to_use (Optional[list[str]]): List of services to use. If None, all available services are used.
            search_iupac_name (Optional[bool]): Whether to search for IUPAC names. Defaults to False.
            minimum_number_of_cross_checks (Optional[int]): Minimum number of services that must agree. Defaults to 1.
            try_to_choose_best_structure (Optional[bool]): Whether to attempt to select the best structure. Defaults to True.
            ignore_exceptions (Optional[bool]): Whether to ignore exceptions during search. Defaults to False.

        Returns:
            Union[Optional[Molecule], list[Optional[Molecule]]]: A single Molecule object if a best structure is chosen,
            a list of Molecule objects if multiple structures are found and not choosing the best,
            or None if no matching molecule is found.

        Raises:
            ValueError: If minimum_number_of_cross_checks exceeds the number of services used.

        Notes:
            - The method searches across multiple services and cross-checks the results.
            - It filters and groups molecules based on structure similarity.
            - If try_to_choose_best_structure is True, it attempts to select the most reliable structure.
            - The method uses various heuristics to resolve conflicts between different services.
            - OPSIN is given preference for name-based searches when available.
            - ChEBI results are given lower priority in case of conflicts.
        """
        if services_to_use is None:
            services_to_use = self._available_services

        if minimum_number_of_cross_checks is None:
            minimum_number_of_cross_checks = 1
        if not minimum_number_of_cross_checks <= len(services_to_use):
            raise ValueError(
                "The minimum_number_of_cross_checks exceeds the number of services that are used."
            )

        molecules = []

        for service in services_to_use:
            molecule = self.find_single_molecule(
                identifiers=identifiers,
                modes=modes,
                required_formula=required_formula,
                required_charge=required_charge,
                required_structure_type=required_structure_type,
                services_to_use=[service],
                search_iupac_name=search_iupac_name,
                ignore_exceptions=ignore_exceptions,
            )

            molecules.append(molecule)

        filtered_molecules = self.filter_molecules(
            molecules, required_formula, required_charge, required_structure_type
        )

        if not filtered_molecules:
            return None

        grouped_molecules = self.group_molecules_by_structure(filtered_molecules, False)

        maximum_number_of_crosschecks_found = max(
            [len(v) for v in grouped_molecules.values()]
        )

        if maximum_number_of_crosschecks_found < minimum_number_of_cross_checks:
            return None

        SMILES_with_highest_number_of_crosschecks = []
        for group_SMILES, group_molecules in grouped_molecules.items():
            if len(group_molecules) >= minimum_number_of_cross_checks:
                if len(group_molecules) == maximum_number_of_crosschecks_found:
                    SMILES_with_highest_number_of_crosschecks.append(group_SMILES)

        if try_to_choose_best_structure:
            SMILES_preferred = sorted(SMILES_with_highest_number_of_crosschecks)[0]
            if len(SMILES_with_highest_number_of_crosschecks) > 1:
                # if SMILES are the same ignoring isomeric info, use the more specific one:
                unique_non_isomeric_SMILES = set(
                    [
                        self.to_SMILES(self.get_from_SMILES(smi), isomeric=False)
                        for smi in SMILES_with_highest_number_of_crosschecks
                    ]
                )
                if len(unique_non_isomeric_SMILES) == 1:
                    SMILES_preferred = sorted(
                        SMILES_with_highest_number_of_crosschecks, key=len
                    )[-1]
                else:
                    # trust opsin algorithm: if not sure and opsin available
                    SMILES_preferred_by_opsin = None
                    for SMILES in SMILES_with_highest_number_of_crosschecks:
                        for molecule in grouped_molecules[SMILES]:
                            if molecule.mode == "name" and molecule.service == "opsin":
                                SMILES_preferred_by_opsin = SMILES

                    # if opsin result not available, or searched by another mode
                    # try getting all structures from the names and see if they agree
                    # with the SMILES found
                    if not SMILES_preferred_by_opsin:
                        SMILES_map = []
                        names_map = []
                        for SMILES in SMILES_with_highest_number_of_crosschecks:
                            for molecule in grouped_molecules[SMILES]:
                                for name in molecule.synonyms:
                                    SMILES_map.append(SMILES)
                                    names_map.append(name)

                        use_opsin_batch = False  # self._OPSIN_tempfolder is not None
                        if use_opsin_batch:
                            opsin_results = self.get_molecule_from_OPSIN_batchmode(
                                names_map
                            )
                        else:
                            opsin_results = [
                                self.get_molecule_from_OPSIN(name) for name in names_map
                            ]

                        SMILES_preferred_by_opsin = []
                        for (
                            original_SMILES_found,
                            molecule_found_by_opsin_from_synonym,
                        ) in zip(SMILES_map, opsin_results, strict=True):
                            if molecule_found_by_opsin_from_synonym:
                                if (
                                    original_SMILES_found
                                    == molecule_found_by_opsin_from_synonym.SMILES
                                ):
                                    SMILES_preferred_by_opsin.append(
                                        original_SMILES_found
                                    )

                        SMILES_preferred_by_opsin = set(SMILES_preferred_by_opsin)
                        if len(SMILES_preferred_by_opsin) == 1:
                            SMILES_preferred_by_opsin = SMILES_preferred_by_opsin.pop()
                        else:
                            SMILES_preferred_by_opsin = None

                    if SMILES_preferred_by_opsin:
                        SMILES_preferred = SMILES_preferred_by_opsin
                    else:
                        # usually when chebi does not agree with others chebi is wrong
                        # this is used to our advantage here
                        SMILES_not_from_chebi = []
                        for (
                            temptative_SMILES,
                            temptative_molecules,
                        ) in grouped_molecules.items():
                            molecules_from_chebi = [
                                t_mol
                                for t_mol in temptative_molecules
                                if "chebi" in t_mol.service
                            ]
                            if len(molecules_from_chebi) == 0:
                                SMILES_not_from_chebi.extend(
                                    [temptative_SMILES] * len(temptative_molecules)
                                )

                        c = collections.Counter(SMILES_not_from_chebi).most_common()
                        if len(c) == 1 or (len(c) > 1 and c[0][1] > c[1][1]):
                            SMILES_preferred = c[0][0]
                        else:
                            if self._show_warning_if_non_unique_structure_was_found:
                                temp = len(SMILES_with_highest_number_of_crosschecks)
                                warnings.warn(
                                    f"\n\n{temp} molecules were found equally as often. First one sorted by SMILES was taken: \n{grouped_molecules}\n"
                                )
            molec = self.combine_molecules(
                SMILES_preferred, grouped_molecules[SMILES_preferred]
            )
            molec.found_molecules.append(grouped_molecules)
            return molec
        else:
            return [
                self.combine_molecules(SMILES, grouped_molecules[SMILES])
                for SMILES in SMILES_with_highest_number_of_crosschecks
            ]

    def find_multiple_molecules_parallelized(
        self,
        identifiers: list[str],
        modes: Optional[list[str]] = ["name"],
        required_formulas: Optional[list[str]] = None,
        required_charges: Optional[list[int]] = None,
        required_structure_types: Optional[list[str]] = None,
        services_to_use: Optional[list[str]] = None,
        search_iupac_name: Optional[bool] = False,
        minimum_number_of_cross_checks: Optional[int] = 1,
        try_to_choose_best_structure: Optional[bool] = True,
        progressbar: Optional[bool] = True,
        max_workers: Optional[int] = 5,
        ignore_exceptions: bool = True,
    ) -> list[Optional[Molecule]]:
        """Finds multiple molecules in parallel based on provided identifiers and criteria.

        This method utilizes multithreading to search for multiple molecules concurrently. It
        checks various attributes such as required formulas, charges, and structure types,
        and fetches data from specified services, with options for progress tracking and
        batch processing.

        Args:
            identifiers (list[str]): A list of identifiers for the molecules to search for.

            modes (Optional[list[str]]): The modes of identification for the identifiers. Defaults to ['name']

            required_formulas (Optional[list[str]]): A list of chemical formulas
            hat the molecules must match. Defaults to None, which means no specific formula
            is required for each identifier.

            required_charges (Optional[list[int]]): A list of charges that the
            molecules must possess. Defaults to None, indicating no specific charge is required.

            required_structure_types (Optional[list[str]]): A list of required
            structure types (e.g., "salt"). Defaults to None, meaning no specific structure
            type is required.

            services_to_use (Optional[list[str]]): Specific services to be used for
            retrieving data. Defaults to None, meaning all available services are considered.

            search_iupac_name (Optional[bool]): If True, attempts to search using the IUPAC name.
            Defaults to False.

            minimum_number_of_cross_checks (Optional[int]): Minimum number of services
            to cross-check for validity. Defaults to 1.

            try_to_choose_best_structure (Optional[bool]): If True, attempts to select the best
            structure among the results. Defaults to True.

            progressbar (Optional[bool]): If True, displays a progress bar during the search.
            Defaults to True.

            max_workers (Optional[int]): Maximum number of threads to use for parallel processing.
            Defaults to 5.

            ignore_exceptions (Optional[bool]): If True, ignores exceptions that may occur during
            the search process. Defaults to True.

        Returns:
            list[Optional[Molecule]]: A list of found molecules, where each molecule is represented
            as an instance of the Molecule class, or None if not found.

        Raises:
            ValueError: If any of the identifiers or parameters are invalid.

        Example:
            >>> find_multiple_molecules_parallelized(
            ...     identifiers=["NaCl", "K2SO4"],
            ...     modes=['name'],
            ...     required_charges=[0, -2]
            ... )
            [<Molecule instance for NaCl>, <Molecule instance for K2SO4>]
        """
        # reinitialize session
        self._session = None
        self._init_session(pool_maxsize=max_workers * 2)

        if services_to_use is None:
            services_to_use = self._available_services

        services_to_use = [service.lower() for service in services_to_use]
        if required_formulas is None:
            required_formulas = [None] * len(identifiers)

        if required_structure_types is not None:
            if isinstance(required_structure_types, str):
                required_structure_types = [required_structure_types] * len(identifiers)
        else:
            required_structure_types = [None] * len(identifiers)
        if required_charges is not None:
            if isinstance(required_charges, int):
                required_charges = [required_charges] * len(identifiers)
        else:
            required_charges = [None] * len(identifiers)
        self._check_parameters(
            modes=modes,
            services=services_to_use,
            required_charges=required_charges,
            required_formulas=required_formulas,
            required_structure_types=required_structure_types,
            context="find_multiple",
        )

        # when lots of molecules, use batch capabilities if possible
        message_getting_from_services = "Getting data from services"
        if len(identifiers) > 100:
            message_getting_from_services = "Getting data from all other services"

            services_asked_for_with_batch_capabilities = set(
                services_to_use
            ).intersection(set(self._available_services_with_batch_capabilities))
            services_asked_for_with_batch_capabilities = sorted(
                list(services_asked_for_with_batch_capabilities)
            )

            # sometimes there are issues with opsin, so always try to get from opsin first
            # in the case of an error you don't loose the rest
            if "opsin" in services_asked_for_with_batch_capabilities:
                services_asked_for_with_batch_capabilities.remove("opsin")
                services_asked_for_with_batch_capabilities.insert(0, "opsin")

            for service in services_asked_for_with_batch_capabilities:
                print("Getting data in batchmode for the following service:", service)

                if not self._is_list_of_list_of_str(
                    identifiers
                ):  # Convert list[str] to list[list[str]] for usage in batchmode
                    identifiers = [[idf] for idf in identifiers]
                if not self._is_list_of_list_of_str(modes):
                    modes = [[md] for md in modes]

                self.get_molecules_using_batchmode_from(
                    identifiers,
                    modes,
                    service,
                    progressbar=progressbar,
                )

        def _find(generator):
            return list(
                tqdm(generator, total=len(identifiers), disable=not progressbar)
            )

        args = []
        for (
            identifier,
            mode,
            required_formula,
            required_charge,
            required_structure_type,
        ) in zip(
            identifiers,
            modes,
            required_formulas,
            required_charges,
            required_structure_types,
            strict=True,
        ):
            if minimum_number_of_cross_checks is None:
                args.append(
                    (
                        identifier,
                        mode,
                        required_formula,
                        required_charge,
                        required_structure_type,
                        services_to_use,
                        search_iupac_name,
                        False,
                        ignore_exceptions,
                    )
                )
            else:
                args.append(
                    (
                        identifier,
                        mode,
                        required_formula,
                        required_charge,
                        required_structure_type,
                        services_to_use,
                        search_iupac_name,
                        minimum_number_of_cross_checks,
                        try_to_choose_best_structure,
                        ignore_exceptions,
                    )
                )

        print(message_getting_from_services)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            if not minimum_number_of_cross_checks:
                results = _find(executor.map(self.find_single_molecule, *zip(*args)))
            else:
                results = _find(
                    executor.map(self.find_single_molecule_cross_checked, *zip(*args))
                )

        return results
